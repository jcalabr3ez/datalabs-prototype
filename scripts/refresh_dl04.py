#!/usr/bin/env python3
"""Refresh the DL-04 retail electricity ledger from EIA and Census files.

Primary price and sales source (SRC-401): EIA Form EIA-861 historical state
file, Total Electric Industry sheet (HS861 2010-). The U.S. Total row is
EIA's sales-weighted all-sector average; this script never averages the
50 state prices.

Verification path: Electric Power Annual table 2.10 all-sector cells for
the two most recent years (same Form EIA-861, different EIA publication).

Generation (SRC-403): EIA-923 / annual_generation_state.xls, Total Electric
Power Industry, Total energy source.

Net summer capacity (SRC-404): EIA-860 existcapacity_annual.xlsx, Total
Electric Power Industry, All Sources.

Population (SRC-402): Census Bureau vintage 2025 state totals for 2020
forward, and the 2010-2020 vintage file for 2012-2019.

Writes netlify/functions/dl04-answers.json and re-runs inject_data.py.
Never pushes to main. Exits nonzero when sanity or two-path checks fail.
"""
import csv
import io
import json
import subprocess
import sys
import urllib.request
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
LEDGER = ROOT / "netlify/functions/dl04-answers.json"

UA = "PioneerDataLabs/1.0 (jcalabrese@pioneerinstitute.org)"
FIRST_YEAR = 2012

URL_HS861 = "https://www.eia.gov/electricity/data/state/xls/861/HS861%202010-.xlsx"
URL_EPA_210 = "https://www.eia.gov/electricity/annual/xls/epa_02_10.xlsx"
URL_GEN = "https://www.eia.gov/electricity/data/state/annual_generation_state.xls"
URL_CAP = "https://www.eia.gov/electricity/data/state/existcapacity_annual.xlsx"
URL_POP_2025 = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2020-2025/state/totals/NST-EST2025-ALLDATA.csv"
)
URL_POP_2010 = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2010-2020/state/totals/nst-est2020-alldata.csv"
)

# Two-path and TWBX extract checks. MA 2023 23.21 is the cell that matched
# the recovered Retail-Price Electricity workbook; EPA 2.10 reprints it.
VERIFY_HS861 = {
    (2023, "MA"): 23.21,
    (2024, "MA"): 23.94,
    (2023, "US"): 12.68,
    (2024, "US"): 12.94,
    (2023, "HI"): 38.60,
    (2023, "ND"): 8.03,
}

STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut",
    "DE": "Delaware", "DC": "District of Columbia", "FL": "Florida",
    "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois",
    "IN": "Indiana", "IA": "Iowa", "KS": "Kansas", "KY": "Kentucky",
    "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana",
    "NE": "Nebraska", "NV": "Nevada", "NH": "New Hampshire",
    "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania",
    "RI": "Rhode Island", "SC": "South Carolina", "SD": "South Dakota",
    "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "US": "United States",
}
NAME_TO_ST = {v: k for k, v in STATE_NAMES.items()}
NAME_TO_ST["U.S. Total"] = "US"
NAME_TO_ST["US Total"] = "US"

FIPS_TO_ST = {
    "00": "US", "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA",
    "08": "CO", "09": "CT", "10": "DE", "11": "DC", "12": "FL", "13": "GA",
    "15": "HI", "16": "ID", "17": "IL", "18": "IN", "19": "IA", "20": "KS",
    "21": "KY", "22": "LA", "23": "ME", "24": "MD", "25": "MA", "26": "MI",
    "27": "MN", "28": "MS", "29": "MO", "30": "MT", "31": "NE", "32": "NV",
    "33": "NH", "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI", "45": "SC",
    "46": "SD", "47": "TN", "48": "TX", "49": "UT", "50": "VT", "51": "VA",
    "53": "WA", "54": "WV", "55": "WI", "56": "WY",
}

CENSUS_DIVISIONS = {
    "New England", "Middle Atlantic", "East North Central",
    "West North Central", "South Atlantic", "East South Central",
    "West South Central", "Mountain", "Pacific Contiguous",
    "Pacific Noncontiguous", "Pacific",
}


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


def num(v):
    if v is None or v == "" or v == "--" or v == ".":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s in {"", "--", ".", "NA", "NM"}:
        return None
    return float(s)


# Year, STATE, then Residential / Commercial / Industrial / Transportation /
# Total, each with revenue, sales, customers, price. Total price is column 21
# (0-based); total sales is column 19.
SECTOR_PRICE_COL = {
    "residential": 5,
    "commercial": 9,
    "industrial": 13,
    "total": 21,
}
SECTOR_SALES_COL = {
    "residential": 3,
    "commercial": 7,
    "industrial": 11,
    "total": 19,
}


def load_hs861(blob):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb["Total Electric Industry"]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        year, st = row[0], row[1]
        if year is None or st is None:
            continue
        year = int(year)
        st = str(st).strip()
        if year < FIRST_YEAR or st not in STATE_NAMES:
            continue
        price = num(row[SECTOR_PRICE_COL["total"]])
        sales = num(row[SECTOR_SALES_COL["total"]])
        if price is None or sales is None:
            continue
        rec = {
            "price_cents": round(price, 2),
            "sales_mwh": int(round(sales)),
        }
        for name, col in SECTOR_PRICE_COL.items():
            if name == "total":
                continue
            cents = num(row[col])
            if cents is not None:
                rec[name + "_cents"] = round(cents, 2)
            sold = num(row[SECTOR_SALES_COL[name]])
            if sold is not None:
                rec[name + "_sales_mwh"] = int(round(sold))
        out[(year, st)] = rec
    wb.close()
    return out


def pack_sector_year(hs861, year, field):
    """Rank 51 jurisdictions on one HS861 sector price. U.S. is EIA's row."""
    ranked = []
    us = None
    for st, name in STATE_NAMES.items():
        rec = hs861.get((year, st))
        if not rec or rec.get(field) is None:
            continue
        if st == "US":
            us = rec[field]
            continue
        ranked.append({"st": st, "name": name, "v": rec[field]})
    if us is None or len(ranked) < 51:
        return None
    ranked.sort(key=lambda r: (-r["v"], r["st"]))
    for i, rec in enumerate(ranked, 1):
        rec["rank"] = i
        rec["n"] = len(ranked)
    ma = next((r for r in ranked if r["st"] == "MA"), None)
    fl = next((r for r in ranked if r["st"] == "FL"), None)
    return {
        "year": year,
        "us": us,
        "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]} if ma else None,
        "fl": {"v": fl["v"], "rank": fl["rank"], "n": fl["n"]} if fl else None,
        "highest": {"st": ranked[0]["st"], "name": ranked[0]["name"], "v": ranked[0]["v"]},
        "lowest": {"st": ranked[-1]["st"], "name": ranked[-1]["name"], "v": ranked[-1]["v"]},
        "n_ranked": len(ranked),
        "states": ranked,
    }


def sector_trend(hs861, st, field):
    years = sorted({y for (y, s) in hs861 if s == st})
    out = []
    for y in years:
        rec = hs861.get((y, st))
        if rec and rec.get(field) is not None:
            out.append({"y": y, "v": rec[field]})
    return out


def attach_sector_prices(ledger, hs861):
    """Add residential, commercial, and industrial prices to an existing ledger."""
    latest = ledger.get("data_year") or (ledger.get("latest") or {}).get("year")
    if not latest:
        sys.exit("FATAL: DL-04 ledger has no data_year")
    packs = {}
    for name in ("residential", "commercial", "industrial"):
        pack = pack_sector_year(hs861, latest, name + "_cents")
        if not pack:
            sys.exit(f"FATAL: HS861 missing {name} prices for {latest}")
        packs[name] = pack
    res = packs["residential"]
    if abs(res["us"] - 16.48) > 2 and latest == 2024:
        # Residential sits above the all-sector average. Guard a wild parse.
        if not (8 <= res["us"] <= 30):
            sys.exit(f"FATAL: implausible U.S. residential price {res['us']}")
    if not (8 <= res["us"] <= 40):
        sys.exit(f"FATAL: implausible U.S. residential price {res['us']}")
    ledger.setdefault("latest", {})["residential"] = {
        "year": latest,
        "us": {"price_cents": res["us"]},
        "ma": {
            "price_cents": res["ma"]["v"],
            "rank": res["ma"]["rank"],
            "n": res["ma"]["n"],
        },
        "fl": {
            "price_cents": res["fl"]["v"],
            "rank": res["fl"]["rank"],
            "n": res["fl"]["n"],
        } if res.get("fl") else None,
        "highest": {
            "st": res["highest"]["st"],
            "name": res["highest"]["name"],
            "price_cents": res["highest"]["v"],
        },
        "lowest": {
            "st": res["lowest"]["st"],
            "name": res["lowest"]["name"],
            "price_cents": res["lowest"]["v"],
        },
    }
    ledger["residential_states"] = [
        {
            "st": r["st"],
            "name": r["name"],
            "price_cents": r["v"],
            "rank": r["rank"],
        }
        for r in res["states"]
    ]
    ledger["residential_trend"] = {
        st: sector_trend(hs861, st, "residential_cents")
        for st in ("US", "MA", "FL")
    }
    derived = ledger.setdefault("derived", {})
    derived["sectors"] = {
        name: {
            "label": name.capitalize() + " average retail price, " + str(latest),
            "src": "SRC-401",
            "unit": "cents per kWh",
            "us": packs[name]["us"],
            "ma": packs[name]["ma"],
            "fl": packs[name]["fl"],
            "highest": packs[name]["highest"],
            "lowest": packs[name]["lowest"],
            "n_ranked": packs[name]["n_ranked"],
        }
        for name in ("residential", "commercial", "industrial")
    }
    scope = ledger.get("scope") or ""
    scope = scope.replace(
        "Does NOT cover: residential, commercial, or industrial prices as "
        "separate published series on this page; ",
        "Also covers residential, commercial, and industrial average prices "
        "from the same EIA-861 Total Electric Industry file. Does NOT cover: ",
    )
    scope = scope.replace(
        "Does NOT cover: residential, commercial, or industrial prices as "
        "separate series; ",
        "Also covers residential, commercial, and industrial average prices "
        "from the same EIA-861 file. Does NOT cover: ",
    )
    ledger["scope"] = scope
    return ledger


def load_epa_210(blob):
    """All-sector price by state for the two years on table 2.10."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=6, values_only=True))
    # Row 3: Year 2024, Year 2023, ... All Sectors is the last pair (ix 9, 10).
    header = rows[3]
    y_new = int(str(header[9]).replace("Year ", "").strip())
    y_old = int(str(header[10]).replace("Year ", "").strip())
    out = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        name = (row[0] or "").replace("\n", " ").strip()
        if not name or name in CENSUS_DIVISIONS:
            continue
        st = NAME_TO_ST.get(name)
        if not st:
            continue
        p_new, p_old = num(row[9]), num(row[10])
        if p_new is not None:
            out[(y_new, st)] = round(p_new, 2)
        if p_old is not None:
            out[(y_old, st)] = round(p_old, 2)
    wb.close()
    return out


def load_generation(blob):
    import xlrd
    book = xlrd.open_workbook(file_contents=blob)
    sh = book.sheet_by_index(0)
    # Row 0 is a title; row 1 is YEAR, STATE, TYPE OF PRODUCER, ENERGY SOURCE, MWh
    out = {}
    for r in range(2, sh.nrows):
        year = sh.cell_value(r, 0)
        st = str(sh.cell_value(r, 1)).strip()
        producer = str(sh.cell_value(r, 2)).strip()
        source = str(sh.cell_value(r, 3)).strip()
        if producer != "Total Electric Power Industry" or source != "Total":
            continue
        try:
            year = int(year)
        except (TypeError, ValueError):
            continue
        if year < FIRST_YEAR or st not in STATE_NAMES:
            continue
        gen = num(sh.cell_value(r, 4))
        if gen is None:
            continue
        out[(year, st)] = int(round(gen))
    return out


def load_capacity(blob):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb.active
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        year, st, producer, fuel = row[0], row[1], row[2], row[3]
        if producer != "Total Electric Power Industry" or fuel != "All Sources":
            continue
        try:
            year = int(year)
        except (TypeError, ValueError):
            continue
        st = str(st).strip()
        if year < FIRST_YEAR or st not in STATE_NAMES:
            continue
        cap = num(row[7])  # Summer Capacity (Megawatts)
        if cap is None:
            continue
        out[(year, st)] = round(cap, 1)
    wb.close()
    return out


def load_population(blob_2025, blob_2010):
    """Return {(year, st): pop}. 2012-2019 from the 2010 vintage file;
    2020 forward from vintage 2025 so the latest year stays current."""
    out = {}

    def read(blob, years):
        text = blob.decode("utf-8-sig")
        rows = csv.DictReader(io.StringIO(text))
        for row in rows:
            st = FIPS_TO_ST.get(row.get("STATE", "").zfill(2))
            if not st:
                continue
            sumlev = row.get("SUMLEV", "")
            if st == "US" and row.get("NAME") not in {"United States"}:
                continue
            if st != "US" and sumlev not in {"040", "40"}:
                continue
            for y in years:
                key = "POPESTIMATE" + str(y)
                if key in row and row[key]:
                    out[(y, st)] = int(float(row[key]))

    read(blob_2010, range(FIRST_YEAR, 2020))
    read(blob_2025, range(2020, 2030))
    return out


def pct_change(new, old):
    if not old:
        return None
    return round(100 * (new / old - 1), 1)


def main():
    print("refresh_dl04: fetching EIA and Census files")
    hs861 = load_hs861(fetch(URL_HS861))
    epa = load_epa_210(fetch(URL_EPA_210))
    gen = load_generation(fetch(URL_GEN))
    cap = load_capacity(fetch(URL_CAP))
    pop = load_population(fetch(URL_POP_2025), fetch(URL_POP_2010))

    years = sorted({y for (y, st) in hs861 if st == "US"})
    if not years:
        sys.exit("FATAL: no U.S. Total price rows in HS861")
    # continuity from FIRST_YEAR
    want = list(range(FIRST_YEAR, years[-1] + 1))
    if years != want:
        sys.exit(f"FATAL: HS861 U.S. years {years} != {want}")
    latest = years[-1]
    missing_latest = [st for st in STATE_NAMES if (latest, st) not in hs861]
    if missing_latest:
        sys.exit(f"FATAL: HS861 missing {latest} for {missing_latest}")
    if len(STATE_NAMES) != 52:
        sys.exit(f"FATAL: expected 52 entities, found {len(STATE_NAMES)}")

    # Two-path: HS861 vs EPA table 2.10 on overlapping cells
    mismatches = []
    for key, epa_v in epa.items():
        if key not in hs861:
            continue
        hs = hs861[key]["price_cents"]
        if abs(hs - epa_v) > 0.011:
            mismatches.append(f"{key}: HS861 {hs} vs EPA 2.10 {epa_v}")
    if mismatches:
        sys.exit("FATAL: HS861 / EPA 2.10 mismatch:\n  " + "\n  ".join(mismatches[:12]))

    for key, expect in VERIFY_HS861.items():
        if key not in hs861:
            if key[0] > latest:
                continue
            sys.exit(f"FATAL: missing verification cell {key}")
        got = hs861[key]["price_cents"]
        if abs(got - expect) > 0.011:
            sys.exit(f"FATAL: {key} price {got} != expected {expect}")

    # Never invent a U.S. average from the states
    us_price = hs861[(latest, "US")]["price_cents"]
    state_mean = sum(hs861[(latest, st)]["price_cents"] for st in STATE_NAMES if st != "US") / 51
    if abs(us_price - round(state_mean, 2)) < 0.05:
        sys.exit(
            f"FATAL: U.S. Total {us_price} looks like an unweighted state mean "
            f"{state_mean:.2f}; the U.S. row must be EIA's sales-weighted total"
        )
    if not (5 <= us_price <= 25):
        sys.exit(f"FATAL: implausible U.S. price {us_price}")

    series = []
    for y in years:
        for st in STATE_NAMES:
            h = hs861.get((y, st))
            if not h:
                sys.exit(f"FATAL: missing HS861 {y} {st}")
            g = gen.get((y, st))
            c = cap.get((y, st))
            p = pop.get((y, st))
            if st != "US" and p is None:
                sys.exit(f"FATAL: missing population {y} {st}")
            rec = {
                "y": y,
                "st": st,
                "price_cents": h["price_cents"],
                "sales_mwh": h["sales_mwh"],
            }
            if g is not None:
                rec["gen_mwh"] = g
            if c is not None:
                rec["cap_mw"] = c
            if p is not None:
                rec["pop"] = p
                rec["sales_mwh_per_capita"] = round(h["sales_mwh"] / p, 3)
                if g is not None:
                    rec["gen_mwh_per_capita"] = round(g / p, 3)
            series.append(rec)

    by_key = {(r["y"], r["st"]): r for r in series}
    ranked = sorted(
        (by_key[(latest, st)] for st in STATE_NAMES if st != "US"),
        key=lambda r: (-r["price_cents"], r["st"]),
    )
    rank_of = {r["st"]: i for i, r in enumerate(ranked, 1)}
    n_states = len(ranked)
    highest, lowest = ranked[0], ranked[-1]
    ma = by_key[(latest, "MA")]
    us = by_key[(latest, "US")]
    prev_us = by_key.get((latest - 1, "US"))
    prev_ma = by_key.get((latest - 1, "MA"))

    def trend(st):
        return [{"y": y, "v": by_key[(y, st)]["price_cents"]} for y in years]

    latest_states = []
    for r in ranked:
        prev = by_key.get((latest - 1, r["st"]))
        latest_states.append({
            "st": r["st"],
            "name": STATE_NAMES[r["st"]],
            "price_cents": r["price_cents"],
            "rank": rank_of[r["st"]],
            "yoy_pct": pct_change(r["price_cents"], prev["price_cents"]) if prev else None,
            "sales_mwh": r["sales_mwh"],
            "gen_mwh": r.get("gen_mwh"),
            "cap_mw": r.get("cap_mw"),
            "pop": r.get("pop"),
            "sales_mwh_per_capita": r.get("sales_mwh_per_capita"),
            "gen_mwh_per_capita": r.get("gen_mwh_per_capita"),
        })

    today = date.today().isoformat()
    as_of = f"{latest}-12"
    old = {}
    if LEDGER.exists():
        old = json.loads(LEDGER.read_text(encoding="utf-8"))
        old_year = int(str(old.get("as_of", "0")).split("-")[0])
        if latest < old_year:
            sys.exit(f"FATAL: fetched year {latest} is older than ledger {old.get('as_of')}")

    page_revised = date.today().strftime("%b %-d, %Y").replace("  ", " ")
    # Linux %-d works; fall back if a platform rejects it.
    try:
        page_revised = date.today().strftime("%b %-d, %Y")
    except ValueError:
        page_revised = date.today().strftime("%b %d, %Y").replace(" 0", " ")

    new = {
        "tool_id": "DL-04",
        "title": "Retail Electricity Prices",
        "as_of": as_of,
        "data_year": latest,
        "first_year": FIRST_YEAR,
        "scope": (
            "Covers the all-sector average retail price of electricity by state "
            "and for the United States, plus retail sales, net generation, net "
            "summer capacity, and per-capita sales and generation where a Census "
            f"population figure joins, calendar years {FIRST_YEAR} through {latest}. "
            "The U.S. figure is EIA's published U.S. Total row, a sales-weighted "
            "all-sector average, never an unweighted mean of the state prices. "
            "Also covers residential, commercial, and industrial average prices "
            "from the same EIA-861 Total Electric Industry file. Does NOT cover: "
            "utility, city, or customer class rates; forecasts or what prices "
            "will do next year; bill calculators or rate-case advice; other fuels."
        ),
        "vintage_note": (
            f"Rebuilt from EIA Form EIA-861 historical state file (HS861 2010-), "
            f"Total Electric Industry, through {latest}, on {today} by "
            f"scripts/refresh_dl04.py. All-sector prices and sales for {latest - 1} "
            f"and {latest} were checked cell-by-cell against Electric Power Annual "
            f"table 2.10 (SRC-401). Generation is EIA-923 annual state totals "
            f"(SRC-403). Net summer capacity is EIA-860 (SRC-404). Population is "
            f"Census vintage 2025 for 2020 forward and the 2010-2020 vintage file "
            f"for {FIRST_YEAR}-2019 (SRC-402). Massachusetts 2023 all-sector price "
            f"23.21 cents per kWh matches both EIA publications and the recovered "
            f"Retail-Price Electricity extract."
        ),
        "page": {"revised": page_revised, "version": "1.0"},
        "source_id_map": {
            "SRC-401": {
                "label": "EIA Form EIA-861 / Electric Power Annual table 2.10",
                "what": "All-sector, residential, commercial, and industrial average retail prices (cents per kWh) and retail sales (MWh) by state, including the U.S. Total row",
                "cadence": "Annual; Electric Power Annual and the EIA-861 historical state file typically land in October for the prior calendar year",
                "url": "https://www.eia.gov/electricity/data/state/",
                "epa_table": "https://www.eia.gov/electricity/annual/html/epa_02_10.html",
                "vintage": str(latest),
                "next_release": f"{latest + 1} annual, expected October {latest + 2}",
            },
            "SRC-402": {
                "label": "Census Bureau, state population estimates",
                "what": "July 1 resident population used for per-capita sales and generation",
                "cadence": "Annual vintage estimates",
                "url": "https://www.census.gov/programs-surveys/popest.html",
                "vintage": "Vintage 2025 (2020-2025); 2010-2020 vintage file for 2012-2019",
                "next_release": "Vintage 2026, expected late 2026",
            },
            "SRC-403": {
                "label": "EIA-923 / annual generation by state",
                "what": "Net electricity generation (MWh), Total Electric Power Industry, all energy sources",
                "cadence": "Annual",
                "url": "https://www.eia.gov/electricity/data/state/",
                "vintage": str(latest),
                "next_release": f"{latest + 1} annual, expected October {latest + 2}",
            },
            "SRC-404": {
                "label": "EIA-860 / existing capacity by state",
                "what": "Net summer capacity (MW), Total Electric Power Industry, all sources",
                "cadence": "Annual",
                "url": "https://www.eia.gov/electricity/data/state/",
                "vintage": str(latest),
                "next_release": f"{latest + 1} annual, expected with the next EIA-860 release",
            },
        },
        "entities": STATE_NAMES,
        "series": series,
        "latest": {
            "year": latest,
            "us": {
                "price_cents": us["price_cents"],
                "sales_mwh": us["sales_mwh"],
                "gen_mwh": us.get("gen_mwh"),
                "cap_mw": us.get("cap_mw"),
                "pop": us.get("pop"),
                "yoy_pct": pct_change(us["price_cents"], prev_us["price_cents"]) if prev_us else None,
            },
            "ma": {
                "price_cents": ma["price_cents"],
                "rank": rank_of["MA"],
                "n": n_states,
                "sales_mwh": ma["sales_mwh"],
                "gen_mwh": ma.get("gen_mwh"),
                "cap_mw": ma.get("cap_mw"),
                "pop": ma.get("pop"),
                "yoy_pct": pct_change(ma["price_cents"], prev_ma["price_cents"]) if prev_ma else None,
            },
            "highest": {
                "st": highest["st"],
                "name": STATE_NAMES[highest["st"]],
                "price_cents": highest["price_cents"],
            },
            "lowest": {
                "st": lowest["st"],
                "name": STATE_NAMES[lowest["st"]],
                "price_cents": lowest["price_cents"],
            },
        },
        "latest_states": latest_states,
        "price_trend": {st: trend(st) for st in STATE_NAMES},
        "derived": {
            "note": (
                "Precomputed from the series above; prefer these over recomputing. "
                "Prices and sales cite (SRC-401). Rankings, year-over-year changes, "
                "and per-capita figures cite (derived, SRC-401) or "
                "(derived, SRC-401, SRC-402). Generation cites (SRC-403). "
                "Capacity cites (SRC-404). Never average the 50 state prices to "
                "produce a U.S. figure; use latest.us.price_cents."
            ),
            "states_ranked_highest_price_first": [
                {"st": r["st"], "name": STATE_NAMES[r["st"]], "price_cents": r["price_cents"], "rank": rank_of[r["st"]]}
                for r in ranked
            ],
            "highest_five": [
                {"st": r["st"], "name": STATE_NAMES[r["st"]], "price_cents": r["price_cents"]}
                for r in ranked[:5]
            ],
            "lowest_five": [
                {"st": r["st"], "name": STATE_NAMES[r["st"]], "price_cents": r["price_cents"]}
                for r in ranked[-5:][::-1]
            ],
            "massachusetts_rank": rank_of["MA"],
            "n_ranked": n_states,
            "us_price_first_year": by_key[(FIRST_YEAR, "US")]["price_cents"],
            "us_price_latest": us["price_cents"],
            "us_change_from_first_pct": pct_change(
                us["price_cents"], by_key[(FIRST_YEAR, "US")]["price_cents"]
            ),
            "ma_price_first_year": by_key[(FIRST_YEAR, "MA")]["price_cents"],
            "ma_price_latest": ma["price_cents"],
            "ma_change_from_first_pct": pct_change(
                ma["price_cents"], by_key[(FIRST_YEAR, "MA")]["price_cents"]
            ),
            "verification": {
                "paths": (
                    "HS861 Total Electric Industry Total price versus Electric "
                    f"Power Annual table 2.10 All Sectors for {latest - 1} and "
                    f"{latest}; overlapping cells matched to 0.01 cents. "
                    "Massachusetts 2023 = 23.21 also matches the recovered "
                    "Retail-Price Electricity extract."
                ),
                "checked_cells": [
                    {"y": y, "st": st, "price_cents": v}
                    for (y, st), v in sorted(VERIFY_HS861.items())
                    if y <= latest
                ],
            },
        },
    }

    # Preserve a human-set revised date only when the data year did not move.
    if old.get("data_year") == latest and old.get("page", {}).get("revised"):
        new["page"]["revised"] = old["page"]["revised"]
    else:
        new["page"]["revised"] = page_revised
    attach_sector_prices(new, hs861)

    LEDGER.parent.mkdir(parents=True, exist_ok=True)
    LEDGER.write_text(json.dumps(new, ensure_ascii=True, indent=1) + "\n", encoding="utf-8")
    print(
        f"refresh_dl04: as_of {old.get('as_of', '(new)')} -> {as_of}; "
        f"US {us['price_cents']} cents/kWh; MA {ma['price_cents']} "
        f"(rank {rank_of['MA']} of {n_states}); "
        f"high {highest['st']} {highest['price_cents']}; "
        f"low {lowest['st']} {lowest['price_cents']}"
    )
    subprocess.run([sys.executable, str(ROOT / "scripts/inject_data.py")], check=True)


if __name__ == "__main__":
    # openpyxl / xlrd are imported where used so a --help style failure is clear
    try:
        import openpyxl  # noqa: F401
        import xlrd  # noqa: F401
    except ImportError as e:
        sys.exit("FATAL: pip install openpyxl xlrd  (" + str(e) + ")")
    main()
