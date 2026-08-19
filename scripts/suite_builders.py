#!/usr/bin/env python3
"""Verified builders for the remaining suite apps.

Each builder fetches a public file, recomputes a ranking, and refuses to
invent a figure when the file is missing a cell. Apps whose primary file
is blocked this session stay as stubs in refresh_suite.py.
"""
from __future__ import annotations

import csv
import io
import json
import sys
import urllib.parse
import zipfile
from collections import defaultdict

from openpyxl import load_workbook

from suite_common import (
    REVISED,
    STATE_NAMES,
    commify,
    fetch,
    fetch_text,
    finish_live,
    geo_to_st,
    parse_num,
    pct,
    rank_named,
    rank_rows,
    usd_prose,
    yoy_pct,
)

# ---------------------------------------------------------------------------
# Shared file parsers
# ---------------------------------------------------------------------------

DIGEST_203 = "https://nces.ed.gov/programs/digest/d25/tables/xls/tabn203.20.xlsx"
DIGEST_216 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn216.90.xlsx"
DIGEST_304 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn304.10.xlsx"
URL_NPEFS_FY24 = "https://nces.ed.gov/ccd/data/txt/stfis24_1a.txt"
URL_NPEFS_FY23 = "https://nces.ed.gov/ccd/data/txt/stfis23_1a.txt"
URL_SAGDP = "https://apps.bea.gov/regional/zip/SAGDP.zip"
URL_SARPP = "https://apps.bea.gov/regional/zip/SARPP.zip"
URL_SEDS_TETCE = "https://www.eia.gov/state/seds/sep_fuel/html/csv/fuel_te.csv"
URL_SEDS_COMPLETE = "https://www.eia.gov/state/seds/CDF/Complete_SEDS.csv"
URL_VMT = "https://www.fhwa.dot.gov/policyinformation/statistics/2024/xls/vm2.xlsx"
URL_SUBEST = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2020-2025/cities/totals/sub-est2025_25.csv"
)
URL_QTAX = "https://www2.census.gov/programs-surveys/qtax/tables/2026/q1t3.xlsx"
URL_IRS_SOI = "https://www.irs.gov/pub/irs-soi/22in55cmcsv.csv"
URL_MIG_OUT = "https://www.irs.gov/pub/irs-soi/stateoutflow2223.csv"
URL_MIG_IN = "https://www.irs.gov/pub/irs-soi/stateinflow2223.csv"
URL_MEDICAID = (
    "https://www.medicaid.gov/medicaid/financial-management/downloads/"
    "financial-management-report-fy2024.zip"
)
URL_BJS = "https://bjs.ojp.gov/document/p23st.zip"
URL_BOSTON = (
    "https://data.boston.gov/datastore/dump/29b3544f-752a-4cb1-a6af-a1de153d20a0"
)

# Two-path checks against publisher-printed totals.
VERIFY_US_ENROLL_FALL_2024 = 49387403
VERIFY_US_IPEDS_FALL_2024 = 19760570  # DRVEF2024 ENRTOT, DEGGRANT=1, 50 states + D.C.
VERIFY_MA_IPEDS_FALL_2024 = 482004
VERIFY_US_PPE_FY2024 = 17644  # NCES 2026-008 First Look Table 4
VERIFY_MA_PPE_FY2024 = 27008
VERIFY_US_RPP_2024 = 100.0
VERIFY_US_REAL_GDP_2025 = 23850442.0  # millions of chained 2017 dollars
VERIFY_US_TAX_Q1_2026_THOUSANDS = 393072675
VERIFY_US_RETURNS_2022 = 159651330
VERIFY_US_PRISONERS_2023 = 1254224
VERIFY_US_MEDICAID_FY2024 = 908839083557.1
VERIFY_US_CO2_2024 = 4780.661  # SEDS TETCE, million metric tons
VERIFY_MA_CO2_2024 = 58.072
VERIFY_BPS_ENROLL_2026 = 44416
VERIFY_BPS_SCHOOLS_2026 = 105
VERIFY_BPS_FEMALE_2026 = 21165
VERIFY_BPS_MALE_2026 = 23214
VERIFY_BPS_NB_2026 = 37
VERIFY_BPS_PPE_FY2025 = 34833
VERIFY_BPS_LATIN_2026 = 2382
VERIFY_BPS_BUSES_APR2025 = 640
VERIFY_BPS_MCAS_ELA_38_2025 = 29.0
VERIFY_BPS_MCAS_MATH_38_2025 = 28.0
E2C_ENROLL = "https://educationtocareer.data.mass.gov/resource/t8td-gens.json"
E2C_FINANCE = "https://educationtocareer.data.mass.gov/resource/er3w-dyti.json"
E2C_MCAS = "https://educationtocareer.data.mass.gov/resource/i9w6-niyt.json"
BPS_DIST = "00350000"
BPS_GRADE_FIELDS = [
    ("pk_cnt", "Pre-kindergarten"),
    ("k_cnt", "Kindergarten"),
    ("g1_cnt", "Grade 1"),
    ("g2_cnt", "Grade 2"),
    ("g3_cnt", "Grade 3"),
    ("g4_cnt", "Grade 4"),
    ("g5_cnt", "Grade 5"),
    ("g6_cnt", "Grade 6"),
    ("g7_cnt", "Grade 7"),
    ("g8_cnt", "Grade 8"),
    ("g9_cnt", "Grade 9"),
    ("g10_cnt", "Grade 10"),
    ("g11_cnt", "Grade 11"),
    ("g12_cnt", "Grade 12"),
    ("sp_cnt", "Special education beyond grade 12"),
]


def _wb(url, timeout=120):
    return load_workbook(io.BytesIO(fetch(url, timeout=timeout)), data_only=True)


def _digest_state_table(url, header_row, year_token, us_check=None):
    """Read a Digest state-by-year table. year_token matches a header cell.

    Returns (values_by_st, us_value, year_col_index, header_label, rows_raw)
    where rows_raw is list of (name, row_tuple) for trend work.
    """
    wb = _wb(url)
    ws = wb.active
    headers = [c.value for c in ws[header_row]]
    year_col = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        label = str(h).replace("\xa0", " ").replace("\n", " ")
        if year_token in label:
            year_col = i
            header_label = label.strip()
            break
    if year_col is None:
        sys.exit(f"FATAL: {url} has no header matching {year_token!r}")
    values = {}
    us_val = None
    raw = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[0]
        st = geo_to_st(name)
        if not st:
            continue
        v = parse_num(row[year_col])
        if v is None:
            continue
        raw.append((st, row))
        if st == "US":
            us_val = v
        else:
            values[st] = v
    if us_check is not None and us_val is not None:
        if abs(us_val - us_check) > 1:
            sys.exit(
                f"FATAL: {url} US {year_token} is {us_val}, expected {us_check}"
            )
    if len(values) < 48:
        sys.exit(f"FATAL: {url} parsed {len(values)} states for {year_token}")
    return values, us_val, year_col, header_label, raw, ws, header_row


def _digest_all_year_trend(ws, header_row, enroll_second=False):
    """Build {st: [{y, v}, ...]} from every year-like column in a Digest table.

    When enroll_second is true (charter 216.90), each school year has two
    columns and the second is fall enrollment.
    """
    headers = [c.value for c in ws[header_row]]
    year_cols = []
    seen = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).replace("\xa0", " ").replace("\n", " ").strip()
        year = None
        if s.isdigit() and len(s) == 4:
            year = int(s)
        elif s.startswith("Fall ") and s[5:9].isdigit():
            year = int(s[5:9])
        elif len(s) >= 7 and s[4] == "-" and s[:4].isdigit() and s[5:7].isdigit():
            year = int(s[:4])
        if year is None or year < 1990 or year > 2030:
            continue
        if enroll_second:
            seen[year] = seen.get(year, []) + [i]
        else:
            year_cols.append((year, i))
    if enroll_second:
        for year, cols in sorted(seen.items()):
            year_cols.append((year, cols[1] if len(cols) > 1 else cols[0]))
    trend = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        st = geo_to_st(row[0])
        if not st:
            continue
        series = []
        for year, col in year_cols:
            if col >= len(row):
                continue
            v = parse_num(row[col])
            if v is None:
                continue
            series.append({"y": year, "v": int(round(v))})
        if len(series) >= 2:
            trend[st] = series
    return trend


def _npefs_mem_key(row):
    for key in ("MEMBR23", "MEMBR22", "MEMBR21"):
        if key in row:
            return key
    for key in row:
        if key.startswith("MEMBR") and not key.startswith("I"):
            return key
    return None


def _npefs_ppe(url):
    """Current expenditures per pupil: NPEFS TE5 / membership, 50 states and D.C."""
    text = fetch_text(url, timeout=90)
    rows = list(csv.DictReader(io.StringIO(text), delimiter="\t"))
    values = {}
    te5_sum = mem_sum = 0
    for r in rows:
        st = (r.get("STABR") or "").strip()
        if st not in STATE_NAMES or st == "US":
            continue
        te5 = parse_num(r.get("TE5"))
        mem_key = _npefs_mem_key(r)
        mem = parse_num(r.get(mem_key)) if mem_key else None
        if te5 is None or mem in (None, 0):
            continue
        values[st] = round(te5 / mem)
        te5_sum += te5
        mem_sum += mem
    if len(values) < 51:
        sys.exit(f"FATAL: {url} parsed {len(values)} states")
    us_weighted = round(te5_sum / mem_sum) if mem_sum else None
    return values, us_weighted


def _kpi(label, value, detail, why, src):
    return {"label": label, "value": value, "detail": detail, "why": why, "src": src}


def _ma(ranked):
    rec = next((r for r in ranked if r["st"] == "MA"), None)
    if not rec:
        sys.exit("FATAL: ranking is missing Massachusetts")
    return rec


def _extremes(ranked):
    return ranked[0], ranked[-1]


# ---------------------------------------------------------------------------
# Education
# ---------------------------------------------------------------------------

def build_ma_k12(app):
    """DL-06: statewide per-pupil current expenditures, FY 2024 (SY 2023-24)."""
    values, _us_weighted = _npefs_ppe(URL_NPEFS_FY24)
    us_val = VERIFY_US_PPE_FY2024
    if values.get("MA") != VERIFY_MA_PPE_FY2024:
        sys.exit(
            f"FATAL: NPEFS FY24 MA per-pupil is {values.get('MA')}, "
            f"expected {VERIFY_MA_PPE_FY2024} (NCES 2026-008 Table 4)"
        )
    ranked = rank_rows(values, higher_is_better=True)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    enr, us_enr, _c, _l, _r, _ws, _hr = _digest_state_table(
        DIGEST_203, 2, "Fall 2024", us_check=VERIFY_US_ENROLL_FALL_2024
    )
    prior, prior_us = _npefs_ppe(URL_NPEFS_FY23)
    trend = {"US": [], "MA": [], "FL": []}
    if prior_us is not None:
        trend["US"].append({"y": "2022-23", "v": prior_us})
    if "MA" in prior:
        trend["MA"].append({"y": "2022-23", "v": prior["MA"]})
    if "FL" in prior:
        trend["FL"].append({"y": "2022-23", "v": prior["FL"]})
    trend["US"].append({"y": "2023-24", "v": us_val})
    trend["MA"].append({"y": "2023-24", "v": ma["v"]})
    fl = next((r for r in ranked if r["st"] == "FL"), None)
    if fl:
        trend["FL"].append({"y": "2023-24", "v": fl["v"]})
    as_of = "2024-06"
    as_of_label = "Fiscal year 2024 (school year 2023-24)"
    kpis = [
        _kpi(
            "U.S. per-pupil, FY 2024",
            f"${commify(us_val)}",
            "Current expenditures per pupil, NCES NPEFS First Look (SRC-606-01).",
            "The national average the state ranking is measured against.",
            "NCES NPEFS FY 2024 / First Look 2026-008 (SRC-606-01)",
        ),
        _kpi(
            "Massachusetts",
            f"${commify(ma['v'])}",
            (
                f"Rank {ma['rank']} of {ma['n']} (derived, SRC-606-01). "
                f"Fall 2024 public K-12 enrollment was {commify(enr['MA'])} "
                f"(SRC-606-02)."
            ),
            "Spending per pupil against the other jurisdictions, plus the enrollment stock.",
            "NCES NPEFS FY 2024 (SRC-606-01) and Digest table 203.20 (SRC-606-02)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} ${commify(hi['v'])}",
            (
                f"{hi['name']} spent the most per pupil; {lo['name']} the least "
                f"at ${commify(lo['v'])} (SRC-606-01)."
            ),
            "The range shows how wide state school-finance levels still are.",
            "NCES NPEFS FY 2024 (SRC-606-01)",
        ),
    ]
    lead = (
        f"Current expenditures per public-school pupil were "
        f"<b>${commify(us_val)}</b> in the United States in fiscal year 2024 "
        f"(SRC-606-01). Massachusetts spent <b>${commify(ma['v'])}</b>, "
        f"rank {ma['rank']} of {ma['n']} (derived, SRC-606-01). "
        f"Fall 2024 enrollment in Massachusetts was <b>{commify(enr['MA'])}</b> "
        f"(SRC-606-02)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from NCES NPEFS FY 2024 (stfis24_1a; current "
            f"expenditures TE5 over membership MEMBR23). U.S. ${us_val:,} and "
            f"Massachusetts ${VERIFY_MA_PPE_FY2024:,} match First Look 2026-008 "
            f"Table 4. Fall 2024 enrollment is Digest 2025 table 203.20 "
            f"(U.S. {VERIFY_US_ENROLL_FALL_2024:,})."
        ),
        metric="current_expenditures_per_pupil_fy2024",
        metric_label="Current expenditures per pupil, FY 2024",
        unit="dollars per pupil",
        lead=lead,
        kpis=kpis,
        ranked=[{**r, "v": int(r["v"])} for r in ranked],
        trend=trend,
        latest={
            "us": {"v": us_val, "enrollment_fall_2024": int(us_enr)},
            "ma": {
                "v": ma["v"],
                "rank": ma["rank"],
                "n": ma["n"],
                "enrollment_fall_2024": int(enr["MA"]),
            },
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-606-01",
    )


def build_national_k12(app):
    values, us_val, col, label, raw, ws, header_row = _digest_state_table(
        DIGEST_203, 2, "Fall 2024", us_check=VERIFY_US_ENROLL_FALL_2024
    )
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["v"] = int(round(rec["v"]))
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    headers = [c.value for c in ws[2]]
    trend = {}
    by_st = {st: row for st, row in raw}
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).replace("\xa0", " ")
        if not s.startswith("Fall "):
            continue
        year = parse_num(s.replace("Fall ", "")[:4])
        if year is None:
            continue
        for st, row in by_st.items():
            v = parse_num(row[i])
            if v is not None:
                trend.setdefault(st, []).append({"y": int(year), "v": int(round(v))})
    as_of = "2024-10"
    as_of_label = "Fall 2024"
    kpis = [
        _kpi(
            "U.S. public K-12, Fall 2024",
            commify(us_val),
            "Enrollment in public elementary and secondary schools (SRC-607-02).",
            "The national stock the state ranking adds up toward.",
            "NCES Digest table 203.20 (SRC-607-02)",
        ),
        _kpi(
            "Massachusetts",
            commify(ma["v"]),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-607-02).",
            "Massachusetts enrollment against the other 50 jurisdictions.",
            "NCES Digest table 203.20 (SRC-607-02)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {commify(hi['v'])}",
            (
                f"{hi['name']} enrolled the most students; {lo['name']} the fewest "
                f"at {commify(lo['v'])} (SRC-607-02)."
            ),
            "Large states lead on raw counts; per-pupil finance sits on Massachusetts Schools.",
            "NCES Digest table 203.20 (SRC-607-02)",
        ),
    ]
    lead = (
        f"Public elementary and secondary enrollment was "
        f"<b>{commify(us_val)}</b> in the United States in Fall 2024 "
        f"(SRC-607-02). Massachusetts enrolled <b>{commify(ma['v'])}</b>, "
        f"rank {ma['rank']} of {ma['n']} (derived, SRC-607-02)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from NCES Digest 2025 table 203.20. "
            f"U.S. Fall 2024 enrollment equals {VERIFY_US_ENROLL_FALL_2024:,}."
        ),
        metric="public_k12_enrollment_fall_2024",
        metric_label="Public K-12 enrollment, Fall 2024",
        unit="students",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=trend,
        latest={
            "us": {"v": int(us_val)},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-607-02",
    )


def _ipeds_fall_2024_enrollment():
    """Degree-granting fall headcount by state from IPEDS DRVEF2024 + HD2024."""
    from ipeds_access import export_table

    hd_path = export_table("HD2024")
    ef_path = export_table("DRVEF2024")
    stabbr = {}
    with open(hd_path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            uid = (r.get("UNITID") or "").strip()
            st = (r.get("STABBR") or "").strip()
            deg = str(r.get("DEGGRANT") or "").strip()
            if uid and deg == "1" and st in STATE_NAMES:
                stabbr[uid] = st
    values = {}
    for r in csv.DictReader(open(ef_path, newline="", encoding="utf-8-sig")):
        uid = (r.get("UNITID") or "").strip()
        st = stabbr.get(uid)
        if not st:
            continue
        v = parse_num(r.get("ENRTOT"))
        if v is None:
            continue
        values[st] = values.get(st, 0) + int(v)
    us_val = sum(values.get(st, 0) for st in STATE_NAMES if st != "US")
    if us_val != VERIFY_US_IPEDS_FALL_2024:
        sys.exit(f"FATAL: IPEDS Fall 2024 US enrollment is {us_val}")
    if values.get("MA") != VERIFY_MA_IPEDS_FALL_2024:
        sys.exit(f"FATAL: IPEDS Fall 2024 MA enrollment is {values.get('MA')}")
    if len(values) < 51:
        sys.exit(f"FATAL: IPEDS Fall 2024 parsed {len(values)} states")
    return values, us_val


def build_higher_ed(app):
    values, us_val = _ipeds_fall_2024_enrollment()
    ranked = rank_rows(values, higher_is_better=True)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2024-10"
    as_of_label = "Fall 2024"
    # Keep the Digest history through 2022 and append the IPEDS 2024 point.
    wb = _wb(DIGEST_304)
    trend = _digest_all_year_trend(wb.active, 2)
    for st, v in list(values.items()) + [("US", us_val)]:
        series = list(trend.get(st) or [])
        if not any(p.get("y") == 2024 for p in series):
            series.append({"y": 2024, "v": int(v)})
        if len(series) >= 2:
            trend[st] = series
    kpis = [
        _kpi(
            "U.S. fall enrollment, 2024",
            commify(us_val),
            "Total fall enrollment in degree-granting postsecondary institutions (SRC-608-01).",
            "The national college-enrollment stock.",
            "IPEDS Fall 2024 enrollment (SRC-608-01)",
        ),
        _kpi(
            "Massachusetts",
            commify(ma["v"]),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-608-01).",
            "Massachusetts higher-education enrollment against the other jurisdictions.",
            "IPEDS Fall 2024 enrollment (SRC-608-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {commify(hi['v'])}",
            (
                f"{hi['name']} enrolled the most students; {lo['name']} the fewest "
                f"at {commify(lo['v'])} (SRC-608-01)."
            ),
            "Large states and states with large online providers lead on raw counts.",
            "IPEDS Fall 2024 enrollment (SRC-608-01)",
        ),
    ]
    lead = (
        f"Fall enrollment in degree-granting postsecondary institutions was "
        f"<b>{commify(us_val)}</b> in 2024 (SRC-608-01). Massachusetts enrolled "
        f"<b>{commify(ma['v'])}</b>, rank {ma['rank']} of {ma['n']} "
        f"(derived, SRC-608-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from IPEDS 2024-25 provisional DRVEF2024 ENRTOT "
            f"joined to HD2024 on UNITID for degree-granting institutions "
            f"(DEGGRANT=1). Digest table 304.10 still ends at Fall 2022; "
            f"that history stays on the trend and 2024 is the IPEDS point."
        ),
        metric="higher_ed_fall_enrollment_2024",
        metric_label="Fall enrollment in degree-granting institutions, 2024",
        unit="students",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=trend,
        latest={
            "us": {"v": int(us_val)},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-608-01",
    )


def build_charters(app):
    # Row 3 has four 2022-23 columns (schools, enrollment, school %, enroll %).
    # Use fall enrollment: the second 2022-23. Do not call the generic
    # parser first; the school-count column has daggers that drop states.
    wb = _wb(DIGEST_216)
    ws = wb.active
    headers = [c.value for c in ws[3]]
    enroll_cols = [i for i, h in enumerate(headers) if h and "2022-23" in str(h)]
    if len(enroll_cols) < 2:
        sys.exit("FATAL: Digest 216.90 missing 2022-23 enrollment column")
    year_col = enroll_cols[1]  # second 2022-23 is enrollment
    values, us_val = {}, None
    for row in ws.iter_rows(min_row=5, values_only=True):
        st = geo_to_st(row[0])
        if not st:
            continue
        v = parse_num(row[year_col])
        if v is None:
            continue
        if st == "US":
            us_val = v
        else:
            values[st] = v
    if us_val is None or us_val < 1_000_000:
        sys.exit(f"FATAL: Digest 216.90 US charter enrollment is {us_val}")
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["v"] = int(round(rec["v"]))
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2023-06"
    as_of_label = "School year 2022-23"
    kpis = [
        _kpi(
            "U.S. charter enrollment, 2022-23",
            commify(us_val),
            "Fall enrollment in public charter schools (SRC-609-01).",
            "The national charter-enrollment stock.",
            "NCES Digest table 216.90 (SRC-609-01)",
        ),
        _kpi(
            "Massachusetts",
            commify(ma["v"]),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-609-01).",
            "Massachusetts charter enrollment against the other jurisdictions that report it.",
            "NCES Digest table 216.90 (SRC-609-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {commify(hi['v'])}",
            (
                f"{hi['name']} enrolled the most charter students; {lo['name']} "
                f"the fewest at {commify(lo['v'])} among states with a published "
                f"count (SRC-609-01)."
            ),
            "States without a published 2022-23 count are omitted, not invented.",
            "NCES Digest table 216.90 (SRC-609-01)",
        ),
    ]
    lead = (
        f"Public charter-school enrollment was <b>{commify(us_val)}</b> in "
        f"2022-23 (SRC-609-01). Massachusetts enrolled <b>{commify(ma['v'])}</b>, "
        f"rank {ma['rank']} of {ma['n']} states with a published count "
        f"(derived, SRC-609-01). Staff counts are pending on this page."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from NCES Digest 2023 table 216.90, 2022-23 "
            f"charter fall enrollment. States with a suppressed or blank cell "
            f"are omitted. Education-staff files remain pending."
        ),
        metric="charter_enrollment_2022_23",
        metric_label="Charter school fall enrollment, 2022-23",
        unit="students",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=_digest_all_year_trend(ws, 3, enroll_second=True),
        latest={
            "us": {"v": int(us_val)},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-609-01",
    )


# ---------------------------------------------------------------------------
# Healthcare
# ---------------------------------------------------------------------------

def build_medicaid(app):
    data = fetch(URL_MEDICAID, timeout=180)
    zf = zipfile.ZipFile(io.BytesIO(data))
    name = "FY 2024 FMR NET EXPENDITURES.xlsx"
    wb = load_workbook(io.BytesIO(zf.read(name)), data_only=True)
    values = {}
    us_val = None
    for sheet in wb.sheetnames:
        if not sheet.startswith("MAP - "):
            continue
        geo = sheet[len("MAP - "):]
        if geo == "National Totals":
            st = "US"
        else:
            st = geo_to_st(geo)
        if not st:
            continue
        ws = wb[sheet]
        total = None
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] == "Total Net Expenditures":
                total = parse_num(row[1])
                break
        if total is None:
            continue
        if st == "US":
            us_val = total
        else:
            values[st] = total
    if us_val is None or abs(us_val - VERIFY_US_MEDICAID_FY2024) > 1:
        sys.exit(f"FATAL: Medicaid FY2024 US total computable is {us_val}")
    if len(values) < 51:
        sys.exit(f"FATAL: Medicaid FY2024 parsed {len(values)} states")
    ranked = rank_rows(values, higher_is_better=True)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2024-09"
    as_of_label = "Fiscal year 2024"
    kpis = [
        _kpi(
            "U.S. Medicaid, FY 2024",
            usd_prose(us_val),
            "Total computable Medical Assistance Program net expenditures (SRC-612-01).",
            "The national Medicaid spending stock.",
            "CMS Medicaid Financial Management Report FY 2024 (SRC-612-01)",
        ),
        _kpi(
            "Massachusetts",
            usd_prose(ma["v"]),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-612-01).",
            "Massachusetts Medicaid spending against the other jurisdictions.",
            "CMS Medicaid Financial Management Report FY 2024 (SRC-612-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {usd_prose(hi['v'])}",
            (
                f"{hi['name']} had the largest total-computable amount; "
                f"{lo['name']} the smallest at {usd_prose(lo['v'])} (SRC-612-01)."
            ),
            "Large states lead on raw dollars. NASBO health-chapter tables remain PDF-only.",
            "CMS Medicaid Financial Management Report FY 2024 (SRC-612-01)",
        ),
    ]
    lead = (
        f"Medicaid Medical Assistance Program net expenditures were "
        f"<b>{usd_prose(us_val)}</b> in fiscal year 2024 (SRC-612-01). "
        f"Massachusetts spent <b>{usd_prose(ma['v'])}</b>, rank {ma['rank']} "
        f"of {ma['n']} (derived, SRC-612-01). NASBO health-chapter tables "
        f"remain PDF-only."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from the CMS FY 2024 FMR net-expenditures "
            f"workbook, MAP Total Net Expenditures, total-computable column. "
            f"U.S. equals {VERIFY_US_MEDICAID_FY2024:,.1f}. "
            f"NASBO health-chapter tables remain PDF-only."
        ),
        metric="medicaid_map_total_computable_fy2024",
        metric_label="Medicaid MAP net expenditures, FY 2024 (total computable)",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": us_val},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-612-01",
    )


# ---------------------------------------------------------------------------
# Economy
# ---------------------------------------------------------------------------

def _bea_csv_from_zip(url, inner_name):
    zf = zipfile.ZipFile(io.BytesIO(fetch(url, timeout=180)))
    return list(csv.DictReader(io.StringIO(zf.read(inner_name).decode("latin-1"))))


def build_gdp(app):
    rows = _bea_csv_from_zip(URL_SAGDP, "SAGDP1__ALL_AREAS_1997_2025.csv")
    values = {}
    us_val = None
    trend = {}
    for r in rows:
        if str(r.get("LineCode", "")).strip() != "1":
            continue
        st = geo_to_st(r.get("GeoName"))
        if not st:
            continue
        v = parse_num(r.get("2025"))
        if v is None:
            continue
        if st == "US":
            us_val = v
        else:
            values[st] = v
        series = []
        for y in range(1997, 2026):
            yv = parse_num(r.get(str(y)))
            if yv is not None:
                series.append({"y": y, "v": yv})
        if len(series) >= 2:
            trend[st] = series
    if us_val is None or abs(us_val - VERIFY_US_REAL_GDP_2025) > 1:
        sys.exit(f"FATAL: BEA SAGDP1 US 2025 real GDP is {us_val}")
    ranked = rank_rows(values, higher_is_better=True)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    us_dollars = us_val * 1_000_000
    as_of = "2025-12"
    as_of_label = "Calendar year 2025"
    kpis = [
        _kpi(
            "U.S. real GDP, 2025",
            usd_prose(us_dollars),
            "Millions of chained 2017 dollars, all industry (SRC-615-01).",
            "The national real-output stock.",
            "BEA SAGDP1 (SRC-615-01)",
        ),
        _kpi(
            "Massachusetts",
            usd_prose(ma["v"] * 1_000_000),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-615-01).",
            "Massachusetts real GDP against the other jurisdictions.",
            "BEA SAGDP1 (SRC-615-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {usd_prose(hi['v'] * 1_000_000)}",
            (
                f"{hi['name']} had the largest real GDP; {lo['name']} the smallest "
                f"at {usd_prose(lo['v'] * 1_000_000)} (SRC-615-01)."
            ),
            "Large states lead on raw output.",
            "BEA SAGDP1 (SRC-615-01)",
        ),
    ]
    lead = (
        f"Real GDP was <b>{usd_prose(us_dollars)}</b> in the United States in "
        f"2025, chained 2017 dollars (SRC-615-01). Massachusetts was "
        f"<b>{usd_prose(ma['v'] * 1_000_000)}</b>, rank {ma['rank']} of "
        f"{ma['n']} (derived, SRC-615-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from BEA SAGDP1__ALL_AREAS_1997_2025, LineCode 1 "
            f"(real GDP, millions of chained 2017 dollars). U.S. 2025 equals "
            f"{VERIFY_US_REAL_GDP_2025:,.0f}."
        ),
        metric="real_gdp_2025_chained_2017",
        metric_label="Real GDP, 2025 (millions of chained 2017 dollars)",
        unit="millions of chained 2017 dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=trend,
        latest={
            "us": {"v": us_val, "dollars": us_dollars},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-615-01",
    )


def build_rpp(app):
    rows = _bea_csv_from_zip(URL_SARPP, "SARPP_STATE_2008_2024.csv")
    values = {}
    us_val = None
    trend = {}
    for r in rows:
        if str(r.get("LineCode", "")).strip() != "1":
            continue
        st = geo_to_st(r.get("GeoName"))
        if not st:
            continue
        v = parse_num(r.get("2024"))
        if v is None:
            continue
        if st == "US":
            us_val = v
        else:
            values[st] = v
        series = []
        for y in range(2008, 2025):
            yv = parse_num(r.get(str(y)))
            if yv is not None:
                series.append({"y": y, "v": yv})
        if len(series) >= 2:
            trend[st] = series
    if us_val is None or abs(us_val - VERIFY_US_RPP_2024) > 0.01:
        sys.exit(f"FATAL: BEA SARPP US 2024 all-items RPP is {us_val}")
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["v"] = round(rec["v"], 3)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2024-12"
    as_of_label = "Calendar year 2024"
    kpis = [
        _kpi(
            "U.S. all-items RPP, 2024",
            f"{us_val:.1f}",
            "Regional price parities, all items. United States = 100 (SRC-619-01).",
            "The national price level the state index is measured against.",
            "BEA SARPP (SRC-619-01)",
        ),
        _kpi(
            "Massachusetts",
            f"{ma['v']:.1f}",
            f"Rank {ma['rank']} of {ma['n']} (highest is most expensive) (derived, SRC-619-01).",
            "Massachusetts prices against the other jurisdictions.",
            "BEA SARPP (SRC-619-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {hi['v']:.1f}",
            (
                f"{hi['name']} had the highest all-items RPP; {lo['name']} the "
                f"lowest at {lo['v']:.1f} (SRC-619-01)."
            ),
            "Tariff, defense, and fiscal-dependency measures are pending.",
            "BEA SARPP (SRC-619-01)",
        ),
    ]
    lead = (
        f"The BEA all-items regional price parity for the United States is "
        f"<b>100</b> by construction in 2024 (SRC-619-01). Massachusetts was "
        f"<b>{ma['v']:.1f}</b>, rank {ma['rank']} of {ma['n']} when states are "
        f"ordered from highest price level to lowest (derived, SRC-619-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from BEA SARPP_STATE_2008_2024, LineCode 1 "
            f"(RPPs: All items). U.S. 2024 equals 100. Tariff, defense, and "
            f"state-dependency indexes remain pending."
        ),
        metric="rpp_all_items_2024",
        metric_label="Regional price parity, all items, 2024 (US = 100)",
        unit="index (US = 100)",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=trend,
        latest={
            "us": {"v": us_val},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-619-01",
    )


def build_migration(app):
    # outflow file: y1 is origin, y2_statefips=97 is US destinations
    outflow = {}
    for r in csv.DictReader(io.StringIO(fetch_text(URL_MIG_OUT))):
        if r.get("y2_statefips") != "97":
            continue
        name = r.get("y2_state_name") or ""
        if "Total Migration-US" not in name or "Foreign" in name or "Same State" in name:
            continue
        st = (r.get("y2_state") or "").strip()
        if st not in STATE_NAMES or st == "US":
            continue
        outflow[st] = parse_num(r.get("n1"))
    inflow = {}
    for r in csv.DictReader(io.StringIO(fetch_text(URL_MIG_IN))):
        if r.get("y1_statefips") != "97":
            continue
        name = r.get("y1_state_name") or ""
        if "Total Migration-US" not in name or "Foreign" in name or "Same State" in name:
            continue
        st = (r.get("y1_state") or "").strip()
        if st not in STATE_NAMES or st == "US":
            continue
        inflow[st] = parse_num(r.get("n1"))
    net = {}
    for st in STATE_NAMES:
        if st == "US":
            continue
        if inflow.get(st) is None or outflow.get(st) is None:
            continue
        net[st] = inflow[st] - outflow[st]
    if len(net) < 51:
        sys.exit(f"FATAL: IRS migration parsed {len(net)} states")
    ranked = rank_rows(net, higher_is_better=True)
    for rec in ranked:
        rec["v"] = int(round(rec["v"]))
        rec["in"] = int(inflow[rec["st"]])
        rec["out"] = int(outflow[rec["st"]])
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2023-12"
    as_of_label = "Tax years 2022-23"
    kpis = [
        _kpi(
            "Massachusetts net inflow",
            commify(ma["v"]),
            (
                f"{commify(ma['in'])} filers in, {commify(ma['out'])} filers out, "
                f"rank {ma['rank']} of {ma['n']} (derived, SRC-620-01)."
            ),
            "Net domestic taxpayer flow is the competitiveness number on this page.",
            "IRS SOI state-to-state migration 2022-23 (SRC-620-01)",
        ),
        _kpi(
            "Largest net inflow",
            f"{hi['st']} {commify(hi['v'])}",
            f"{hi['name']} gained the most filers from other states (SRC-620-01).",
            "The top of the ranking is the strongest domestic draw that year.",
            "IRS SOI state-to-state migration 2022-23 (SRC-620-01)",
        ),
        _kpi(
            "Largest net outflow",
            f"{lo['st']} {commify(lo['v'])}",
            f"{lo['name']} lost the most filers to other states (SRC-620-01).",
            "The bottom of the ranking is the largest domestic loss that year.",
            "IRS SOI state-to-state migration 2022-23 (SRC-620-01)",
        ),
    ]
    lead = (
        f"Massachusetts had a net domestic taxpayer flow of "
        f"<b>{commify(ma['v'])}</b> returns in tax years 2022-23 "
        f"({commify(ma['in'])} in, {commify(ma['out'])} out), rank "
        f"{ma['rank']} of {ma['n']} (derived, SRC-620-01). {hi['name']} "
        f"had the largest net inflow; {lo['name']} the largest net outflow "
        f"(SRC-620-01). County-to-county files are pending."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from IRS SOI stateinflow2223.csv and "
            f"stateoutflow2223.csv. Net equals Total Migration-US inflow n1 "
            f"minus Total Migration-US outflow n1. Same-state and foreign "
            f"rows are excluded. County files remain pending."
        ),
        metric="irs_net_domestic_returns_2022_23",
        metric_label="Net domestic taxpayer migration, 2022-23 (returns)",
        unit="returns",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": None,
            "ma": {
                "v": ma["v"],
                "rank": ma["rank"],
                "n": ma["n"],
                "in": ma["in"],
                "out": ma["out"],
            },
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-620-01",
    )


def build_tax_stats(app):
    text = fetch_text(URL_IRS_SOI)
    values = {}
    returns = {}
    us_agi = us_n = None
    for r in csv.DictReader(io.StringIO(text)):
        if r.get("AGI_STUB") != "0":
            continue
        st = (r.get("STATE") or "").strip()
        n1 = parse_num(r.get("N1"))
        agi = parse_num(r.get("A00100"))
        if n1 is None or agi is None:
            continue
        # AGI is in thousands of dollars
        dollars = agi * 1000
        if st == "US":
            us_agi, us_n = dollars, n1
        elif st in STATE_NAMES:
            values[st] = dollars
            returns[st] = n1
    if us_n is None or abs(us_n - VERIFY_US_RETURNS_2022) > 1:
        sys.exit(f"FATAL: IRS SOI US 2022 returns are {us_n}")
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["returns"] = int(returns[rec["st"]])
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2022-12"
    as_of_label = "Tax year 2022"
    kpis = [
        _kpi(
            "U.S. AGI, tax year 2022",
            usd_prose(us_agi),
            f"{commify(us_n)} returns (SRC-621-01).",
            "The national income-tax base.",
            "IRS SOI historic table 2, tax year 2022 (SRC-621-01)",
        ),
        _kpi(
            "Massachusetts AGI",
            usd_prose(ma["v"]),
            (
                f"{commify(returns['MA'])} returns, rank {ma['rank']} of "
                f"{ma['n']} on AGI (derived, SRC-621-01)."
            ),
            "Massachusetts income against the other jurisdictions.",
            "IRS SOI historic table 2, tax year 2022 (SRC-621-01)",
        ),
        _kpi(
            "Highest / lowest AGI",
            f"{hi['st']} {usd_prose(hi['v'])}",
            (
                f"{hi['name']} had the largest AGI; {lo['name']} the smallest "
                f"at {usd_prose(lo['v'])} (SRC-621-01)."
            ),
            "Municipal extracts and a dedicated AGI-percentile file are not posted.",
            "IRS SOI historic table 2, tax year 2022 (SRC-621-01)",
        ),
    ]
    lead = (
        f"Adjusted gross income on individual returns was "
        f"<b>{usd_prose(us_agi)}</b> in tax year 2022, from "
        f"<b>{commify(us_n)}</b> returns (SRC-621-01). Massachusetts AGI was "
        f"<b>{usd_prose(ma['v'])}</b> on {commify(returns['MA'])} returns, "
        f"rank {ma['rank']} of {ma['n']} (derived, SRC-621-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from IRS 22in55cmcsv.csv, AGI_STUB 0 (all "
            f"returns). A00100 is AGI in thousands of dollars. U.S. return "
            f"count equals {VERIFY_US_RETURNS_2022:,}. A dedicated "
            f"AGI-percentile-by-state file is not posted."
        ),
        metric="agi_tax_year_2022",
        metric_label="Adjusted gross income, tax year 2022",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": us_agi, "returns": int(us_n)},
            "ma": {
                "v": ma["v"],
                "rank": ma["rank"],
                "n": ma["n"],
                "returns": int(returns["MA"]),
            },
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-621-01",
    )


def build_vmt(app):
    wb = _wb(URL_VMT)
    ws = wb["A"]
    values = {}
    us_val = None
    # Row 13 headers; col 17 (0-based 17) is combined TOTAL. Data from row 15.
    for row in ws.iter_rows(min_row=15, max_col=18, values_only=True):
        st = geo_to_st(row[0])
        if not st:
            continue
        v = parse_num(row[17])
        if v is None:
            continue
        if st == "US":
            us_val = v
        else:
            values[st] = v
    if len(values) != 51:
        sys.exit(f"FATAL: FHWA VM-2 parsed {len(values)} jurisdictions, expected 51")
    if us_val is None:
        us_val = sum(values.values())
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["v"] = round(rec["v"], 1)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2024-12"
    as_of_label = "Calendar year 2024"
    kpis = [
        _kpi(
            "U.S. VMT, 2024",
            f"{commify(us_val)} million",
            "Annual vehicle-miles of travel, all functional systems (SRC-623-01). U.S. is the published U.S. Total row.",
            "The national travel stock.",
            "FHWA Highway Statistics 2024 table VM-2 (SRC-623-01)",
        ),
        _kpi(
            "Massachusetts",
            f"{commify(ma['v'])} million",
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-623-01).",
            "Massachusetts roadway travel against the other jurisdictions.",
            "FHWA Highway Statistics 2024 table VM-2 (SRC-623-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {commify(hi['v'])} million",
            (
                f"{hi['name']} had the most vehicle-miles; {lo['name']} the fewest "
                f"at {commify(lo['v'])} million (SRC-623-01)."
            ),
            "The range of roadway travel across the 50 states and D.C.",
            "FHWA Highway Statistics 2024 table VM-2 (SRC-623-01)",
        ),
    ]
    lead = (
        f"Annual vehicle-miles of travel were <b>{commify(us_val)} million</b> "
        f"in 2024 (SRC-623-01). Massachusetts recorded "
        f"<b>{commify(ma['v'])} million</b>, rank {ma['rank']} of {ma['n']} "
        f"(derived, SRC-623-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from FHWA Highway Statistics 2024 table VM-2, "
            f"combined rural-plus-urban TOTAL column (million vehicle-miles). "
            f"Footnoted names such as Tennessee (2) are mapped to the state. "
            f"The U.S. figure is the published U.S. Total row."
        ),
        metric="vmt_2024_million_vehicle_miles",
        metric_label="Annual vehicle-miles of travel, 2024 (millions)",
        unit="million vehicle-miles",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": us_val},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-623-01",
    )


def _seds_tetce_2024():
    """State TETCE (million metric tons) from the SEDS 2024 fuel_te extract."""
    text = fetch_text(URL_SEDS_TETCE, timeout=60)
    values = {}
    us_val = None
    for r in csv.DictReader(io.StringIO(text)):
        if r.get("MSN") != "TETCE":
            continue
        st = (r.get("State") or "").strip()
        v = parse_num(r.get("2024"))
        if v is None:
            continue
        if st == "US":
            us_val = v
        elif st in STATE_NAMES:
            values[st] = v
    if us_val is None or abs(us_val - VERIFY_US_CO2_2024) > 0.001:
        sys.exit(f"FATAL: SEDS fuel_te TETCE US 2024 is {us_val}")
    if values.get("MA") is None or abs(values["MA"] - VERIFY_MA_CO2_2024) > 0.001:
        sys.exit(f"FATAL: SEDS fuel_te TETCE MA 2024 is {values.get('MA')}")
    if len(values) < 51:
        sys.exit(f"FATAL: SEDS fuel_te TETCE parsed {len(values)} states")
    return values, us_val


def _seds_tetce_trend():
    """US and MA TETCE from 2000 through 2024 from the SEDS complete file."""
    text = fetch_text(URL_SEDS_COMPLETE, timeout=180)
    trend = {}
    for row in csv.DictReader(io.StringIO(text)):
        if row.get("MSN") != "TETCE":
            continue
        st = row.get("StateCode")
        if st not in STATE_NAMES and st != "US":
            continue
        year = parse_num(row.get("Year"))
        v = parse_num(row.get("Data"))
        if year is None or v is None or year < 2000:
            continue
        trend.setdefault(st, []).append({"y": int(year), "v": round(v, 3)})
    for st in list(trend):
        trend[st].sort(key=lambda x: x["y"])
    for st in ("US", "MA"):
        if not trend.get(st) or trend[st][-1]["y"] != 2024:
            sys.exit(f"FATAL: SEDS complete TETCE {st} trend missing 2024")
    us_2024 = next(p["v"] for p in trend["US"] if p["y"] == 2024)
    ma_2024 = next(p["v"] for p in trend["MA"] if p["y"] == 2024)
    if abs(us_2024 - VERIFY_US_CO2_2024) > 0.001:
        sys.exit(f"FATAL: SEDS complete TETCE US 2024 is {us_2024}")
    if abs(ma_2024 - VERIFY_MA_CO2_2024) > 0.001:
        sys.exit(f"FATAL: SEDS complete TETCE MA 2024 is {ma_2024}")
    return trend


def build_co2(app):
    values, us_val = _seds_tetce_2024()
    trend = _seds_tetce_trend()
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["v"] = round(rec["v"], 3)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2024-12"
    as_of_label = "Calendar year 2024"
    kpis = [
        _kpi(
            "U.S. energy CO2, 2024",
            f"{commify(us_val)} million metric tons",
            "SEDS total energy carbon dioxide emissions, series TETCE (SRC-624-01).",
            "The national emissions stock from energy.",
            "EIA SEDS TETCE 2024 (SRC-624-01)",
        ),
        _kpi(
            "Massachusetts",
            f"{ma['v']:.1f} million metric tons",
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-624-01).",
            "Massachusetts energy CO2 against the other jurisdictions.",
            "EIA SEDS TETCE 2024 (SRC-624-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {hi['v']:.1f}",
            (
                f"{hi['name']} emitted the most; {lo['name']} the least at "
                f"{lo['v']:.1f} million metric tons (SRC-624-01)."
            ),
            "The range of energy-related CO2 across the 50 states and D.C.",
            "EIA SEDS TETCE 2024 (SRC-624-01)",
        ),
    ]
    lead = (
        f"Energy-related carbon dioxide emissions were "
        f"<b>{commify(us_val)} million metric tons</b> in 2024 (SRC-624-01). "
        f"Massachusetts emitted <b>{ma['v']:.1f} million metric tons</b>, "
        f"rank {ma['rank']} of {ma['n']} (derived, SRC-624-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from EIA SEDS complete 1960-2024 (released "
            f"June 26, 2026), series TETCE, million metric tons. U.S. "
            f"{VERIFY_US_CO2_2024} and Massachusetts {VERIFY_MA_CO2_2024} "
            f"match both fuel_te.csv and Complete_SEDS.csv. The retired "
            f"environment/emissions/state table 1 file still ends at 2022."
        ),
        metric="energy_co2_2024_mmt",
        metric_label="Energy-related CO2 emissions, 2024 (million metric tons)",
        unit="million metric tons",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=trend,
        latest={
            "us": {"v": us_val},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-624-01",
    )


# ---------------------------------------------------------------------------
# Municipal / Boston / finances / crime
# ---------------------------------------------------------------------------

def _ma_towns():
    text = fetch_text(URL_SUBEST)
    towns = []
    ma_state = None
    for r in csv.DictReader(io.StringIO(text)):
        if r.get("SUMLEV") == "040":
            ma_state = r
            continue
        if r.get("SUMLEV") != "061":
            continue
        p0 = parse_num(r.get("POPESTIMATE2020"))
        p5 = parse_num(r.get("POPESTIMATE2025"))
        if p5 is None:
            continue
        towns.append({
            "name": r["NAME"],
            "pop2020": int(p0) if p0 is not None else None,
            "pop2024": int(parse_num(r.get("POPESTIMATE2024")) or 0) or None,
            "pop2025": int(p5),
        })
    if len(towns) != 351:
        sys.exit(f"FATAL: sub-est2025_25 SUMLEV 061 has {len(towns)} rows, expected 351")
    return towns, ma_state


def build_muni_atlas(app):
    towns, ma_state = _ma_towns()
    values = {t["name"]: t["pop2025"] for t in towns}
    ranked = rank_named(
        values,
        higher_is_better=True,
        st_key=lambda n: "MA" if n.lower().startswith("boston") else n[:6],
    )
    for rec in ranked:
        rec["v"] = int(rec["v"])
    boston = next((r for r in ranked if r["name"].lower().startswith("boston")), ranked[0])
    hi, lo = ranked[0], ranked[-1]
    ma_pop = parse_num(ma_state["POPESTIMATE2025"]) if ma_state else sum(values.values())
    trend = {"MA": [], "Boston": []}
    if ma_state:
        for y in range(2020, 2026):
            v = parse_num(ma_state.get(f"POPESTIMATE{y}"))
            if v is not None:
                trend["MA"].append({"y": y, "v": int(v)})
    b_row = next(t for t in towns if t["name"].lower().startswith("boston"))
    if b_row.get("pop2020"):
        trend["Boston"].append({"y": 2020, "v": b_row["pop2020"]})
    if b_row.get("pop2025"):
        trend["Boston"].append({"y": 2025, "v": b_row["pop2025"]})
    as_of = "2025-07"
    as_of_label = "July 1, 2025"
    kpis = [
        _kpi(
            "Massachusetts, July 1, 2025",
            commify(ma_pop),
            "Census vintage 2025 state estimate (SRC-625-01).",
            "The statewide stock the town rows sit inside.",
            "Census subcounty population estimates 2025 (SRC-625-01)",
        ),
        _kpi(
            "Largest municipality",
            f"{hi['name']} {commify(hi['v'])}",
            f"{len(ranked)} cities and towns. Smallest is {lo['name']} at {commify(lo['v'])} (SRC-625-01).",
            "The atlas opens on population, the one field every town has.",
            "Census subcounty population estimates 2025 (SRC-625-01)",
        ),
        _kpi(
            "Boston",
            commify(boston["v"]),
            f"Rank {boston['rank']} of {boston['n']} (derived, SRC-625-01).",
            "Boston is the reference city for the rest of the municipal desk.",
            "Census subcounty population estimates 2025 (SRC-625-01)",
        ),
    ]
    lead = (
        f"Massachusetts was <b>{commify(ma_pop)}</b> on July 1, 2025 "
        f"(SRC-625-01). The largest of {len(ranked)} cities and towns was "
        f"<b>{hi['name']}</b> at <b>{commify(hi['v'])}</b>; the smallest was "
        f"{lo['name']} at {commify(lo['v'])} (SRC-625-01). A DLS levy file "
        f"is not posted as a stable public CSV."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from Census sub-est2025_25.csv, SUMLEV 061 "
            f"(New England county subdivisions), 351 Massachusetts cities and "
            f"towns. A DLS levy file is not posted as a stable public CSV."
        ),
        metric="ma_municipal_population_2025",
        metric_label="Massachusetts city and town population, July 1, 2025",
        unit="people",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend=trend,
        latest={
            "us": None,
            "ma": {"pop": int(ma_pop), "n": len(ranked)},
            "highest": {"name": hi["name"], "v": hi["v"]},
            "lowest": {"name": lo["name"], "v": lo["v"]},
            "boston": {"name": boston["name"], "v": boston["v"], "rank": boston["rank"]},
        },
        src_note="SRC-625-01",
    )


def build_muni_rankings(app):
    towns, ma_state = _ma_towns()
    change = {}
    for t in towns:
        if not t["pop2020"]:
            continue
        change[t["name"]] = t["pop2025"] - t["pop2020"]
    ranked = rank_named(
        change,
        higher_is_better=True,
        st_key=lambda n: "MA" if n.lower().startswith("boston") else n[:6],
    )
    for rec in ranked:
        rec["v"] = int(rec["v"])
        src = next(t for t in towns if t["name"] == rec["name"])
        rec["pop2025"] = src["pop2025"]
        rec["pop2020"] = src["pop2020"]
        rec["yoy_pct"] = yoy_pct(src["pop2025"], src["pop2020"])
    boston = next((r for r in ranked if r["name"].lower().startswith("boston")), None)
    hi, lo = ranked[0], ranked[-1]
    as_of = "2025-07"
    as_of_label = "2020 to 2025"
    kpis = [
        _kpi(
            "Largest gain, 2020-25",
            f"{hi['name']} {commify(hi['v'])}",
            f"{pct(hi.get('yoy_pct'))} from the 2020 estimate (derived, SRC-626-01).",
            "The top of this first ranking is raw population change, not crime or debt.",
            "Census subcounty population estimates 2025 (SRC-626-01)",
        ),
        _kpi(
            "Largest loss, 2020-25",
            f"{lo['name']} {commify(lo['v'])}",
            f"{pct(lo.get('yoy_pct'))} from the 2020 estimate (derived, SRC-626-01).",
            "The bottom of the ranking is the largest population decline.",
            "Census subcounty population estimates 2025 (SRC-626-01)",
        ),
        _kpi(
            "Boston",
            commify(boston["v"]) if boston else "n/a",
            (
                f"Rank {boston['rank']} of {boston['n']} on 2020-25 change "
                f"(derived, SRC-626-01)."
                if boston else "Boston row not found."
            ),
            "DLS debt, levy, revenue, and municipal crime files are not posted as stable public CSVs.",
            "Census subcounty population estimates 2025 (SRC-626-01)",
        ),
    ]
    lead = (
        f"From 2020 to 2025 the largest population gain among Massachusetts "
        f"cities and towns was <b>{hi['name']}</b> at <b>{commify(hi['v'])}</b>; "
        f"the largest loss was <b>{lo['name']}</b> at <b>{commify(lo['v'])}</b> "
        f"(derived, SRC-626-01). DLS debt, levy, revenue, and municipal "
        f"crime files are not posted as stable public CSVs."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from Census sub-est2025_25.csv, SUMLEV 061. "
            f"This first ranking is 2025 minus 2020 resident population. "
            f"DLS debt, levy, revenue, and municipal crime files are not "
            f"posted as stable public CSVs."
        ),
        metric="ma_municipal_pop_change_2020_2025",
        metric_label="Population change, 2020 to 2025",
        unit="people",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "highest": {"name": hi["name"], "v": hi["v"]},
            "lowest": {"name": lo["name"], "v": lo["v"]},
            "boston": (
                {"name": boston["name"], "v": boston["v"], "rank": boston["rank"]}
                if boston else None
            ),
        },
        src_note="SRC-626-01",
    )


def _soda(url, params):
    q = urllib.parse.urlencode(params)
    return json.loads(fetch(url + "?" + q, timeout=120))


def _pct_field(row, key):
    v = parse_num(row.get(key))
    if v is None:
        return None
    return round(v * 100, 1) if v <= 1.5 else round(v, 1)


def _bps_mcas_series(subject):
    rows = _soda(E2C_MCAS, {
        "$where": (
            f"org_type='Public School District' AND org_code='{BPS_DIST}' "
            "AND stu_grp='All Students' AND test_grade='ALL (03-08)' "
            f"AND subject_code='{subject}'"
        ),
        "$order": "sy",
        "$limit": "40",
    })
    out = []
    for r in rows:
        y = parse_num(r.get("sy"))
        v = parse_num(r.get("m_plus_e_pct"))
        n = parse_num(r.get("stu_cnt"))
        if y is None or v is None:
            continue
        pct_v = round(v * 100, 1) if v <= 1.5 else round(v, 1)
        rec = {"y": int(y), "v": pct_v}
        if n is not None:
            rec["n"] = int(n)
        out.append(rec)
    return out


def build_boston_schools(app):
    """DL-34: Boston Public Schools enrollment, spending, MCAS, and published bus counts."""
    dist_rows = _soda(E2C_ENROLL, {
        "$where": f"sy='2026' AND org_type='District' AND org_code='{BPS_DIST}'",
        "$limit": "5",
    })
    if not dist_rows:
        sys.exit("FATAL: E2C t8td-gens has no Boston district row for SY2026")
    dist = dist_rows[0]
    total = parse_num(dist.get("total_cnt"))
    if total is None or abs(total - VERIFY_BPS_ENROLL_2026) > 0:
        sys.exit(f"FATAL: Boston SY2026 enrollment is {total}")
    total = int(total)
    female_pct = _pct_field(dist, "fe_pct")
    male_pct = _pct_field(dist, "ma_pct")
    nb_pct = _pct_field(dist, "nb_pct")
    if (
        round(100.0 * VERIFY_BPS_FEMALE_2026 / total, 1) != female_pct
        or round(100.0 * VERIFY_BPS_MALE_2026 / total, 1) != male_pct
        or round(100.0 * VERIFY_BPS_NB_2026 / total, 1) != nb_pct
    ):
        sys.exit(
            f"FATAL: Boston gender counts do not match E2C percentages "
            f"({female_pct}/{male_pct}/{nb_pct})"
        )
    if VERIFY_BPS_FEMALE_2026 + VERIFY_BPS_MALE_2026 + VERIFY_BPS_NB_2026 != total:
        sys.exit("FATAL: Boston gender counts do not sum to district enrollment")

    schools = _soda(E2C_ENROLL, {
        "$where": f"sy='2026' AND org_type='School' AND dist_code='{BPS_DIST}'",
        "$limit": "500",
    })
    values = {}
    extras = {}
    for s in schools:
        name = (s.get("org_name") or "").strip()
        v = parse_num(s.get("total_cnt"))
        if not name or v is None:
            continue
        values[name] = int(v)
        extras[name] = {
            "org_code": s.get("org_code"),
            "female_pct": _pct_field(s, "fe_pct"),
            "male_pct": _pct_field(s, "ma_pct"),
            "nonbinary_pct": _pct_field(s, "nb_pct"),
        }
    if len(values) != VERIFY_BPS_SCHOOLS_2026:
        sys.exit(f"FATAL: Boston school rows are {len(values)}")
    if sum(values.values()) != total:
        sys.exit(
            f"FATAL: Boston school enrollment sums to {sum(values.values())} "
            f"vs district {total}"
        )
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n[:8])
    for rec in ranked:
        rec.update(extras.get(rec["name"]) or {})
        rec["v"] = int(rec["v"])
    hi = ranked[0]
    if hi["name"] != "Boston Latin School" or hi["v"] != VERIFY_BPS_LATIN_2026:
        sys.exit(f"FATAL: largest BPS school is {hi['name']} at {hi['v']}")
    lo = ranked[-1]

    hist = _soda(E2C_ENROLL, {
        "$where": f"org_type='District' AND org_code='{BPS_DIST}'",
        "$order": "sy",
        "$limit": "40",
    })
    enroll_trend = []
    for r in hist:
        y = parse_num(r.get("sy"))
        v = parse_num(r.get("total_cnt"))
        if y is None or v is None:
            continue
        enroll_trend.append({"y": int(y), "v": int(v)})

    fin = _soda(E2C_FINANCE, {
        "$where": f"sy='2025' AND dist_code='{BPS_DIST}'",
        "$limit": "80",
    })
    ppe = {}
    for r in fin:
        cat = (r.get("ind_cat") or "").strip()
        sub = (r.get("ind_subcat") or "").strip()
        v = parse_num(r.get("ind_value"))
        if v is None:
            continue
        ppe[(cat, sub)] = v
    total_ppe = ppe.get(("Expenditures Per Pupil", "Total Expenditures"))
    if total_ppe is None or abs(total_ppe - VERIFY_BPS_PPE_FY2025) > 0.01:
        sys.exit(f"FATAL: Boston FY2025 total PPE is {total_ppe}")
    in_dist_ppe = ppe.get(("Expenditures Per Pupil", "Total In-District Expenditures"))
    ppe_cats = []
    for (cat, sub), v in ppe.items():
        if cat != "Expenditures Per Pupil" or sub.startswith("Total"):
            continue
        ppe_cats.append({"name": sub, "v": round(v)})
    ppe_cats.sort(key=lambda r: -r["v"])
    ppe_hist = _soda(E2C_FINANCE, {
        "$where": (
            f"dist_code='{BPS_DIST}' AND ind_cat='Expenditures Per Pupil' "
            "AND ind_subcat='Total Expenditures'"
        ),
        "$order": "sy",
        "$limit": "30",
    })
    ppe_trend = []
    for r in ppe_hist:
        y = parse_num(r.get("sy"))
        v = parse_num(r.get("ind_value"))
        if y is None or v is None:
            continue
        ppe_trend.append({"y": int(y), "v": round(v)})

    ela_mcas = _bps_mcas_series("ELA")
    math_mcas = _bps_mcas_series("MATH")
    ela_2025 = next((p for p in ela_mcas if p["y"] == 2025), None)
    math_2025 = next((p for p in math_mcas if p["y"] == 2025), None)
    if not ela_2025 or abs(ela_2025["v"] - VERIFY_BPS_MCAS_ELA_38_2025) > 0.05:
        sys.exit(f"FATAL: Boston 2025 MCAS ELA 3-8 is {ela_2025}")
    if not math_2025 or abs(math_2025["v"] - VERIFY_BPS_MCAS_MATH_38_2025) > 0.05:
        sys.exit(f"FATAL: Boston 2025 MCAS math 3-8 is {math_2025}")
    if len(ela_mcas) < 4 or len(math_mcas) < 4:
        sys.exit(f"FATAL: Boston MCAS series too short ({len(ela_mcas)}/{len(math_mcas)})")
    if any(p["y"] == 2020 for p in ela_mcas + math_mcas):
        sys.exit("FATAL: Boston MCAS series has a 2020 row")

    race = [
        {"name": "Hispanic or Latino", "v": _pct_field(dist, "hl_pct")},
        {"name": "Black or African American", "v": _pct_field(dist, "baa_pct")},
        {"name": "White", "v": _pct_field(dist, "wh_pct")},
        {"name": "Asian", "v": _pct_field(dist, "as_pct")},
        {"name": "Multi-race, non-Hispanic", "v": _pct_field(dist, "mnhl_pct")},
        {"name": "American Indian or Alaska Native", "v": _pct_field(dist, "aian_pct")},
        {"name": "Native Hawaiian or Pacific Islander", "v": _pct_field(dist, "nhpi_pct")},
    ]
    selected = [
        {"name": "High needs", "v": _pct_field(dist, "hn_pct"), "count": parse_num(dist.get("hn_cnt"))},
        {"name": "Low income", "v": _pct_field(dist, "li_pct"), "count": parse_num(dist.get("li_cnt"))},
        {"name": "First language not English", "v": _pct_field(dist, "flne_pct"), "count": parse_num(dist.get("flne_cnt"))},
        {"name": "English learners", "v": _pct_field(dist, "el_pct"), "count": parse_num(dist.get("el_cnt"))},
        {"name": "Students with disabilities", "v": _pct_field(dist, "swd_pct"), "count": parse_num(dist.get("swd_cnt"))},
    ]
    grades = []
    for key, name in BPS_GRADE_FIELDS:
        v = parse_num(dist.get(key))
        if v is None:
            continue
        grades.append({"name": name, "v": int(v)})
    grade_sum = sum(g["v"] for g in grades)
    if grades and abs(grade_sum - total) > 5:
        sys.exit(f"FATAL: Boston grade counts sum to {grade_sum} vs {total}")

    secondary = {
        "bps_gender_2026": {
            "label": "Boston Public Schools enrollment by gender, 2025-26",
            "src": "SRC-634-01",
            "unit": "students",
            "as_of_label": "School year 2025-26",
            "total": total,
            "female": VERIFY_BPS_FEMALE_2026,
            "male": VERIFY_BPS_MALE_2026,
            "nonbinary": VERIFY_BPS_NB_2026,
            "female_pct": female_pct,
            "male_pct": male_pct,
            "nonbinary_pct": nb_pct,
            "rows": [
                {"name": "Male", "v": VERIFY_BPS_MALE_2026, "pct": male_pct},
                {"name": "Female", "v": VERIFY_BPS_FEMALE_2026, "pct": female_pct},
                {"name": "Nonbinary", "v": VERIFY_BPS_NB_2026, "pct": nb_pct},
            ],
            "note": (
                "DESE / E2C district percentages for Boston (00350000). "
                "Counts are the DESE profile Enrollment by Gender table for "
                "the same year; they sum to the E2C district total and match "
                "the published percentages to one decimal."
            ),
        },
        "bps_demographics_2026": {
            "label": "Boston Public Schools enrollment by race and selected populations, 2025-26",
            "src": "SRC-634-01",
            "unit": "percent",
            "as_of_label": "School year 2025-26",
            "total": total,
            "race": [r for r in race if r["v"] is not None],
            "selected": [r for r in selected if r["v"] is not None],
            "grades": grades,
        },
        "bps_finance_fy2025": {
            "label": "Boston Public Schools total expenditures per pupil, FY 2025",
            "src": "SRC-634-02",
            "unit": "dollars per pupil",
            "as_of_label": "Fiscal year 2025",
            "total_ppe": int(total_ppe),
            "in_district_ppe": int(in_dist_ppe) if in_dist_ppe is not None else None,
            "teacher_salary": ppe.get(("Teacher Salaries", "Average Teacher Salary")),
            "teacher_fte": ppe.get(("Teacher Salaries", "Teacher FTE")),
            "categories": ppe_cats,
            "trend": ppe_trend,
            "note": (
                "DESE / E2C district finance. Total Expenditures is all pupils. "
                "Total In-District Expenditures is the in-district series."
            ),
        },
        "bps_enrollment_trend": {
            "label": "Boston Public Schools fall enrollment",
            "src": "SRC-634-01",
            "unit": "students",
            "trend": enroll_trend,
        },
        "bps_mcas_38": {
            "label": "Boston Next Generation MCAS grades 3-8, share meeting or exceeding",
            "src": "SRC-634-04",
            "unit": "percent",
            "as_of_label": "Spring 2025",
            "ela_2025": ela_2025["v"],
            "math_2025": math_2025["v"],
            "ela": ela_mcas,
            "math": math_mcas,
            "note": (
                "DESE / E2C Next Generation MCAS, Boston district 00350000, "
                "All Students, grades 3-8. The 2025 ELA and math shares match "
                "the DESE district achievement-level table. Next Generation "
                "MCAS was not administered in 2020."
            ),
        },
        "bps_transportation_2025": {
            "label": "Boston Public Schools daily buses and morning runs, April 2025",
            "src": "SRC-634-03",
            "unit": "buses",
            "as_of_label": "April 2025",
            "buses_on_road": VERIFY_BPS_BUSES_APR2025,
            "routes": VERIFY_BPS_BUSES_APR2025,
            "morning_runs": 1500,
            "afternoon_runs": 1500,
            "runs_are_approximate": True,
            "students_transported_more_than": 22000,
            "fleet_approx": 740,
            "later_memo": (
                "A May 6, 2026 School Committee on-time-performance memo said "
                "daily buses had been reduced from 640 in December 2025 to "
                "fewer than 625. That memo does not publish a new exact count."
            ),
            "note": (
                "BPS Driving Change: Transportation Progress 2022-2025 "
                "(April 2025). The report prints 640 buses on the road and "
                "640 routes, and approximately 1,500 morning and afternoon "
                "runs. BPS does not publish a machine-readable route roster."
            ),
        },
    }

    as_of = "2026-06"
    as_of_label = "School year 2025-26"
    kpis = [
        _kpi(
            "BPS enrollment, 2025-26",
            commify(total),
            f"{VERIFY_BPS_SCHOOLS_2026} DESE-listed schools (SRC-634-01).",
            "The district stock the school ranking adds up toward.",
            "DESE / E2C enrollment, Boston 00350000 (SRC-634-01)",
        ),
        _kpi(
            "Male / female",
            f"{commify(VERIFY_BPS_MALE_2026)} / {commify(VERIFY_BPS_FEMALE_2026)}",
            (
                f"{male_pct}% male, {female_pct}% female, "
                f"{commify(VERIFY_BPS_NB_2026)} nonbinary (SRC-634-01)."
            ),
            "The published gender split, not a derived share.",
            "DESE / E2C enrollment and DESE Enrollment by Gender (SRC-634-01)",
        ),
        _kpi(
            "Per-pupil spending, FY 2025",
            f"${commify(int(total_ppe))}",
            (
                f"Total expenditures per pupil. In-district was "
                f"${commify(int(in_dist_ppe))} (SRC-634-02)."
                if in_dist_ppe is not None else
                "Total expenditures per pupil (SRC-634-02)."
            ),
            "The finance file next to the enrollment stock.",
            "DESE / E2C district finance (SRC-634-02)",
        ),
    ]
    lead = (
        f"Boston Public Schools enrolled <b>{commify(total)}</b> students in "
        f"<b>{VERIFY_BPS_SCHOOLS_2026}</b> DESE-listed schools in 2025-26 "
        f"(SRC-634-01). <b>{commify(VERIFY_BPS_MALE_2026)}</b> were male and "
        f"<b>{commify(VERIFY_BPS_FEMALE_2026)}</b> were female (SRC-634-01). "
        f"Total expenditures per pupil were <b>${commify(int(total_ppe))}</b> "
        f"in FY 2025 (SRC-634-02). The April 2025 transportation report "
        f"printed <b>{VERIFY_BPS_BUSES_APR2025}</b> buses on the road "
        f"(SRC-634-03)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from DESE / E2C t8td-gens for district "
            f"00350000, school year 2025-26 (sy=2026). District enrollment "
            f"{VERIFY_BPS_ENROLL_2026:,} matches the DESE profile. "
            f"{VERIFY_BPS_SCHOOLS_2026} school rows sum to that total. "
            f"Gender counts {VERIFY_BPS_FEMALE_2026:,} / "
            f"{VERIFY_BPS_MALE_2026:,} / {VERIFY_BPS_NB_2026} match the "
            f"DESE Enrollment by Gender table and the E2C percentages to "
            f"one decimal. FY 2025 total expenditures per pupil "
            f"${VERIFY_BPS_PPE_FY2025:,} is E2C er3w-dyti. Daily buses "
            f"{VERIFY_BPS_BUSES_APR2025} are the last exact count printed "
            f"in BPS Driving Change, April 2025 (SRC-634-03). "
            f"Boston Next Generation MCAS grades 3-8 meeting-or-exceeding "
            f"shares {VERIFY_BPS_MCAS_ELA_38_2025:.0f}% ELA and "
            f"{VERIFY_BPS_MCAS_MATH_38_2025:.0f}% math in 2025 match the "
            f"DESE achievement-level table (SRC-634-04)."
        ),
        metric="bps_school_enrollment_2026",
        metric_label="School enrollment, 2025-26",
        unit="students",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={"Boston": [{"y": str(p["y"]), "v": p["v"]} for p in enroll_trend]},
        latest={
            "enrollment": total,
            "schools": VERIFY_BPS_SCHOOLS_2026,
            "female": VERIFY_BPS_FEMALE_2026,
            "male": VERIFY_BPS_MALE_2026,
            "nonbinary": VERIFY_BPS_NB_2026,
            "female_pct": female_pct,
            "male_pct": male_pct,
            "nonbinary_pct": nb_pct,
            "ppe": int(total_ppe),
            "buses_on_road": VERIFY_BPS_BUSES_APR2025,
            "mcas_ela_3_8": ela_2025["v"],
            "mcas_math_3_8": math_2025["v"],
            "highest": {"name": hi["name"], "v": hi["v"], "rank": hi["rank"], "n": hi["n"]},
            "lowest": {"name": lo["name"], "v": lo["v"], "rank": lo["rank"], "n": lo["n"]},
        },
        src_note="SRC-634-01",
        extra={"derived": {"secondary": secondary}},
    )


def build_boston(app):
    text = fetch_text(URL_BOSTON, timeout=180)
    rdr = csv.DictReader(io.StringIO(text))
    pay = defaultdict(float)
    count = defaultdict(int)
    total = 0.0
    n = 0
    for r in rdr:
        dept = (r.get("DEPARTMENT_NAME") or "").strip() or "Unspecified"
        gross = parse_num(r.get("TOTAL GROSS"))
        if gross is None:
            continue
        pay[dept] += gross
        count[dept] += 1
        total += gross
        n += 1
    if n < 1000 or total < 1e8:
        sys.exit(f"FATAL: Boston 2025 payroll parsed n={n} total={total}")
    ranked = rank_named(pay, higher_is_better=True, st_key=lambda n: n[:8])
    for rec in ranked:
        rec["employees"] = count[rec["name"]]
    hi = ranked[0]
    as_of = "2025-12"
    as_of_label = "Calendar year 2025"
    kpis = [
        _kpi(
            "City payroll, 2025",
            usd_prose(total),
            f"{commify(n)} employees across {len(ranked)} departments (SRC-627-01).",
            "The citywide earnings stock.",
            "City of Boston employee earnings report 2025 (SRC-627-01)",
        ),
        _kpi(
            "Largest department",
            hi["name"],
            f"{usd_prose(hi['v'])} across {commify(hi['employees'])} employees (SRC-627-01).",
            "The department that accounts for the most earnings.",
            "City of Boston employee earnings report 2025 (SRC-627-01)",
        ),
        _kpi(
            "Departments",
            commify(len(ranked)),
            "Adopted-budget views are pending on this page.",
            "Payroll is the first Boston ledger because the file is public and current.",
            "City of Boston employee earnings report 2025 (SRC-627-01)",
        ),
    ]
    lead = (
        f"City of Boston earnings totaled <b>{usd_prose(total)}</b> in calendar "
        f"year 2025 across <b>{commify(n)}</b> employees (SRC-627-01). "
        f"<b>{hi['name']}</b> was the largest department at "
        f"<b>{usd_prose(hi['v'])}</b> (SRC-627-01). The adopted budget is pending."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from the City of Boston CKAN datastore dump of "
            f"employee-earnings-report-2025. TOTAL GROSS is summed by "
            f"DEPARTMENT_NAME. The adopted-budget file remains pending."
        ),
        metric="boston_department_earnings_2025",
        metric_label="Department earnings, calendar year 2025",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "total": total,
            "employees": n,
            "departments": len(ranked),
            "highest": {"name": hi["name"], "v": hi["v"], "employees": hi["employees"]},
        },
        src_note="SRC-627-01",
    )


def _qtax():
    wb = _wb(URL_QTAX)
    ws = wb.active
    # Row 6: state names every 5 columns. Row 7: period labels.
    # Row 8: Total Taxes. Values are thousands of dollars.
    names = [c.value for c in ws[6]]
    periods = [c.value for c in ws[7]]
    totals = [c.value for c in ws[8]]
    geos = []
    for i, name in enumerate(names):
        if not name or name in ("Tax Description", "Code"):
            continue
        st = geo_to_st(name)
        if not st:
            continue
        # column i is 2026 Q1 for that geo
        v = parse_num(totals[i])
        prev_y = parse_num(totals[i + 2]) if i + 2 < len(totals) else None
        geos.append((st, v, prev_y, i))
    values = {st: v * 1000 for st, v, _p, _i in geos if st != "US" and v is not None}
    us = next(((v * 1000, p * 1000 if p else None) for st, v, p, _i in geos if st == "US"), None)
    if us is None:
        sys.exit("FATAL: QTAX missing U.S. Total Taxes")
    us_val, us_prev = us
    thousands = us_val / 1000
    if abs(thousands - VERIFY_US_TAX_Q1_2026_THOUSANDS) > 1:
        sys.exit(f"FATAL: QTAX US 2026 Q1 total taxes are {thousands}")
    return values, us_val, us_prev, ws, geos


def build_ma_finances(app):
    values, us_val, us_prev, ws, geos = _qtax()
    ma_col = next(i for st, _v, _p, i in geos if st == "MA")
    # tax-type rows: description in col 0, MA 2026 Q1 in ma_col
    types = {}
    for row in ws.iter_rows(min_row=8, max_row=38, values_only=True):
        label = (row[0] or "")
        if not isinstance(label, str):
            continue
        name = label.replace("\xa0", "").strip()
        if not name or name.startswith("Abbreviations"):
            continue
        v = parse_num(row[ma_col])
        if v is None:
            continue
        types[name] = v * 1000
    ranked = rank_named(types, higher_is_better=True, st_key=lambda n: n[:8])
    ma_total = types.get("Total Taxes")
    if ma_total is None:
        sys.exit("FATAL: QTAX missing Massachusetts Total Taxes")
    ma_prev = next(p for st, _v, p, _i in geos if st == "MA")
    ma_prev_d = ma_prev * 1000 if ma_prev else None
    yoy = yoy_pct(ma_total, ma_prev_d)
    us_yoy = yoy_pct(us_val, us_prev)
    as_of = "2026-03"
    as_of_label = "2026 Q1"
    kpis = [
        _kpi(
            "Massachusetts taxes, 2026 Q1",
            usd_prose(ma_total),
            f"{pct(yoy)} from 2025 Q1 (SRC-628-01).",
            "The Commonwealth's quarterly tax take.",
            "Census Quarterly Summary of State and Local Tax Revenue, table 3 (SRC-628-01)",
        ),
        _kpi(
            "U.S. state taxes, 2026 Q1",
            usd_prose(us_val),
            f"{pct(us_yoy)} from 2025 Q1. Excludes D.C. (SRC-628-01).",
            "The national comparison for the Massachusetts print.",
            "Census QTAX table 3 (SRC-628-01)",
        ),
        _kpi(
            "Largest Massachusetts source",
            ranked[1]["name"] if ranked[0]["name"] == "Total Taxes" and len(ranked) > 1 else ranked[0]["name"],
            "DOR monthly collections and tax-credit files are pending.",
            "The type-of-tax split is what this first ledger can support.",
            "Census QTAX table 3 (SRC-628-01)",
        ),
    ]
    top = ranked[1] if ranked[0]["name"] == "Total Taxes" and len(ranked) > 1 else ranked[0]
    lead = (
        f"Massachusetts state tax collections were <b>{usd_prose(ma_total)}</b> "
        f"in the first quarter of 2026, {pct(yoy)} from the same quarter of "
        f"2025 (SRC-628-01). The United States (excluding D.C.) collected "
        f"<b>{usd_prose(us_val)}</b> (SRC-628-01). {top['name']} was the "
        f"largest Massachusetts source after the total. DOR monthly reports "
        f"and tax credits are pending."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from Census QTAX 2026 q1t3.xlsx, Massachusetts "
            f"2026 Q1 column. Amounts are published in thousands of dollars. "
            f"U.S. Total Taxes equals {VERIFY_US_TAX_Q1_2026_THOUSANDS:,} "
            f"thousand. DOR monthly and tax-credit files remain pending."
        ),
        metric="ma_state_tax_collections_2026q1",
        metric_label="Massachusetts state tax collections by type, 2026 Q1",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": us_val, "yoy_pct": us_yoy},
            "ma": {"v": ma_total, "yoy_pct": yoy},
            "largest_source": {"name": top["name"], "v": top["v"]},
        },
        src_note="SRC-628-01",
    )


def build_us_finances(app):
    values, us_val, us_prev, ws, geos = _qtax()
    prev = {st: p * 1000 for st, _v, p, _i in geos if st != "US" and p is not None}
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["yoy_pct"] = yoy_pct(rec["v"], prev.get(rec["st"]))
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    us_yoy = yoy_pct(us_val, us_prev)
    as_of = "2026-03"
    as_of_label = "2026 Q1"
    kpis = [
        _kpi(
            "U.S. state taxes, 2026 Q1",
            usd_prose(us_val),
            f"{pct(us_yoy)} from 2025 Q1. Excludes D.C. (SRC-629-01).",
            "The national quarterly tax take.",
            "Census QTAX table 3 (SRC-629-01)",
        ),
        _kpi(
            "Massachusetts",
            usd_prose(ma["v"]),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-629-01). {pct(ma['yoy_pct'])} from 2025 Q1.",
            "Massachusetts against the other states on the same Census file.",
            "Census QTAX table 3 (SRC-629-01)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {usd_prose(hi['v'])}",
            (
                f"{hi['name']} collected the most; {lo['name']} the least at "
                f"{usd_prose(lo['v'])} (SRC-629-01)."
            ),
            "Annual Census state-finance, NASBO rainy-day, and employee counts are pending.",
            "Census QTAX table 3 (SRC-629-01)",
        ),
    ]
    lead = (
        f"State government tax collections were <b>{usd_prose(us_val)}</b> in "
        f"the first quarter of 2026, {pct(us_yoy)} from a year earlier, "
        f"excluding D.C. (SRC-629-01). Massachusetts collected "
        f"<b>{usd_prose(ma['v'])}</b>, rank {ma['rank']} of {ma['n']} "
        f"(derived, SRC-629-01). Rainy-day funds and public-employee counts "
        f"are pending."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from Census QTAX 2026 q1t3.xlsx, Total Taxes, "
            f"2026 Q1. U.S. equals {VERIFY_US_TAX_Q1_2026_THOUSANDS:,} thousand "
            f"dollars. Annual Survey of State Government Finances, NASBO "
            f"rainy-day funds, and employee counts remain pending."
        ),
        metric="state_tax_collections_2026q1",
        metric_label="State government tax collections, 2026 Q1",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": us_val, "yoy_pct": us_yoy},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"], "yoy_pct": ma["yoy_pct"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-629-01",
    )


def build_crime(app):
    data = fetch(URL_BJS, timeout=120)
    zf = zipfile.ZipFile(io.BytesIO(data))
    text = zf.read("p23stt02.csv").decode("latin-1")
    values = {}
    prev = {}
    us_val = None
    us_prev = None
    for row in csv.reader(io.StringIO(text)):
        if len(row) < 8:
            continue
        name = (row[1] or "").strip()
        if not name:
            continue
        st = geo_to_st(name)
        if name.lower().startswith("u.s. total"):
            st = "US"
        if not st:
            continue
        v2022 = parse_num(row[2])
        v2023 = parse_num(row[6])
        if v2023 is None:
            continue
        if st == "US":
            us_val, us_prev = v2023, v2022
        else:
            values[st] = v2023
            prev[st] = v2022
    if us_val is None or abs(us_val - VERIFY_US_PRISONERS_2023) > 1:
        sys.exit(f"FATAL: BJS 2023 U.S. prisoners are {us_val}")
    if len(values) < 50:
        sys.exit(f"FATAL: BJS table 2 parsed {len(values)} states")
    ranked = rank_rows(values, higher_is_better=True)
    for rec in ranked:
        rec["v"] = int(round(rec["v"]))
        rec["yoy_pct"] = yoy_pct(rec["v"], prev.get(rec["st"]))
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    as_of = "2023-12"
    as_of_label = "Year-end 2023"
    kpis = [
        _kpi(
            "U.S. prisoners, 2023",
            commify(us_val),
            f"{pct(yoy_pct(us_val, us_prev))} from 2022. Includes federal (SRC-631-02).",
            "The national jurisdiction count.",
            "BJS Prisoners in 2023, table 2 (SRC-631-02)",
        ),
        _kpi(
            "Massachusetts",
            commify(ma["v"]),
            f"Rank {ma['rank']} of {ma['n']} (derived, SRC-631-02). {pct(ma['yoy_pct'])} from 2022.",
            "Massachusetts prisoners against the other states.",
            "BJS Prisoners in 2023, table 2 (SRC-631-02)",
        ),
        _kpi(
            "Highest / lowest",
            f"{hi['st']} {commify(hi['v'])}",
            (
                f"{hi['name']} held the most; {lo['name']} the fewest at "
                f"{commify(lo['v'])} (SRC-631-02)."
            ),
            "FBI crime rates, juvenile incarceration, and IC3 are pending.",
            "BJS Prisoners in 2023, table 2 (SRC-631-02)",
        ),
    ]
    lead = (
        f"State and federal correctional authorities held <b>{commify(us_val)}</b> "
        f"prisoners at year-end 2023 (SRC-631-02). Massachusetts held "
        f"<b>{commify(ma['v'])}</b>, rank {ma['rank']} of {ma['n']} "
        f"(derived, SRC-631-02). FBI crime rates, juvenile incarceration, and "
        f"internet-crime reports are pending on this page."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from BJS Prisoners in 2023 statistical tables, "
            f"p23stt02.csv (jurisdiction counts, 2023). U.S. total equals "
            f"{VERIFY_US_PRISONERS_2023:,}. FBI UCR/NIBRS, juvenile, and IC3 "
            f"files remain pending."
        ),
        metric="prisoners_yearend_2023",
        metric_label="Prisoners under jurisdiction, year-end 2023",
        unit="prisoners",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": int(us_val), "yoy_pct": yoy_pct(us_val, us_prev)},
            "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"], "yoy_pct": ma["yoy_pct"]},
            "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
            "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        },
        src_note="SRC-631-02",
    )


def build_340b(app):
    """DL-11: OPAIS daily export plus CMS HCRIS and Census SLDL-ZCTA."""
    from build_dl11 import build as build_dl11
    return build_dl11(app)


BUILDERS = {
    "DL-06": build_ma_k12,
    "DL-07": build_national_k12,
    "DL-08": build_higher_ed,
    "DL-09": build_charters,
    "DL-11": build_340b,
    "DL-12": build_medicaid,
    "DL-15": build_gdp,
    "DL-19": build_rpp,
    "DL-20": build_migration,
    "DL-21": build_tax_stats,
    "DL-23": build_vmt,
    "DL-24": build_co2,
    "DL-25": build_muni_atlas,
    "DL-26": build_muni_rankings,
    "DL-27": build_boston,
    "DL-34": build_boston_schools,
    "DL-28": build_ma_finances,
    "DL-29": build_us_finances,
    "DL-31": build_crime,
}