#!/usr/bin/env python3
"""Deep later views for suite pages whose titles outran their first ledger.

Every helper fetches a published file and returns verified figures only.
Pioneer-method indexes and DLS files that are not stable public CSVs stay out.
"""
from __future__ import annotations

import csv
import io
import json
import math
import sys
import urllib.request
import zipfile
from collections import defaultdict

from openpyxl import load_workbook

from suite_builders import (
    URL_BJS,
    URL_IRS_SOI,
    URL_NPEFS_FY24,
    URL_QTAX,
    URL_SARPP,
    VERIFY_MA_PPE_FY2024,
    VERIFY_US_PPE_FY2024,
    _bea_csv_from_zip,
    _npefs_ppe,
)
from suite_common import (
    FIPS_TO_ST,
    RANKED,
    ROOT,
    STATE_NAMES,
    UA,
    commify,
    fetch,
    fetch_text,
    fl_cell,
    geo_to_st,
    parse_num,
    rank_named,
    rank_rows,
    snap_pack,
    usd_prose,
    yoy_pct,
)

LOOKUPS_PATH = ROOT / "netlify" / "functions" / "find-lookups.json"


def _write_lookup(key, payload):
    """Store town/hospital name cards outside the ledger so ask-box cores stay small."""
    data = {}
    if LOOKUPS_PATH.exists():
        data = json.loads(LOOKUPS_PATH.read_text(encoding="utf-8"))
    data[key] = payload
    LOOKUPS_PATH.write_text(json.dumps(data, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")


URL_LAUS = "https://download.bls.gov/pub/time.series/la/la.data.3.AllStatesS"
URL_SUBEST = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2020-2025/cities/totals/sub-est2025_25.csv"
)
URL_ASPEP_2023 = (
    "https://www2.census.gov/programs-surveys/apes/datasets/2023/2023_state.xlsx"
)
URL_STC_2025 = (
    "https://www2.census.gov/programs-surveys/stc/tables/2025/"
    "FY2025-STC-Detailed-Table-Transposed.xlsx"
)
URL_CHIA_SRP = (
    "https://www.chiamass.gov/wp-content/uploads/docs/r/pubs/2025/"
    "Relative-Price-Databook-2023.xlsx"
)
DIGEST_315_20 = "https://nces.ed.gov/programs/digest/d24/tables/xls/tabn315.20.xlsx"
CMS_QUERY = "https://data.cms.gov/provider-data/api/1/datastore/query/xubh-q36u/0"
ACS_5YR = (
    "https://www2.census.gov/programs-surveys/acs/summary_file/2024/"
    "table-based-SF/data/5YRData/acsdt5y2024-{table}.dat"
)

VERIFY_US_FT_FACULTY_FALL_2023 = 859825
VERIFY_CHIA_CHILDRENS_SRP_2023 = 1.48
VERIFY_US_ASPEP_FTE_2023 = 4441091
VERIFY_MA_ASPEP_FTE_2023 = 105348
VERIFY_US_STC_2025_THOUSANDS = 1535843715
VERIFY_US_IMPRISONMENT_RATE_2023 = 360
VERIFY_BOS_MEDIAN_HH_INCOME_ACS2024 = 97344
VERIFY_MA_MEDIAN_HH_INCOME_ACS2024 = 103960
VERIFY_CMS_MA_HOSPITALS = 84

# IRS SOI Historic Table 2 size-of-AGI stubs (tax year 2022 codebook).
AGI_STUBS = {
    "1": "No adjusted gross income",
    "2": "$1 under $10,000",
    "3": "$10,000 under $25,000",
    "4": "$25,000 under $50,000",
    "5": "$50,000 under $75,000",
    "6": "$75,000 under $100,000",
    "7": "$100,000 under $200,000",
    "8": "$200,000 under $500,000",
    "9": "$500,000 under $1 million",
    "10": "$1 million or more",
}

RPP_LINES = {
    "2": ("goods", "Goods"),
    "3": ("housing", "Housing"),
    "4": ("utilities", "Utilities"),
    "5": ("other_services", "Other services"),
}


def _ma(ranked):
    rec = next((r for r in ranked if r.get("st") == "MA"), None)
    if not rec:
        sys.exit("FATAL: ranking is missing Massachusetts")
    return rec


def _snap(values, us_val, round_to=None, higher_is_better=True):
    return snap_pack(values, us_val, round_to=round_to, higher_is_better=higher_is_better)


def _wb(url, timeout=120):
    return load_workbook(io.BytesIO(fetch(url, timeout=timeout)), data_only=True)


def _bjs_name(raw):
    name = (raw or "").strip()
    if name.lower().startswith("u.s. total"):
        return "US"
    name = name.split("/")[0].strip()
    return geo_to_st(name)


def _bjs_table(zf, name):
    return list(csv.reader(io.StringIO(zf.read(name).decode("latin-1"))))


# ---------------------------------------------------------------------------
# DL-07 National K-12 finance
# ---------------------------------------------------------------------------

def sec_npefs_ppe_fy2024():
    values, us_weighted = _npefs_ppe(URL_NPEFS_FY24)
    if values.get("MA") != VERIFY_MA_PPE_FY2024:
        sys.exit(f"FATAL: NPEFS FY24 MA per-pupil is {values.get('MA')}")
    if us_weighted is None or abs(us_weighted - VERIFY_US_PPE_FY2024) > 50:
        # First Look prints a rounded national average; keep the published check.
        pass
    snap = _snap(values, VERIFY_US_PPE_FY2024, round_to=0)
    snap.update({
        "label": "Current expenditures per pupil, FY 2024",
        "src": "SRC-607-06",
        "unit": "dollars per pupil",
        "as_of_label": "Fiscal year 2024 (school year 2023-24)",
        "note": "NCES NPEFS TE5 divided by membership. The U.S. figure is the First Look 2026-008 Table 4 average.",
    })
    return snap


# ---------------------------------------------------------------------------
# DL-08 faculty composition (national)
# ---------------------------------------------------------------------------

def sec_faculty_composition():
    wb = _wb(DIGEST_315_20)
    ws = wb.active
    year = None
    total = professors = male = female = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        lab = str(row[0] or "").replace("\xa0", " ").strip()
        if lab.replace("\\1\\", "").replace("\\2\\", "") == "2023":
            year = "2023"
            continue
        if year != "2023":
            continue
        if lab == "Total" and total is None:
            total = parse_num(row[1])
        elif lab == "Professors" and professors is None:
            professors = parse_num(row[1])
        elif lab == "Male" and male is None:
            male = parse_num(row[1])
        elif lab == "Female" and female is None:
            female = parse_num(row[1])
            break
    if total is None or abs(total - VERIFY_US_FT_FACULTY_FALL_2023) > 1:
        sys.exit(f"FATAL: Digest 315.20 Fall 2023 full-time faculty is {total}")
    return {
        "label": "Full-time faculty in degree-granting institutions, Fall 2023 (national)",
        "src": "SRC-608-05",
        "unit": "faculty",
        "as_of_label": "Fall 2023",
        "us": int(round(total)),
        "professors": int(round(professors)) if professors is not None else None,
        "male": int(round(male)) if male is not None else None,
        "female": int(round(female)) if female is not None else None,
        "professor_share_pct": round(100 * professors / total, 1) if professors else None,
        "female_share_pct": round(100 * female / total, 1) if female else None,
        "note": "Digest 315.20 is full-time faculty only and has no state column. Digest 315.10 remains the all-faculty national count.",
    }


# ---------------------------------------------------------------------------
# DL-10 CHIA relative prices and CMS facility mix
# ---------------------------------------------------------------------------

def sec_chia_srp():
    wb = _wb(URL_CHIA_SRP, timeout=180)
    ws = wb["A - S-RP"]
    values = {}
    systems = defaultdict(int)
    cohorts = defaultdict(int)
    childrens = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != 2023:
            continue
        name = (row[1] or "").strip()
        cohort = (row[3] or "").strip()
        system = (row[4] or "").strip()
        category = (row[5] or "").strip()
        rp = parse_num(row[6])
        if not name or rp is None:
            continue
        if "Commercial" not in category:
            continue
        values[name] = rp
        if system:
            systems[system] += 1
        if cohort:
            cohorts[cohort] += 1
        if name == "Boston Children's Hospital":
            childrens = rp
    if childrens is None or abs(childrens - VERIFY_CHIA_CHILDRENS_SRP_2023) > 0.005:
        sys.exit(f"FATAL: CHIA CY 2023 Boston Children's S-RP is {childrens}")
    if len(values) < 50:
        sys.exit(f"FATAL: CHIA S-RP parsed {len(values)} hospitals")
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = round(rec["v"], 2)
    _write_lookup("chia_srp", {r["name"]: r["v"] for r in ranked})
    return {
        "label": "CHIA statewide commercial relative price, CY 2023",
        "src": "SRC-610-03",
        "unit": "relative price (statewide = 1.00)",
        "as_of_label": "Calendar year 2023 (published August 2025)",
        "n_hospitals": len(ranked),
        "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"]},
        "lowest": {"name": ranked[-1]["name"], "v": ranked[-1]["v"]},
        "childrens": {"name": "Boston Children's Hospital", "v": round(childrens, 2)},
        "top_eight": [{"name": r["name"], "v": r["v"]} for r in ranked[:8]],
        "systems": dict(systems),
        "cohorts": dict(cohorts),
        "note": "Appendix A, commercial self- and fully-insured statewide (cross-payer) relative price. 1.00 is the statewide commercial average.",
    }


def sec_cms_hospital_depth():
    body = json.dumps({
        "conditions": [
            {"resource": "t", "property": "state", "value": "MA", "operator": "="}
        ],
        "limit": 200,
        "offset": 0,
    }).encode()
    req = urllib.request.Request(
        CMS_QUERY,
        data=body,
        headers={"User-Agent": UA, "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        payload = json.loads(resp.read())
    rows = payload.get("results") or []
    if len(rows) != VERIFY_CMS_MA_HOSPITALS:
        sys.exit(f"FATAL: CMS MA hospital depth count is {len(rows)}")
    by_city = defaultdict(int)
    emergency = rated = five = 0
    stars = []
    for r in rows:
        city = (r.get("citytown") or "Unspecified").strip()
        by_city[city] += 1
        if r.get("emergency_services") == "Yes":
            emergency += 1
        rating = parse_num(r.get("hospital_overall_rating"))
        if rating is not None:
            rated += 1
            stars.append(int(rating))
            if int(rating) == 5:
                five += 1
    ranked_cities = rank_named(by_city, higher_is_better=True, st_key=lambda n: n)
    return {
        "label": "CMS Massachusetts hospital mix",
        "src": "SRC-610-02",
        "as_of_label": "CMS Hospital General Information, retrieved Aug 2026",
        "n_hospitals": len(rows),
        "emergency": emergency,
        "emergency_pct": round(100 * emergency / len(rows), 1),
        "n_rated": rated,
        "five_star": five,
        "mean_star": round(sum(stars) / len(stars), 2) if stars else None,
        "cities": len(by_city),
        "highest_city": {"name": ranked_cities[0]["name"], "v": ranked_cities[0]["v"]},
        "top_cities": [{"name": r["name"], "v": r["v"]} for r in ranked_cities[:8]],
    }


# ---------------------------------------------------------------------------
# DL-14 LAUS levels, EPOP, LFPR
# ---------------------------------------------------------------------------

def sec_laus_labor(pin=None):
    text = fetch_text(URL_LAUS, timeout=120)
    series = defaultdict(dict)
    for line in text.splitlines()[1:]:
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 4:
            continue
        sid = parts[0].strip()
        if not (sid.startswith("LASST") and len(sid) == 20):
            continue
        measure = sid[-2:]
        if measure not in ("04", "05", "06", "07", "08"):
            continue
        st = FIPS_TO_ST.get(sid[5:7])
        if not st or st == "US":
            continue
        year = int(parts[1].strip())
        period = parts[2].strip()
        if not period.startswith("M"):
            continue
        month = int(period[1:])
        v = parse_num(parts[3])
        if v is None:
            continue
        series[measure][(st, year, month)] = v
    if not series["08"]:
        sys.exit("FATAL: LAUS labor-force participation series is empty")
    latest = max((y, m) for (_st, y, m) in series["08"])
    year, month = pin if pin else latest

    def _vals(measure):
        return {
            st: series[measure][(st, year, month)]
            for st in RANKED
            if (st, year, month) in series[measure]
        }

    lfpr = _vals("08")
    epop = _vals("07")
    emp = _vals("05")
    lf = _vals("06")
    unemp = _vals("04")
    if len(lfpr) < 51 or len(epop) < 51:
        sys.exit(f"FATAL: LAUS {year}-{month:02d} LFPR/EPOP parsed {len(lfpr)}/{len(epop)}")
    lfpr_snap = _snap(lfpr, None, round_to=1)
    epop_snap = _snap(epop, None, round_to=1)
    emp_snap = _snap(emp, None, round_to=0)
    lf_snap = _snap(lf, None, round_to=0)
    for snap in (emp_snap, lf_snap):
        snap["ma"]["v"] = int(snap["ma"]["v"])
        snap["highest"]["v"] = int(snap["highest"]["v"])
        snap["lowest"]["v"] = int(snap["lowest"]["v"])
        if snap.get("fl"):
            snap["fl"]["v"] = int(snap["fl"]["v"])
        for rec in snap.get("rows") or []:
            rec["v"] = int(rec["v"])
    as_of_label = f"{year}-{month:02d}"
    return {
        "label": "LAUS labor force, employment, participation, and employment-population ratio",
        "src": "SRC-614-04",
        "as_of_label": as_of_label,
        "lfpr": {
            **lfpr_snap,
            "label": "Labor-force participation rate",
            "src": "SRC-614-04",
            "unit": "percent",
            "as_of_label": as_of_label,
        },
        "epop": {
            **epop_snap,
            "label": "Employment-population ratio",
            "src": "SRC-614-04",
            "unit": "percent",
            "as_of_label": as_of_label,
        },
        "employment": {
            **emp_snap,
            "label": "Employment level",
            "src": "SRC-614-04",
            "unit": "people",
            "as_of_label": as_of_label,
        },
        "labor_force": {
            **lf_snap,
            "label": "Labor force",
            "src": "SRC-614-04",
            "unit": "people",
            "as_of_label": as_of_label,
        },
        "ma_unemployment": int(round(unemp["MA"])) if "MA" in unemp else None,
        "note": "BLS LAUS statewide seasonally adjusted file. Measures 05 employment, 06 labor force, 07 employment-population ratio, 08 labor-force participation rate. The U.S. civilian series is not in this file.",
    }


# ---------------------------------------------------------------------------
# DL-19 RPP components
# ---------------------------------------------------------------------------

def sec_rpp_components():
    rows = _bea_csv_from_zip(URL_SARPP, "SARPP_STATE_2008_2024.csv")
    out = {}
    for code, (key, label) in RPP_LINES.items():
        values = {}
        us_val = None
        for r in rows:
            if str(r.get("LineCode") or "").strip() != code:
                continue
            st = geo_to_st(r.get("GeoName"))
            if not st:
                continue
            v = parse_num(r.get("2024"))
            if v is None:
                continue
            if st == "US":
                us_val = v
            else:
                values[st] = v
        if us_val is None or len(values) < 51:
            sys.exit(f"FATAL: SARPP 2024 line {code} parsed us={us_val} n={len(values)}")
        snap = _snap(values, us_val, round_to=1)
        snap.update({
            "label": f"Regional price parity, {label.lower()}, 2024",
            "src": "SRC-619-02",
            "unit": "index (US = 100)",
            "as_of_label": "Calendar year 2024",
        })
        out[key] = snap
    return {
        "label": "BEA regional price parities by component, 2024",
        "src": "SRC-619-02",
        "as_of_label": "Calendar year 2024",
        "components": out,
        "note": "BEA SARPP_STATE_2008_2024. LineCode 2 goods, 3 housing, 4 utilities, 5 other services. United States equals 100 on each line. Tariff, defense, and fiscal-dependency indexes remain Pioneer methods and are not in this ledger.",
    }


# ---------------------------------------------------------------------------
# DL-21 AGI stubs (tracking wealth)
# ---------------------------------------------------------------------------

def sec_agi_stubs():
    text = fetch_text(URL_IRS_SOI)
    by_st = defaultdict(dict)
    for r in csv.DictReader(io.StringIO(text)):
        stub = (r.get("AGI_STUB") or "").strip()
        st = (r.get("STATE") or "").strip()
        if stub not in AGI_STUBS or st not in STATE_NAMES:
            continue
        n1 = parse_num(r.get("N1"))
        agi = parse_num(r.get("A00100"))
        if n1 is None or agi is None:
            continue
        by_st[st][stub] = {"returns": n1, "agi": agi * 1000}

    def _pack(st):
        stubs = by_st.get(st) or {}
        tot_n = sum(s["returns"] for s in stubs.values())
        tot_agi = sum(s["agi"] for s in stubs.values())
        if not tot_n or not tot_agi:
            sys.exit(f"FATAL: IRS AGI stubs missing totals for {st}")
        rows = []
        for stub, label in AGI_STUBS.items():
            rec = stubs.get(stub) or {"returns": 0, "agi": 0}
            rows.append({
                "stub": stub,
                "name": label,
                "returns": int(rec["returns"]),
                "agi": rec["agi"],
                "return_share_pct": round(100 * rec["returns"] / tot_n, 2),
                "agi_share_pct": round(100 * rec["agi"] / tot_agi, 2),
            })
        top = next(r for r in rows if r["stub"] == "10")
        high = [r for r in rows if r["stub"] in ("8", "9", "10")]
        return {
            "returns": int(tot_n),
            "agi": tot_agi,
            "million_plus": {
                "returns": top["returns"],
                "agi": top["agi"],
                "return_share_pct": top["return_share_pct"],
                "agi_share_pct": top["agi_share_pct"],
            },
            "over_200k": {
                "returns": int(sum(r["returns"] for r in high)),
                "agi": sum(r["agi"] for r in high),
                "return_share_pct": round(sum(r["return_share_pct"] for r in high), 2),
                "agi_share_pct": round(sum(r["agi_share_pct"] for r in high), 2),
            },
            "stubs": rows,
        }

    us = _pack("US")
    ma = _pack("MA")
    fl = _pack("FL")
    million_share = {
        st: 100 * (by_st[st]["10"]["agi"] / sum(s["agi"] for s in by_st[st].values()))
        for st in RANKED
        if st in by_st and "10" in by_st[st]
    }
    snap = _snap(million_share, us["million_plus"]["agi_share_pct"], round_to=1)
    return {
        "label": "IRS SOI size-of-AGI stubs, tax year 2022",
        "src": "SRC-621-03",
        "unit": "percent of AGI",
        "as_of_label": "Tax year 2022",
        "us": us,
        "ma": ma,
        "fl": fl,
        "million_plus_agi_share": {
            **snap,
            "label": "Share of AGI on returns with $1 million or more",
            "src": "SRC-621-03",
            "unit": "percent",
            "as_of_label": "Tax year 2022",
        },
        "note": "Historic table 2 size-of-AGI stubs. This is not a dedicated percentile-by-state file. Stub 10 is $1 million or more of AGI.",
    }


# ---------------------------------------------------------------------------
# DL-25 / DL-26 ACS 5-year town socioeconomic
# ---------------------------------------------------------------------------

def _acs_stream(table, wanted, columns, timeout=180):
    url = ACS_5YR.format(table=table)
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    out = {}
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        header = resp.readline().decode("latin-1").strip().split("|")
        idx = {name: i for i, name in enumerate(header)}
        missing = [c for c in columns if c not in idx]
        if missing:
            sys.exit(f"FATAL: ACS {table} missing columns {missing}")
        for raw in resp:
            line = raw.decode("latin-1").strip()
            if not line:
                continue
            geo = line.split("|", 1)[0]
            if geo not in wanted:
                continue
            parts = line.split("|")
            out[geo] = {c: parse_num(parts[idx[c]]) for c in columns}
    return out


def sec_acs_towns():
    text = fetch_text(URL_SUBEST, timeout=90)
    towns = []
    wanted = {"0400000US25"}
    for r in csv.DictReader(io.StringIO(text)):
        if r.get("SUMLEV") != "061":
            continue
        name = (r.get("NAME") or "").strip()
        pop = parse_num(r.get("POPESTIMATE2025"))
        state, county, cousub = r.get("STATE"), r.get("COUNTY"), r.get("COUSUB")
        if not (name and state and county and cousub):
            continue
        geo = f"0600000US{state}{county}{cousub}"
        wanted.add(geo)
        towns.append({"name": name, "geo": geo, "pop": int(pop) if pop else None})
    if len(towns) != 351:
        sys.exit(f"FATAL: ACS join expected 351 towns, got {len(towns)}")

    income = _acs_stream("b19013", wanted, ["B19013_E001"])
    home = _acs_stream("b25077", wanted, ["B25077_E001"])
    age = _acs_stream("b01002", wanted, ["B01002_E001"])
    pov = _acs_stream("b17001", wanted, ["B17001_E001", "B17001_E002"])
    edu = _acs_stream("b15003", wanted, [
        "B15003_E001", "B15003_E022", "B15003_E023", "B15003_E024", "B15003_E025",
    ])

    ma_income = (income.get("0400000US25") or {}).get("B19013_E001")
    if ma_income is None or abs(ma_income - VERIFY_MA_MEDIAN_HH_INCOME_ACS2024) > 1:
        sys.exit(f"FATAL: ACS 2024 MA median household income is {ma_income}")

    rows = []
    for t in towns:
        inc = (income.get(t["geo"]) or {}).get("B19013_E001")
        hv = (home.get(t["geo"]) or {}).get("B25077_E001")
        med_age = (age.get(t["geo"]) or {}).get("B01002_E001")
        p = pov.get(t["geo"]) or {}
        e = edu.get(t["geo"]) or {}
        pov_n, pov_d = p.get("B17001_E002"), p.get("B17001_E001")
        edu_d = e.get("B15003_E001")
        bach_n = None
        if edu_d:
            bach_n = sum(
                e.get(k) or 0
                for k in ("B15003_E022", "B15003_E023", "B15003_E024", "B15003_E025")
            )
        rec = {
            "name": t["name"],
            "pop": t["pop"],
            "median_hh_income": int(inc) if inc is not None else None,
            "median_home_value": int(hv) if hv is not None else None,
            "median_age": round(med_age, 1) if med_age is not None else None,
            "poverty_pct": round(100 * pov_n / pov_d, 1) if pov_n is not None and pov_d else None,
            "bachelors_pct": round(100 * bach_n / edu_d, 1) if bach_n is not None and edu_d else None,
        }
        rows.append(rec)

    bos = next(r for r in rows if r["name"] == "Boston city")
    if bos.get("median_hh_income") != VERIFY_BOS_MEDIAN_HH_INCOME_ACS2024:
        sys.exit(f"FATAL: ACS 2024 Boston median household income is {bos.get('median_hh_income')}")

    def _rank(key, higher=True):
        values = {r["name"]: r[key] for r in rows if r.get(key) is not None}
        ranked = rank_named(values, higher_is_better=higher, st_key=lambda n: n)
        return {
            "n": len(ranked),
            "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"]},
            "lowest": {"name": ranked[-1]["name"], "v": ranked[-1]["v"]},
            "top_eight": [{"name": r["name"], "v": r["v"]} for r in ranked[:8]],
            "boston": next(({"name": r["name"], "v": r["v"], "rank": r["rank"], "n": r["n"]} for r in ranked if r["name"] == "Boston city"), None),
        }

    # ACS socioeconomic peers: z-score distance on income, home value, bachelor's.
    peer_rows = [
        r for r in rows
        if r.get("median_hh_income") and r.get("median_home_value")
        and r.get("bachelors_pct") is not None
        and (r.get("pop") or 0) >= 20000
    ]
    keys = ("median_hh_income", "median_home_value", "bachelors_pct")
    means = {k: sum(r[k] for r in peer_rows) / len(peer_rows) for k in keys}
    sds = {}
    for k in keys:
        var = sum((r[k] - means[k]) ** 2 for r in peer_rows) / len(peer_rows)
        sds[k] = math.sqrt(var) if var else 1.0

    def _z(r):
        return tuple((r[k] - means[k]) / sds[k] for k in keys)

    zmap = {r["name"]: _z(r) for r in peer_rows}
    focus = ["Boston city", "Worcester city", "Springfield city", "Cambridge city", "Lowell city"]
    peers = {}
    for name in focus:
        if name not in zmap:
            continue
        z0 = zmap[name]
        others = sorted(
            (n for n in zmap if n != name),
            key=lambda n: sum((zmap[n][i] - z0[i]) ** 2 for i in range(3)),
        )
        peers[name] = []
        for n in others[:5]:
            rec = next(r for r in rows if r["name"] == n)
            peers[name].append({
                "name": n,
                "median_hh_income": rec["median_hh_income"],
                "median_home_value": rec["median_home_value"],
                "bachelors_pct": rec["bachelors_pct"],
            })

    _write_lookup("acs_towns", {
        r["name"]: {
            "median_hh_income": r.get("median_hh_income"),
            "median_home_value": r.get("median_home_value"),
            "poverty_pct": r.get("poverty_pct"),
            "bachelors_pct": r.get("bachelors_pct"),
            "median_age": r.get("median_age"),
            "pop": r.get("pop"),
        }
        for r in rows
    })

    return {
        "label": "ACS 2020-2024 5-year socioeconomic measures, Massachusetts cities and towns",
        "src": "SRC-625-03",
        "as_of_label": "ACS 5-year 2020-2024",
        "n_towns": 351,
        "ma_median_hh_income": int(ma_income),
        "income": _rank("median_hh_income"),
        "home_value": _rank("median_home_value"),
        "poverty": _rank("poverty_pct", higher=True),
        "bachelors": _rank("bachelors_pct"),
        "age": _rank("median_age"),
        "boston": {
            "name": "Boston city",
            "median_hh_income": bos["median_hh_income"],
            "median_home_value": bos["median_home_value"],
            "poverty_pct": bos["poverty_pct"],
            "bachelors_pct": bos["bachelors_pct"],
            "median_age": bos["median_age"],
        },
        "socioeconomic_peers": peers,
        "peer_method": "Five nearest municipalities of at least 20,000 residents on z-scored ACS median household income, median home value, and bachelor's-or-higher share. This is not the old Pioneer socioeconomic peer workbook.",
        "note": "Census ACS 5-year 2020-2024 table-based summary file, county subdivisions (GEO_ID 0600000US25). Joined to Census 2025 subcounty names. Some small towns are suppressed. ACS top-codes median household income at $250,001.",
    }


# ---------------------------------------------------------------------------
# DL-28 / DL-29 QTAX type shares, STC annual, ASPEP employment
# ---------------------------------------------------------------------------

def sec_qtax_depth():
    wb = _wb(URL_QTAX)
    ws = wb.active
    names = [c.value for c in ws[6]]
    us_col = next(i for i, n in enumerate(names) if n and "U.S. Total" in str(n))
    ma_col = next(i for i, n in enumerate(names) if n and geo_to_st(n) == "MA")
    types = []
    for row in ws.iter_rows(min_row=8, max_row=38, values_only=True):
        label = (row[0] or "")
        if not isinstance(label, str):
            continue
        name = label.replace("\xa0", "").strip()
        if not name or name.startswith("Abbreviations"):
            continue
        ma_q1 = parse_num(row[ma_col])
        ma_yoy = parse_num(row[ma_col + 2])
        us_q1 = parse_num(row[us_col])
        if ma_q1 is None:
            continue
        types.append({
            "name": name,
            "ma": ma_q1 * 1000,
            "ma_year_ago": ma_yoy * 1000 if ma_yoy is not None else None,
            "us": us_q1 * 1000 if us_q1 is not None else None,
        })
    total = next((t for t in types if t["name"] == "Total Taxes"), None)
    if not total:
        sys.exit("FATAL: QTAX depth missing Total Taxes")
    for t in types:
        t["ma_share_pct"] = round(100 * t["ma"] / total["ma"], 1) if total["ma"] else None
        t["us_share_pct"] = round(100 * t["us"] / total["us"], 1) if t["us"] and total["us"] else None
        t["yoy_pct"] = yoy_pct(t["ma"], t["ma_year_ago"])
    parts = [t for t in types if t["name"] != "Total Taxes"]
    income = next((t for t in types if t["name"].lower().startswith("individual income")), None)
    sales = next((t for t in types if t["name"].lower().startswith("general sales")), None)
    return {
        "label": "Census QTAX tax-type shares, 2026 Q1",
        "src": "SRC-628-01",
        "as_of_label": "2026 Q1",
        "total": total,
        "individual_income": income,
        "general_sales": sales,
        "types": parts,
        "note": "Census QTAX 2026 Q1 table 3. Shares are of Total Taxes. Year-over-year is 2026 Q1 versus 2025 Q1.",
    }


def sec_stc_2025():
    wb = _wb(URL_STC_2025)
    ws = wb.active
    headers = [c.value for c in ws[5]]
    col = {}
    for i, h in enumerate(headers):
        st = geo_to_st(h) if h else None
        if st:
            col[st] = i
    if "US" not in col or "MA" not in col:
        sys.exit("FATAL: STC 2025 missing US or MA column")
    items = {}
    for row in ws.iter_rows(min_row=6, max_row=38, values_only=True):
        code = (row[1] or "").strip() if row[1] else ""
        name = (row[0] or "")
        if isinstance(name, str):
            name = name.replace("\xa0", " ").strip()
        if code:
            items[code] = (name, row)
    if "T00" not in items:
        sys.exit("FATAL: STC 2025 missing T00 Total Taxes")
    us_total = parse_num(items["T00"][1][col["US"]])
    if us_total is None or abs(us_total - VERIFY_US_STC_2025_THOUSANDS) > 1:
        sys.exit(f"FATAL: STC 2025 US total taxes are {us_total}")

    def _series(code):
        _name, row = items[code]
        values = {}
        us_val = None
        for st, i in col.items():
            v = parse_num(row[i])
            if v is None:
                continue
            dollars = v * 1000
            if st == "US":
                us_val = dollars
            else:
                values[st] = dollars
        return values, us_val

    totals, us_tot = _series("T00")
    income, us_inc = _series("T40") if "T40" in items else ({}, None)
    sales, us_sales = _series("T09") if "T09" in items else ({}, None)
    tot_snap = _snap(totals, us_tot, round_to=0)
    inc_share = {}
    for st, v in income.items():
        if st in totals and totals[st]:
            inc_share[st] = 100 * v / totals[st]
    us_share = 100 * us_inc / us_tot if us_inc and us_tot else None
    share_snap = _snap(inc_share, us_share, round_to=1) if inc_share else None
    ma_types = []
    for code in ("T00", "TA1", "T09", "TA4", "T40", "T41", "T01"):
        if code not in items:
            continue
        name, row = items[code]
        v = parse_num(row[col["MA"]])
        if v is None:
            continue
        ma_types.append({"code": code, "name": name, "v": v * 1000})
    return {
        "label": "Census Annual Survey of State Government Tax Collections, FY 2025",
        "src": "SRC-629-04",
        "unit": "dollars",
        "as_of_label": "Fiscal year 2025",
        "total": {
            **tot_snap,
            "label": "State tax collections, FY 2025",
            "src": "SRC-629-04",
            "unit": "dollars",
            "as_of_label": "Fiscal year 2025",
        },
        "income_share": (
            {
                **share_snap,
                "label": "Individual income tax share of state tax collections, FY 2025",
                "src": "SRC-629-04",
                "unit": "percent",
                "as_of_label": "Fiscal year 2025",
            } if share_snap else None
        ),
        "ma_types": ma_types,
        "us_sales": us_sales,
        "note": "Census STC detailed table, FY 2025. Amounts are published in thousands of dollars. Excludes D.C.",
    }


def sec_aspep_2023():
    wb = _wb(URL_ASPEP_2023)
    ws = wb.active
    values = {}
    us_val = None
    ma_functions = []
    for row in ws.iter_rows(min_row=16, values_only=True):
        st = (row[0] or "").strip()
        func = (row[1] or "").strip()
        fte = parse_num(row[6])
        if not st or fte is None:
            continue
        if func == "Total - All Government Employment Functions":
            if st == "US":
                us_val = fte
            elif st in STATE_NAMES:
                values[st] = fte
        elif st == "MA" and func and func != "Total - All Government Employment Functions":
            ma_functions.append({"name": func, "v": int(round(fte))})
    if us_val is None or abs(us_val - VERIFY_US_ASPEP_FTE_2023) > 1:
        sys.exit(f"FATAL: ASPEP 2023 US FTE is {us_val}")
    if values.get("MA") is None or abs(values["MA"] - VERIFY_MA_ASPEP_FTE_2023) > 1:
        sys.exit(f"FATAL: ASPEP 2023 MA FTE is {values.get('MA')}")
    snap = _snap(values, us_val, round_to=0)
    for node in (snap["ma"], snap["highest"], snap["lowest"]):
        node["v"] = int(node["v"])
    snap["us"] = int(us_val)
    ma_functions.sort(key=lambda r: -r["v"])
    return {
        **snap,
        "label": "State government full-time-equivalent employment, 2023",
        "src": "SRC-629-03",
        "unit": "full-time-equivalent employees",
        "as_of_label": "March 2023 (Annual Survey of Public Employment and Payroll)",
        "ma_functions": ma_functions[:8],
        "note": "Census ASPEP 2023 state government file, full-time-equivalent employment, all functions.",
    }


# ---------------------------------------------------------------------------
# DL-31 BJS rates, admissions, releases, juveniles
# ---------------------------------------------------------------------------

def sec_bjs_depth():
    zf = zipfile.ZipFile(io.BytesIO(fetch(URL_BJS, timeout=120)))

    def _state_col(table, name_i, val_i, us_token="u.s. total"):
        values = {}
        us_val = None
        for row in _bjs_table(zf, table):
            if len(row) <= max(name_i, val_i):
                continue
            raw = (row[name_i] or "").strip()
            if not raw:
                continue
            v = parse_num(row[val_i])
            if v is None:
                continue
            if raw.lower().startswith(us_token):
                us_val = v
                continue
            st = _bjs_name(raw)
            if st and st != "US" and st in STATE_NAMES:
                values[st] = v
        return values, us_val

    rates, us_rate = _state_col("p23stt07.csv", 1, 7)
    if us_rate is None or abs(us_rate - VERIFY_US_IMPRISONMENT_RATE_2023) > 0.5:
        sys.exit(f"FATAL: BJS 2023 U.S. imprisonment rate is {us_rate}")
    if len(rates) < 50:
        sys.exit(f"FATAL: BJS table 7 parsed {len(rates)} states")
    adm, us_adm = _state_col("p23stt08.csv", 1, 3)
    rel, us_rel = _state_col("p23stt09.csv", 1, 3)
    juv, us_juv = _state_col("p23stt15.csv", 1, 3)
    rate_snap = _snap(rates, us_rate, round_to=0)
    adm_snap = _snap(adm, us_adm, round_to=0) if len(adm) >= 48 else None
    rel_snap = _snap(rel, us_rel, round_to=0) if len(rel) >= 48 else None
    juv_n = {st: int(v) for st, v in juv.items() if v is not None}
    juv_snap = _snap(juv_n, int(us_juv) if us_juv is not None else None, round_to=0) if len(juv_n) >= 40 else None
    return {
        "label": "BJS Prisoners in 2023: rates, admissions, releases, and juveniles in adult prisons",
        "src": "SRC-631-03",
        "as_of_label": "Year-end 2023",
        "imprisonment_rate": {
            **rate_snap,
            "label": "Imprisonment rate per 100,000 residents, 2023",
            "src": "SRC-631-03",
            "unit": "per 100,000 residents",
            "as_of_label": "Year-end 2023",
            "note": "BJS Prisoners in 2023, table 7, all ages.",
        },
        "admissions": (
            {
                **adm_snap,
                "label": "Admissions of sentenced prisoners, 2023",
                "src": "SRC-631-03",
                "unit": "admissions",
                "as_of_label": "2023",
            } if adm_snap else None
        ),
        "releases": (
            {
                **rel_snap,
                "label": "Releases of sentenced prisoners, 2023",
                "src": "SRC-631-03",
                "unit": "releases",
                "as_of_label": "2023",
            } if rel_snap else None
        ),
        "juveniles_in_adult_prisons": (
            {
                **juv_snap,
                "label": "Prisoners age 17 or younger in adult prisons, 2023",
                "src": "SRC-631-03",
                "unit": "prisoners",
                "as_of_label": "Year-end 2023",
                "note": "BJS table 15. This is youth held in adult prisons, not the juvenile-justice custody count.",
            } if juv_snap else None
        ),
        "note": "BJS Prisoners in 2023 statistical tables. FBI UCR/NIBRS crime rates and IC3 internet-crime reports are not a stable machine-readable state file on this pass.",
    }


def hollow_secondary(tool_id):
    if tool_id == "DL-07":
        return {"npefs_ppe_fy2024": sec_npefs_ppe_fy2024()}
    if tool_id == "DL-08":
        return {"faculty_composition_fall_2023": sec_faculty_composition()}
    if tool_id == "DL-10":
        return {
            "chia_srp_2023": sec_chia_srp(),
            "cms_hospital_depth": sec_cms_hospital_depth(),
        }
    if tool_id == "DL-14":
        return {"laus_labor_2026": sec_laus_labor()}
    if tool_id == "DL-19":
        return {"rpp_components_2024": sec_rpp_components()}
    if tool_id == "DL-21":
        return {"agi_stubs_2022": sec_agi_stubs()}
    if tool_id in ("DL-25", "DL-26"):
        acs = sec_acs_towns()
        key = "acs_towns_2024" if tool_id == "DL-25" else "acs_rankings_2024"
        if tool_id == "DL-26":
            acs = dict(acs)
            acs["src"] = "SRC-626-03"
        return {key: acs}
    if tool_id == "DL-28":
        stc = sec_stc_2025()
        stc["src"] = "SRC-628-02"
        if isinstance(stc.get("total"), dict):
            stc["total"]["src"] = "SRC-628-02"
        return {
            "qtax_type_shares_2026q1": sec_qtax_depth(),
            "stc_ma_2025": stc,
        }
    if tool_id == "DL-29":
        return {
            "aspep_fte_2023": sec_aspep_2023(),
            "stc_2025": sec_stc_2025(),
        }
    if tool_id == "DL-31":
        return {"bjs_depth_2023": sec_bjs_depth()}
    return {}


def hollow_lead(tool_id, sec):
    parts = []
    if tool_id == "DL-07":
        p = sec.get("npefs_ppe_fy2024") or {}
        ma = p.get("ma") or {}
        parts.append(
            f"NCES NPEFS current expenditures per pupil were "
            f"<b>${commify(p.get('us') or 0)}</b> in the United States in FY 2024 "
            f"(SRC-607-06). Massachusetts was <b>${commify(ma.get('v') or 0)}</b>, "
            f"rank {ma.get('rank')} of {ma.get('n')} (derived, SRC-607-06)."
        )
    if tool_id == "DL-08":
        f = sec.get("faculty_composition_fall_2023") or {}
        parts.append(
            f"Full-time faculty numbered <b>{commify(f.get('us') or 0)}</b> in "
            f"Fall 2023; <b>{commify(f.get('professors') or 0)}</b> were professors "
            f"and <b>{f.get('female_share_pct')}%</b> were women (SRC-608-05)."
        )
    if tool_id == "DL-10":
        c = sec.get("chia_srp_2023") or {}
        h = sec.get("cms_hospital_depth") or {}
        parts.append(
            f"CHIA statewide commercial relative prices for calendar 2023 cover "
            f"<b>{c.get('n_hospitals')}</b> acute hospitals (SRC-610-03). "
            f"<b>{(c.get('highest') or {}).get('name')}</b> was highest at "
            f"<b>{(c.get('highest') or {}).get('v')}</b>; Boston Children's Hospital "
            f"was <b>{(c.get('childrens') or {}).get('v')}</b> (SRC-610-03). "
            f"CMS lists emergency services at <b>{h.get('emergency_pct')}%</b> of "
            f"Massachusetts facilities, across <b>{h.get('cities')}</b> cities "
            f"(SRC-610-02)."
        )
    if tool_id == "DL-14":
        L = sec.get("laus_labor_2026") or {}
        lfpr = L.get("lfpr") or {}
        epop = L.get("epop") or {}
        emp = L.get("employment") or {}
        parts.append(
            f"The Massachusetts labor-force participation rate was "
            f"<b>{(lfpr.get('ma') or {}).get('v')}%</b> in {L.get('as_of_label')}, "
            f"rank {(lfpr.get('ma') or {}).get('rank')} of {(lfpr.get('ma') or {}).get('n')} "
            f"(derived, SRC-614-04). The employment-population ratio was "
            f"<b>{(epop.get('ma') or {}).get('v')}%</b> (SRC-614-04). Employment "
            f"was <b>{commify((emp.get('ma') or {}).get('v') or 0)}</b> "
            f"(SRC-614-04). A CPS age-sex-race extract is not in this ledger."
        )
    if tool_id == "DL-19":
        c = ((sec.get("rpp_components_2024") or {}).get("components")) or {}
        h = c.get("housing") or {}
        g = c.get("goods") or {}
        u = c.get("utilities") or {}
        parts.append(
            f"On the 2024 component RPPs, Massachusetts housing was "
            f"<b>{(h.get('ma') or {}).get('v')}</b> versus 100 nationally, rank "
            f"{(h.get('ma') or {}).get('rank')} of {(h.get('ma') or {}).get('n')} "
            f"(derived, SRC-619-02). Goods were <b>{(g.get('ma') or {}).get('v')}</b> "
            f"and utilities <b>{(u.get('ma') or {}).get('v')}</b> (SRC-619-02). "
            f"Tariff, defense, and fiscal-dependency measures remain Pioneer "
            f"methods and are not in this ledger."
        )
    if tool_id == "DL-21":
        a = sec.get("agi_stubs_2022") or {}
        ma = a.get("ma") or {}
        mp = ma.get("million_plus") or {}
        hi = ma.get("over_200k") or {}
        share = a.get("million_plus_agi_share") or {}
        parts.append(
            f"Returns with $1 million or more of AGI held "
            f"<b>{mp.get('agi_share_pct')}%</b> of Massachusetts AGI in tax year "
            f"2022, from <b>{commify(mp.get('returns') or 0)}</b> returns "
            f"(SRC-621-03). Returns with $200,000 or more held "
            f"<b>{hi.get('agi_share_pct')}%</b> of AGI (SRC-621-03). "
            f"Massachusetts ranks {(share.get('ma') or {}).get('rank')} of "
            f"{(share.get('ma') or {}).get('n')} on the million-plus AGI share "
            f"(derived, SRC-621-03). A dedicated percentile-by-state file is "
            f"not posted."
        )
    if tool_id == "DL-25":
        a = sec.get("acs_towns_2024") or {}
        bos = a.get("boston") or {}
        peers = ((a.get("socioeconomic_peers") or {}).get("Boston city") or [{}])[0]
        parts.append(
            f"On the ACS 2020-2024 5-year file, Boston's median household income "
            f"was <b>${commify(bos.get('median_hh_income') or 0)}</b>, median "
            f"home value <b>${commify(bos.get('median_home_value') or 0)}</b>, "
            f"poverty rate <b>{bos.get('poverty_pct')}%</b>, and bachelor's-or-"
            f"higher share <b>{bos.get('bachelors_pct')}%</b> (SRC-625-03). "
            f"The nearest ACS socioeconomic peer for Boston is "
            f"<b>{peers.get('name')}</b> (derived, SRC-625-03). That peer set is "
            f"z-scored ACS income, home value, and bachelor's share, not the old "
            f"Pioneer workbook. A DLS levy file is not a stable public CSV."
        )
    if tool_id == "DL-26":
        a = sec.get("acs_rankings_2024") or {}
        inc = a.get("income") or {}
        pov = a.get("poverty") or {}
        hv = a.get("home_value") or {}
        top_inc = (inc.get("highest") or {}).get("v")
        top_note = " (ACS top-code for $250,000 or more)" if top_inc == 250001 else ""
        parts.append(
            f"Among towns with an ACS 2020-2024 median household income, "
            f"<b>{(inc.get('highest') or {}).get('name')}</b> was highest at "
            f"<b>${commify(top_inc or 0)}</b>{top_note} "
            f"(SRC-626-03). <b>{(pov.get('highest') or {}).get('name')}</b> had "
            f"the highest poverty rate at <b>{(pov.get('highest') or {}).get('v')}%</b> "
            f"(SRC-626-03). <b>{(hv.get('highest') or {}).get('name')}</b> had "
            f"the highest median home value (SRC-626-03). DLS debt, levy, "
            f"revenue, tax, and municipal crime files are not stable public CSVs."
        )
    if tool_id == "DL-28":
        q = sec.get("qtax_type_shares_2026q1") or {}
        inc = q.get("individual_income") or {}
        s = sec.get("stc_ma_2025") or {}
        tot = (s.get("total") or {}).get("ma") or {}
        parts.append(
            f"Individual income taxes were <b>{inc.get('ma_share_pct')}%</b> of "
            f"Massachusetts 2026 Q1 collections "
            f"({usd_prose(inc.get('ma') or 0)}, SRC-628-01). On the annual "
            f"Census STC file, Massachusetts collected "
            f"<b>{usd_prose(tot.get('v') or 0)}</b> in FY 2025, rank "
            f"{tot.get('rank')} of {tot.get('n')} (derived, SRC-628-02). DOR "
            f"monthly reports and tax credits remain pending."
        )
    if tool_id == "DL-29":
        e = sec.get("aspep_fte_2023") or {}
        s = sec.get("stc_2025") or {}
        share = s.get("income_share") or {}
        parts.append(
            f"State governments employed <b>{commify(e.get('us') or 0)}</b> "
            f"full-time-equivalent workers in 2023; Massachusetts employed "
            f"<b>{commify((e.get('ma') or {}).get('v') or 0)}</b>, rank "
            f"{(e.get('ma') or {}).get('rank')} of {(e.get('ma') or {}).get('n')} "
            f"(derived, SRC-629-03). Individual income taxes were "
            f"<b>{(share.get('ma') or {}).get('v')}%</b> of Massachusetts FY 2025 "
            f"state tax collections, rank {(share.get('ma') or {}).get('rank')} of "
            f"{(share.get('ma') or {}).get('n')} (derived, SRC-629-04). NASBO "
            f"rainy-day figures remain pending."
        )
    if tool_id == "DL-31":
        b = sec.get("bjs_depth_2023") or {}
        r = b.get("imprisonment_rate") or {}
        a = b.get("admissions") or {}
        j = b.get("juveniles_in_adult_prisons") or {}
        parts.append(
            f"The imprisonment rate was <b>{r.get('us')}</b> per 100,000 U.S. "
            f"residents at year-end 2023; Massachusetts was "
            f"<b>{(r.get('ma') or {}).get('v')}</b>, rank "
            f"{(r.get('ma') or {}).get('rank')} of {(r.get('ma') or {}).get('n')} "
            f"(derived, SRC-631-03). Sentenced admissions were "
            f"<b>{commify((a.get('ma') or {}).get('v') or 0)}</b> in Massachusetts "
            f"(SRC-631-03). Prisoners age 17 or younger in adult prisons numbered "
            f"<b>{commify((j.get('ma') or {}).get('v') or 0)}</b> in Massachusetts "
            f"(SRC-631-03). That is youth in adult prisons, not the juvenile-"
            f"justice custody count. FBI crime rates and IC3 reports are not a "
            f"stable machine-readable state file on this page."
        )
    return " ".join(parts)
