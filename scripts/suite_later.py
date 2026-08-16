#!/usr/bin/env python3
"""Later views and remaining stub builders for the suite.

Each helper fetches a public file and returns verified figures only.
Nothing here invents a cell. Apps whose files are still blocked stay stubs.
"""
from __future__ import annotations

import csv
import io
import json
import sys
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict

from openpyxl import load_workbook

from suite_common import (
    FIPS_TO_ST,
    REVISED,
    STATE_NAMES,
    UA,
    commify,
    fetch,
    fetch_text,
    finish_live,
    fl_cell,
    geo_to_st,
    parse_num,
    pct,
    rank_named,
    rank_rows,
    usd_prose,
    yoy_pct,
)
from suite_public_later import MORE_SECONDARY, MORE_STRIP, more_lead
from suite_windows import attach_windows, windows_from_trend
from suite_substance import sec_boston_earners, sec_k12_staff

DIGEST_219 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn219.46.xlsx"
DIGEST_208 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn208.30.xlsx"
DIGEST_226 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn226.40.xlsx"
DIGEST_233 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn233.40.xlsx"
DIGEST_315 = "https://nces.ed.gov/programs/digest/d24/tables/xls/tabn315.10.xlsx"
URL_SAINC = "https://apps.bea.gov/regional/zip/SAINC.zip"
URL_FHFA = "https://www.fhfa.gov/hpi/download/annual/hpi_at_state.csv"
URL_SEDS = "https://www.eia.gov/state/seds/sep_use/total/csv/use_all_btu.csv"
URL_QCEW = "https://data.bls.gov/cew/data/api/2025/4/industry/10.csv"
URL_CO_OUT = "https://www.irs.gov/pub/irs-soi/countyoutflow2223.csv"
URL_CO_IN = "https://www.irs.gov/pub/irs-soi/countyinflow2223.csv"
URL_ST_OUT = "https://www.irs.gov/pub/irs-soi/stateoutflow2223.csv"
URL_ST_IN = "https://www.irs.gov/pub/irs-soi/stateinflow2223.csv"
URL_NTD = "https://data.transportation.gov/resource/8bui-9xvu.json"
URL_CTHRU = "https://cthru.data.socrata.com/resource/9ttk-7vz6.json"
URL_SPEND = "https://cthru.data.socrata.com/resource/pegc-naaa.json"
URL_BOS_BUDGET = (
    "https://data.boston.gov/datastore/dump/8f2971f0-7a0d-401d-8376-0289e3b810ba"
)
CMS_QUERY = "https://data.cms.gov/provider-data/api/1/datastore/query/xubh-q36u/0"
URL_E2C_PATHWAYS = "https://educationtocareer.data.mass.gov/resource/9p45-t37j.json"
CH74_PATHWAY = "Career Technical Education (Chapter 74 Programs)"

# Two-path checks against publisher-printed or previously verified cells.
VERIFY_US_ACGR_2021_22 = 86.6
VERIFY_US_SAT_2023 = 1028
VERIFY_US_TEACHERS_FALL_2022 = 3228894.87690338
VERIFY_US_FACULTY_FALL_2023 = 1530513
VERIFY_US_OSS_2020_21 = 1.3248361838221818
VERIFY_US_PI_2025 = 26109831.2  # millions of dollars
VERIFY_NTD_US_JUN_2026 = 690656393
VERIFY_CMS_MA_HOSPITALS = 84
VERIFY_MA_CH74_SY2026 = 59962


def _wb(url, timeout=120):
    return load_workbook(io.BytesIO(fetch(url, timeout=timeout)), data_only=True)


def _kpi(label, value, detail, why, src):
    return {"label": label, "value": value, "detail": detail, "why": why, "src": src}


def _ma(ranked):
    rec = next((r for r in ranked if r.get("st") == "MA"), None)
    if not rec:
        sys.exit("FATAL: ranking is missing Massachusetts")
    return rec


def _extremes(ranked):
    return ranked[0], ranked[-1]


def _soda(url, params):
    q = urllib.parse.urlencode(params)
    return json.loads(fetch(url + "?" + q, timeout=120))


def _digest_year_col(ws, header_row, token, occurrence=0):
    headers = [c.value for c in ws[header_row]]
    hits = []
    for i, h in enumerate(headers):
        if h is None:
            continue
        label = str(h).replace("\xa0", " ").replace("\n", " ").replace(" ", "")
        want = token.replace(" ", "").replace("\n", "")
        if want in label:
            hits.append(i)
    if occurrence >= len(hits):
        sys.exit(f"FATAL: {ws.title} has no header {token!r} occurrence {occurrence}")
    return hits[occurrence]


def _digest_state_col(ws, header_row, year_col, start_row):
    values, us_val = {}, None
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        st = geo_to_st(row[0])
        if not st:
            continue
        v = parse_num(row[year_col])
        if v is None:
            continue
        if st == "US":
            us_val = v
        else:
            values[st] = v
    if len(values) < 48:
        sys.exit(f"FATAL: {ws.title} parsed {len(values)} states")
    return values, us_val


def _state_snapshot(values, us_val, round_to=None):
    ranked = rank_rows(values, higher_is_better=True)
    if round_to is not None:
        for rec in ranked:
            rec["v"] = round(rec["v"], round_to)
        if us_val is not None:
            us_val = round(us_val, round_to)
    ma = _ma(ranked)
    hi, lo = _extremes(ranked)
    out = {
        "us": us_val,
        "ma": {"v": ma["v"], "rank": ma["rank"], "n": ma["n"]},
        "highest": {"st": hi["st"], "name": hi["name"], "v": hi["v"]},
        "lowest": {"st": lo["st"], "name": lo["name"], "v": lo["v"]},
        "n_ranked": ma["n"],
    }
    fl = fl_cell(ranked)
    if fl:
        out["fl"] = fl
    return out


# ---------------------------------------------------------------------------
# Secondary series for live apps
# ---------------------------------------------------------------------------

def sec_acgr():
    wb = _wb(DIGEST_219)
    ws = wb.active
    col = _digest_year_col(ws, 3, "2021-22")
    values, us_val = _digest_state_col(ws, 3, col, 7)
    if us_val is None or abs(us_val - VERIFY_US_ACGR_2021_22) > 0.05:
        sys.exit(f"FATAL: Digest 219.46 US ACGR 2021-22 is {us_val}")
    snap = _state_snapshot(values, us_val, round_to=1)
    snap.update({
        "label": "Public high school 4-year adjusted cohort graduation rate, 2021-22",
        "src": "SRC-607-03",
        "unit": "percent",
        "as_of_label": "School year 2021-22",
    })
    return snap


def sec_discipline():
    wb = _wb(DIGEST_233)
    ws = wb.active
    values, us_val = _digest_state_col(ws, 2, 1, 6)
    if us_val is None or abs(us_val - VERIFY_US_OSS_2020_21) > 0.01:
        sys.exit(f"FATAL: Digest 233.40 US OSS 2020-21 is {us_val}")
    snap = _state_snapshot(values, us_val, round_to=2)
    snap.update({
        "label": "Share of public-school students with an out-of-school suspension, 2020-21",
        "src": "SRC-607-04",
        "unit": "percent",
        "as_of_label": "School year 2020-21",
        "note": "Higher is more students suspended, not a performance ranking.",
    })
    return snap


def sec_sat():
    wb = _wb(DIGEST_226)
    ws = wb.active
    # Row 2 year labels; 2023 total SAT mean is the first column under 2023.
    headers = [c.value for c in ws[2]]
    year_col = None
    for i, h in enumerate(headers):
        if h == 2023 or str(h).strip() == "2023":
            year_col = i
            break
    if year_col is None:
        sys.exit("FATAL: Digest 226.40 has no 2023 column")
    values, us_val = _digest_state_col(ws, 2, year_col, 7)
    if us_val is None or abs(us_val - VERIFY_US_SAT_2023) > 0.5:
        sys.exit(f"FATAL: Digest 226.40 US SAT 2023 is {us_val}")
    # participation is 6 columns after the mean
    part, us_part = {}, None
    for row in ws.iter_rows(min_row=7, values_only=True):
        st = geo_to_st(row[0])
        if not st:
            continue
        v = parse_num(row[year_col + 6]) if year_col + 6 < len(row) else None
        if v is None:
            continue
        if st == "US":
            us_part = v
        else:
            part[st] = v
    snap = _state_snapshot(values, us_val, round_to=0)
    snap.update({
        "label": "SAT mean total score, 2023 high-school graduates",
        "src": "SRC-608-02",
        "unit": "points",
        "as_of_label": "2023 graduates",
        "participation_pct": {
            "us": round(us_part, 1) if us_part is not None else None,
            "ma": round(part["MA"], 1) if "MA" in part else None,
        },
        "note": "Means are not comparable across states with very different participation rates.",
    })
    return snap


def sec_faculty_us():
    wb = _wb(DIGEST_315)
    ws = wb.active
    us_val = None
    for row in ws.iter_rows(min_row=6, values_only=True):
        year = str(row[0] or "").replace("\\2\\", "").strip()
        if year.replace("\\1\\", "").replace("\\2\\", "").strip() == "2023":
            us_val = parse_num(row[1])
            break
    if us_val is None or abs(us_val - VERIFY_US_FACULTY_FALL_2023) > 1:
        sys.exit(f"FATAL: Digest 315.10 Fall 2023 faculty is {us_val}")
    return {
        "label": "Faculty in degree-granting institutions, Fall 2023 (national)",
        "src": "SRC-608-03",
        "unit": "faculty",
        "as_of_label": "Fall 2023",
        "us": int(round(us_val)),
        "ma": None,
        "note": "This Digest table is national only; it has no state column.",
    }


def _ch74_rollup_where(sy, org_type):
    return (
        f"sy='{sy}' AND org_type='{org_type}' AND pathway='{CH74_PATHWAY}' AND "
        f"(program='{CH74_PATHWAY}' OR program='Career Technical Education Programs')"
    )


def sec_ma_voctech():
    """Massachusetts Chapter 74 CTE from the E2C pathways enrollment file."""
    trend = []
    state_by_year = {}
    for sy in ("2022", "2023", "2024", "2025", "2026"):
        rows = _soda(URL_E2C_PATHWAYS, {
            "$where": _ch74_rollup_where(sy, "State"),
            "$limit": "2",
        })
        if len(rows) != 1:
            sys.exit(f"FATAL: E2C Chapter 74 state rollup for SY {sy} has {len(rows)} rows")
        r = rows[0]
        n = parse_num(r.get("program_cnt"))
        hs = parse_num(r.get("totalstu"))
        if n is None or hs is None:
            sys.exit(f"FATAL: E2C Chapter 74 state rollup for SY {sy} is missing counts")
        rec = {
            "sy": int(sy),
            "label": f"School year {int(sy) - 1}-{str(sy)[2:]}",
            "v": int(n),
            "hs_enrollment": int(hs),
            "share": round(n / hs, 4),
        }
        state_by_year[sy] = {**rec, "raw": r}
        trend.append({"y": rec["label"], "v": rec["v"]})
    latest = state_by_year["2026"]
    if latest["v"] != VERIFY_MA_CH74_SY2026:
        sys.exit(
            f"FATAL: E2C Chapter 74 SY 2026 is {latest['v']}, "
            f"expected {VERIFY_MA_CH74_SY2026}"
        )
    districts = _soda(URL_E2C_PATHWAYS, {
        "$where": _ch74_rollup_where("2026", "District"),
        "$select": "dist_name,program_cnt,program_pct,totalstu",
        "$order": "program_cnt DESC",
        "$limit": "200",
    })
    schools = _soda(URL_E2C_PATHWAYS, {
        "$where": _ch74_rollup_where("2026", "School"),
        "$select": "dist_name,org_name,program_cnt,program_pct,totalstu",
        "$order": "program_cnt DESC",
        "$limit": "300",
    })
    d_sum = sum(int(parse_num(r.get("program_cnt")) or 0) for r in districts)
    s_sum = sum(int(parse_num(r.get("program_cnt")) or 0) for r in schools)
    if d_sum != latest["v"] or s_sum != latest["v"]:
        sys.exit(
            f"FATAL: E2C Chapter 74 SY 2026 state {latest['v']} "
            f"!= districts {d_sum} or schools {s_sum}"
        )
    if len(districts) < 80 or len(schools) < 90:
        sys.exit(
            f"FATAL: E2C Chapter 74 SY 2026 parsed {len(districts)} districts "
            f"and {len(schools)} schools"
        )

    def _named(rows, name_key):
        values = {}
        extra = {}
        for r in rows:
            name = (r.get(name_key) or "").strip()
            v = parse_num(r.get("program_cnt"))
            if not name or v is None:
                continue
            key = name if name not in values else f"{name} ({r.get('dist_name')})"
            values[key] = int(v)
            extra[key] = {
                "share_of_org": parse_num(r.get("program_pct")),
                "org_enrollment": int(parse_num(r.get("totalstu")) or 0) or None,
                "district": (r.get("dist_name") or "").strip() or None,
            }
        ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
        for rec in ranked:
            rec.update(extra.get(rec["name"]) or {})
        return ranked

    district_ranked = _named(districts, "dist_name")
    school_ranked = _named(schools, "org_name")
    regional = [
        r for r in district_ranked
        if (r.get("share_of_org") or 0) >= 0.99
        and (
            "vocational" in r["name"].lower()
            or "agricultural" in r["name"].lower()
        )
    ]
    programs = _soda(URL_E2C_PATHWAYS, {
        "$where": (
            "sy='2026' AND org_type='State' AND "
            f"pathway='{CH74_PATHWAY}' AND "
            f"program!='{CH74_PATHWAY}' AND "
            "program!='Career Technical Education Programs'"
        ),
        "$select": "program,program_cnt,g9_cnt,g10_cnt,g11_cnt,g12_cnt",
        "$order": "program_cnt DESC",
        "$limit": "80",
    })
    program_rows = []
    for r in programs:
        v = parse_num(r.get("program_cnt"))
        if v is None:
            continue
        program_rows.append({
            "name": r["program"],
            "v": int(v),
            "g9": int(parse_num(r.get("g9_cnt")) or 0),
            "g10": int(parse_num(r.get("g10_cnt")) or 0),
            "g11": int(parse_num(r.get("g11_cnt")) or 0),
            "g12": int(parse_num(r.get("g12_cnt")) or 0),
        })
    exploratory = next((p for p in program_rows if p["name"] == "Exploratory"), None)
    occupational = [p for p in program_rows if p["name"] != "Exploratory"]

    def _state_program(pathway, program):
        rows = _soda(URL_E2C_PATHWAYS, {
            "$where": (
                f"sy='2026' AND org_type='State' AND pathway='{pathway}' "
                f"AND program='{program}'"
            ),
            "$select": "program_cnt,program_pct,totalstu",
            "$limit": "1",
        })
        if not rows:
            sys.exit(f"FATAL: E2C missing {pathway} / {program} for SY 2026")
        return {
            "v": int(parse_num(rows[0]["program_cnt"])),
            "share": parse_num(rows[0].get("program_pct")),
            "hs_enrollment": int(parse_num(rows[0].get("totalstu"))),
        }

    after_dark = _state_program(
        "After Dark", "Career Technical Education Partnership Programs"
    )
    connections = _state_program(
        "Career Connections (Non-Chapter 74 Programs)",
        "Career Connections Programs",
    )
    early_college = _state_program("Early College", "Early College Programs")
    innovation = _state_program(
        "Innovation Career Pathways", "Innovation Career Pathway Programs"
    )
    all_pathways = _state_program(
        "All (Pathways/Programs)", "All (Pathways & Programs)"
    )
    raw = latest["raw"]
    yoy = yoy_pct(latest["v"], state_by_year["2025"]["v"])
    since_2022 = yoy_pct(latest["v"], state_by_year["2022"]["v"])
    hi = district_ranked[0]
    return {
        "label": "Massachusetts Chapter 74 career technical education enrollment, 2025-26",
        "src": "SRC-606-03",
        "unit": "students",
        "as_of_label": "School year 2025-26",
        "v": latest["v"],
        "hs_enrollment": latest["hs_enrollment"],
        "share": latest["share"],
        "yoy_pct": yoy,
        "change_since_2021_22_pct": since_2022,
        "change_since_2021_22": latest["v"] - state_by_year["2022"]["v"],
        "districts": len(district_ranked),
        "schools": len(school_ranked),
        "regional_vocational_districts": len(regional),
        "highest": {"name": hi["name"], "v": hi["v"]},
        "female_share": parse_num(raw.get("fe_pct")),
        "male_share": parse_num(raw.get("ma_pct")),
        "low_income": int(parse_num(raw.get("li_cnt")) or 0),
        "low_income_share": parse_num(raw.get("li_pct")),
        "swd": int(parse_num(raw.get("swd_cnt")) or 0),
        "swd_share": parse_num(raw.get("swd_pct")),
        "english_learners": int(parse_num(raw.get("el_cnt")) or 0),
        "english_learner_share": parse_num(raw.get("el_pct")),
        "after_dark": after_dark,
        "career_connections": connections,
        "early_college": early_college,
        "innovation_pathways": innovation,
        "all_pathways": all_pathways,
        "exploratory": exploratory,
        "top_occupational_programs": occupational[:8],
        "trend": trend,
        "district_rows": [
            {"name": r["name"], "v": r["v"], "rank": r["rank"], "n": r["n"],
             "share_of_org": r.get("share_of_org")}
            for r in district_ranked
        ],
        "school_rows": [
            {"name": r["name"], "district": r.get("district"), "v": r["v"],
             "rank": r["rank"], "n": r["n"]}
            for r in school_ranked[:40]
        ],
        "note": (
            "After Dark enrollment is a published subset of Chapter 74. "
            "Students may appear in more than one college-and-career pathway, "
            "so pathway totals do not add to the all-pathways count. "
            "Waitlists and lottery outcomes are not in this file."
        ),
    }


def sec_teachers():
    wb = _wb(DIGEST_208)
    ws = wb.active
    col = _digest_year_col(ws, 3, "Fall 2022")
    values, us_val = _digest_state_col(ws, 3, col, 6)
    if us_val is None or abs(us_val - VERIFY_US_TEACHERS_FALL_2022) > 1:
        sys.exit(f"FATAL: Digest 208.30 Fall 2022 teachers are {us_val}")
    values = {st: round(v) for st, v in values.items()}
    us_val = round(us_val)
    snap = _state_snapshot(values, us_val)
    snap.update({
        "label": "Public elementary and secondary teachers (FTE), Fall 2022",
        "src": "SRC-609-02",
        "unit": "full-time-equivalent teachers",
        "as_of_label": "Fall 2022",
    })
    return snap


def sec_qcew():
    text = fetch_text(URL_QCEW, timeout=120)
    values = {}
    wages = {}
    emp = {}
    for r in csv.DictReader(io.StringIO(text)):
        if not (
            r.get("own_code") == "0"
            and r.get("agglvl_code") == "50"
            and r.get("industry_code") == "10"
        ):
            continue
        fips = (r.get("area_fips") or "").zfill(5)
        st = FIPS_TO_ST.get(fips[:2])
        if not st or st == "US":
            continue
        wage = parse_num(r.get("avg_wkly_wage"))
        e1 = parse_num(r.get("month1_emplvl"))
        e2 = parse_num(r.get("month2_emplvl"))
        e3 = parse_num(r.get("month3_emplvl"))
        tw = parse_num(r.get("total_qtrly_wages"))
        if wage is None:
            continue
        values[st] = wage
        wages[st] = tw
        if None not in (e1, e2, e3):
            emp[st] = (e1 + e2 + e3) / 3
    if "MA" not in values or len(values) < 50:
        sys.exit(f"FATAL: QCEW 2025 Q4 parsed {len(values)} states")
    # US average weekly wage is not a published row here; derive from sums.
    us_wages = sum(v for v in wages.values() if v is not None)
    us_emp = sum(emp.values())
    us_wage = round(us_wages / us_emp / 13) if us_emp else None
    snap = _state_snapshot(values, us_wage, round_to=0)
    snap.update({
        "label": "Average weekly wage, all industries, 2025 Q4",
        "src": "SRC-614-02",
        "unit": "dollars per week",
        "as_of_label": "2025 Q4",
        "ma_employment": int(round(emp["MA"])) if "MA" in emp else None,
        "note": "The U.S. weekly wage is derived from the sum of published state wage and employment cells (derived, SRC-614-02). It is not a separate BLS U.S. line.",
    })
    return snap


def sec_personal_income():
    zf = zipfile.ZipFile(io.BytesIO(fetch(URL_SAINC, timeout=180)))
    pi, pcpi = {}, {}
    us_pi = us_pcpi = None
    for st, name in STATE_NAMES.items():
        if st == "US":
            fn = "SAINC4_US_1929_2025.csv"
        else:
            fn = f"SAINC4_{st}_1929_2025.csv"
        if fn not in zf.namelist():
            continue
        rows = csv.DictReader(io.StringIO(zf.read(fn).decode("latin-1")))
        for r in rows:
            code = str(r.get("LineCode") or "").strip()
            v = parse_num(r.get("2025"))
            if v is None:
                continue
            if code == "10":
                if st == "US":
                    us_pi = v
                else:
                    pi[st] = v
            elif code == "30":
                if st == "US":
                    us_pcpi = v
                else:
                    pcpi[st] = v
    if us_pi is None or abs(us_pi - VERIFY_US_PI_2025) > 0.2:
        sys.exit(f"FATAL: SAINC4 US 2025 personal income is {us_pi}")
    if "MA" not in pi or "MA" not in pcpi:
        sys.exit("FATAL: SAINC4 missing Massachusetts 2025")
    snap = _state_snapshot(pi, us_pi, round_to=1)
    pc = _state_snapshot(pcpi, us_pcpi, round_to=0)
    return {
        "label": "Personal income, 2025",
        "src": "SRC-615-02",
        "unit": "millions of dollars",
        "as_of_label": "Calendar year 2025",
        "us": us_pi,
        "us_dollars": us_pi * 1_000_000,
        "ma": snap["ma"],
        "highest": snap["highest"],
        "lowest": snap["lowest"],
        "n_ranked": snap["n_ranked"],
        "per_capita": {
            "label": "Per capita personal income, 2025",
            "us": us_pcpi,
            "ma": pc["ma"],
            "highest": pc["highest"],
            "lowest": pc["lowest"],
            "unit": "dollars",
        },
    }


def sec_fhfa():
    wb = _wb(URL_FHFA)
    ws = wb.active
    latest = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        st = (row[1] or "").strip() if row[1] else None
        year = parse_num(row[3])
        chg = parse_num(row[4])
        if st not in STATE_NAMES or st == "US" or year is None or chg is None:
            continue
        year = int(year)
        if st not in latest or year > latest[st][0]:
            latest[st] = (year, chg)
    years = {y for y, _c in latest.values()}
    year = max(years) if years else None
    values = {st: chg for st, (y, chg) in latest.items() if y == year}
    if year != 2025 or "MA" not in values or len(values) < 50:
        sys.exit(f"FATAL: FHFA HPI latest year {year} parsed {len(values)} states")
    snap = _state_snapshot(values, None, round_to=2)
    snap.update({
        "label": f"FHFA all-transactions house-price index, annual change, {year}",
        "src": "SRC-616-02",
        "unit": "percent",
        "as_of_label": f"Calendar year {year}",
        "us": None,
        "note": "FHFA does not publish a U.S. row in this state file. The index is developmental (FHFA note, March 31, 2026).",
    })
    return snap


def sec_seds():
    text = fetch_text(URL_SEDS, timeout=120)
    values, us_val = {}, None
    for r in csv.DictReader(io.StringIO(text)):
        if r.get("MSN") != "TETCB":
            continue
        st = (r.get("State") or "").strip()
        v = parse_num(r.get("2024"))
        if v is None:
            continue
        if st == "US":
            us_val = v
        elif st in STATE_NAMES:
            values[st] = v
    if us_val is None or "MA" not in values or us_val < 1e7:
        sys.exit(f"FATAL: SEDS TETCB 2024 US is {us_val}")
    snap = _state_snapshot(values, us_val, round_to=0)
    snap.update({
        "label": "Total energy consumption, 2024 (billion Btu)",
        "src": "SRC-624-02",
        "unit": "billion Btu",
        "as_of_label": "Calendar year 2024",
        "note": "SEDS series TETCB. Production (TEPRB) was not in this consumption file.",
    })
    return snap


def _county_mig_name(raw):
    return (raw or "").replace(" Total Migration-US", "").strip()


def _load_county_migration(url, state_fips=None):
    """Total Migration-US inflow or outflow by county.

    Outflow file: origin county -> 97/000. Inflow file: 97/000 -> destination
    county. When state_fips is set, only that state's counties are kept.
    """
    text = fetch_text(url, timeout=180)
    outflow = url.endswith("countyoutflow2223.csv")
    out = {}
    for r in csv.DictReader(io.StringIO(text)):
        if outflow:
            st, cty = r.get("y1_statefips"), r.get("y1_countyfips")
            if r.get("y2_statefips") != "97" or r.get("y2_countyfips") != "000":
                continue
            name = _county_mig_name(r.get("y1_countyname") or r.get("y2_countyname"))
        else:
            st, cty = r.get("y2_statefips"), r.get("y2_countyfips")
            if r.get("y1_statefips") != "97" or r.get("y1_countyfips") != "000":
                continue
            name = _county_mig_name(r.get("y2_countyname") or r.get("y1_countyname"))
        if state_fips and st != state_fips:
            continue
        n1 = parse_num(r.get("n1"))
        if n1 is None or not cty or cty == "000" or not st or st == "00":
            continue
        if st not in FIPS_TO_ST or FIPS_TO_ST[st] == "US":
            continue
        key = st + cty
        out[key] = {
            "name": name,
            "st": FIPS_TO_ST[st],
            "fips": key,
            "n1": int(n1),
        }
    return out


def _county_migration_nets(state_fips=None):
    outflow = _load_county_migration(URL_CO_OUT, state_fips)
    inflow = _load_county_migration(URL_CO_IN, state_fips)
    rows = []
    for key, o in outflow.items():
        inn = inflow.get(key)
        if not inn:
            continue
        st = o["st"]
        name = o["name"] or inn["name"]
        label = name if state_fips else f"{name}, {st}"
        rows.append({
            "name": label,
            "county": name,
            "st": st,
            "fips": key,
            "v": inn["n1"] - o["n1"],
            "in": inn["n1"],
            "out": o["n1"],
        })
    rows.sort(key=lambda r: r["v"], reverse=True)
    for i, rec in enumerate(rows, 1):
        rec["rank"] = i
        rec["n"] = len(rows)
    return rows


def _rerank_county_migration(rows, name_key="name"):
    ranked = sorted(rows, key=lambda r: r["v"], reverse=True)
    out = []
    for i, rec in enumerate(ranked, 1):
        item = dict(rec)
        item["name"] = rec[name_key]
        item["rank"] = i
        item["n"] = len(ranked)
        out.append(item)
    return out


def _pack_county_rows(ranked, state_name, min_counties):
    if len(ranked) < min_counties:
        sys.exit(
            f"FATAL: IRS county migration parsed {len(ranked)} {state_name} counties"
        )
    hi, lo = ranked[0], ranked[-1]
    return {
        "label": f"{state_name} county net domestic taxpayer migration, 2022-23",
        "src": "SRC-620-02",
        "unit": "returns",
        "as_of_label": "Tax years 2022-23",
        "n_counties": len(ranked),
        "highest": {"name": hi["name"], "v": hi["v"], "st": hi["st"]},
        "lowest": {"name": lo["name"], "v": lo["v"], "st": lo["st"]},
        "counties": [
            {"name": r["name"], "v": r["v"], "rank": r["rank"], "st": r["st"]}
            for r in ranked
        ],
        "note": (
            "Net equals Total Migration-US inflow n1 minus outflow n1. "
            "Same-state and foreign rows are excluded."
        ),
    }


def _state_pair_rows(url):
    """IRS SOI state outflow: y1 is origin FIPS, y2 is destination FIPS."""
    rows = []
    for r in csv.DictReader(io.StringIO(fetch_text(url))):
        o_fips = str(r.get("y1_statefips") or "").strip().zfill(2)
        d_fips = str(r.get("y2_statefips") or "").strip().zfill(2)
        if d_fips in {"96", "97", "98"} or o_fips in {"96", "97", "98"}:
            continue
        origin = FIPS_TO_ST.get(o_fips)
        dest = FIPS_TO_ST.get(d_fips)
        if origin not in STATE_NAMES or dest not in STATE_NAMES:
            continue
        if origin == dest or origin == "US" or dest == "US":
            continue
        n1 = parse_num(r.get("n1"))
        agi = parse_num(r.get("AGI") or r.get("agi"))
        if n1 is None:
            continue
        rows.append({
            "origin": origin,
            "dest": dest,
            "returns": int(round(n1)),
            "agi": int(round(agi)) if agi is not None else None,
        })
    return rows


def _top_from(rows, origin, key="dest", n=8):
    picked = [r for r in rows if r["origin"] == origin]
    picked.sort(key=lambda r: r["returns"], reverse=True)
    out = []
    for r in picked[:n]:
        st = r[key]
        item = {
            "st": st,
            "name": STATE_NAMES.get(st, st),
            "returns": r["returns"],
        }
        if r.get("agi") is not None:
            item["agi"] = r["agi"]
        out.append(item)
    return out


def _top_into(rows, dest, n=8):
    picked = [r for r in rows if r["dest"] == dest]
    picked.sort(key=lambda r: r["returns"], reverse=True)
    out = []
    for r in picked[:n]:
        item = {
            "st": r["origin"],
            "name": STATE_NAMES.get(r["origin"], r["origin"]),
            "returns": r["returns"],
        }
        if r.get("agi") is not None:
            item["agi"] = r["agi"]
        out.append(item)
    return out


def _pair_cell(rows, origin, dest):
    for r in rows:
        if r["origin"] == origin and r["dest"] == dest:
            cell = {"returns": r["returns"]}
            if r.get("agi") is not None:
                cell["agi"] = r["agi"]
            return cell
    return None


def sec_state_pair_flows():
    """Origin-destination taxpayer returns from the same IRS SOI state files."""
    outflow = _state_pair_rows(URL_ST_OUT)
    if len(outflow) < 2000:
        sys.exit(f"FATAL: IRS state outflow pairs parsed {len(outflow)} rows")
    ma_fl = _pair_cell(outflow, "MA", "FL")
    fl_ma = _pair_cell(outflow, "FL", "MA")
    if not ma_fl or not fl_ma:
        sys.exit("FATAL: IRS state outflow missing the Massachusetts-Florida pair")
    ma_out = _top_from(outflow, "MA")
    fl_in = _top_into(outflow, "FL")
    if not ma_out or ma_out[0]["st"] not in STATE_NAMES:
        sys.exit("FATAL: IRS state outflow missing Massachusetts destinations")
    return {
        "label": "State-to-state taxpayer destinations, tax years 2022-23",
        "src": "SRC-620-01",
        "unit": "returns",
        "as_of_label": "Tax years 2022-23",
        "note": (
            "Origin-destination returns from IRS SOI stateoutflow2223.csv. "
            "Same-state, foreign, and Total Migration-US aggregate rows are "
            "excluded. AGI is the published IRS column, in thousands of dollars."
        ),
        "ma_to_fl": ma_fl,
        "fl_to_ma": fl_ma,
        "ma_out_top": ma_out,
        "ma_in_top": _top_into(outflow, "MA"),
        "fl_in_top": fl_in,
        "fl_out_top": _top_from(outflow, "FL"),
    }


def sec_county_migration_bundle():
    """Parse the two IRS county files once; return MA, FL, and U.S. packs."""
    ranked = _county_migration_nets()
    if len(ranked) < 2000:
        sys.exit(f"FATAL: IRS county migration parsed {len(ranked)} U.S. counties")
    ma = _rerank_county_migration(
        [r for r in ranked if r["st"] == "MA"], name_key="county"
    )
    fl = _rerank_county_migration(
        [r for r in ranked if r["st"] == "FL"], name_key="county"
    )
    hi, lo = ranked[0], ranked[-1]
    gains = ranked[:15]
    losses = list(reversed(ranked[-15:]))
    return {
        "ma_county_taxpayer_migration_2022_23": _pack_county_rows(
            ma, "Massachusetts", 10
        ),
        "fl_county_taxpayer_migration_2022_23": _pack_county_rows(
            fl, "Florida", 50
        ),
        "us_county_taxpayer_migration_2022_23": {
            "label": "U.S. county net domestic taxpayer migration, 2022-23",
            "src": "SRC-620-02",
            "unit": "returns",
            "as_of_label": "Tax years 2022-23",
            "n_counties": len(ranked),
            "highest": {"name": hi["name"], "v": hi["v"], "st": hi["st"]},
            "lowest": {"name": lo["name"], "v": lo["v"], "st": lo["st"]},
            "top_gains": [
                {"name": r["name"], "v": r["v"], "st": r["st"]} for r in gains
            ],
            "top_losses": [
                {"name": r["name"], "v": r["v"], "st": r["st"]} for r in losses
            ],
            "chart": [
                {"name": r["name"], "v": r["v"], "st": r["st"]}
                for r in gains + losses
            ],
            "note": (
                "Every U.S. county with a published Total Migration-US inflow "
                "and outflow. The figure shows the 15 largest net gains and "
                "the 15 largest net losses. Net equals inflow n1 minus "
                "outflow n1. Same-state and foreign rows are excluded."
            ),
        },
    }


def sec_boston_budget():
    text = fetch_text(URL_BOS_BUDGET, timeout=90)
    rdr = csv.DictReader(io.StringIO(text))
    by_dept = defaultdict(float)
    fy26 = fy25 = fy27 = 0.0
    n = 0
    for r in rdr:
        dept = (r.get("Dept") or "").strip()
        # The file includes a citywide total row with every field blank.
        if not dept:
            continue
        a26 = parse_num(r.get("FY26 Appropriation"))
        a25 = parse_num(r.get("FY25 Actual Expense"))
        a27 = parse_num(r.get("FY27 Budget"))
        if a26 is not None:
            by_dept[dept] += a26
            fy26 += a26
        if a25 is not None:
            fy25 += a25
        if a27 is not None:
            fy27 += a27
        n += 1
    if n < 100 or fy26 < 1e9 or fy26 > 8e9:
        sys.exit(f"FATAL: Boston operating budget parsed n={n} fy26={fy26}")
    ranked = rank_named(by_dept, higher_is_better=True, st_key=lambda n: n[:8])
    hi = ranked[0]
    return {
        "label": "City of Boston adopted operating budget, FY26 appropriation",
        "src": "SRC-627-02",
        "unit": "dollars",
        "as_of_label": "FY26 appropriation",
        "fy25_actual": fy25,
        "fy26_appropriation": fy26,
        "fy27_budget": fy27,
        "departments": len(ranked),
        "highest": {"name": hi["name"], "v": hi["v"]},
        "top_five": [{"name": r["name"], "v": r["v"]} for r in ranked[:5]],
    }


SECONDARY = {
    "DL-06": lambda: {"ma_chapter74_cte": sec_ma_voctech()},
    "DL-07": lambda: {"acgr_2021_22": sec_acgr(), "oss_suspension_2020_21": sec_discipline()},
    "DL-08": lambda: {"sat_2023": sec_sat(), "faculty_fall_2023_us": sec_faculty_us()},
    "DL-09": lambda: {"teachers_fte_fall_2022": sec_teachers(), **sec_k12_staff()},
    "DL-14": lambda: {"qcew_avg_weekly_wage_2025q4": sec_qcew()},
    "DL-15": lambda: {"personal_income_2025": sec_personal_income()},
    "DL-16": lambda: {"fhfa_hpi_annual_change_2025": sec_fhfa()},
    "DL-20": lambda: {**sec_county_migration_bundle(), "state_pair_flows_2022_23": sec_state_pair_flows()},
    "DL-24": lambda: {"seds_consumption_2024": sec_seds()},
    "DL-27": lambda: {"boston_operating_budget_fy26": sec_boston_budget(), **sec_boston_earners()},
}


def _fmt_pct_points(v):
    return f"{v:.1f}%" if isinstance(v, float) and not float(v).is_integer() else f"{v}%"


def lead_appendix(tool_id, sec):
    if tool_id == "DL-06":
        v = sec["ma_chapter74_cte"]
        share_pct = f"{v['share'] * 100:.1f}"
        expl = v.get("exploratory") or {}
        top = (v.get("top_occupational_programs") or [{}])[0]
        return (
            f"Chapter 74 career technical education enrolled "
            f"<b>{commify(v['v'])}</b> Massachusetts high-school students in "
            f"2025-26, {share_pct} percent of high-school enrollment "
            f"(SRC-606-03). That is <b>{commify(v['change_since_2021_22'])}</b> "
            f"more students than in 2021-22 (derived, SRC-606-03). "
            f"<b>{v['districts']}</b> districts and <b>{v['schools']}</b> schools "
            f"reported a Chapter 74 program; <b>{v['regional_vocational_districts']}</b> "
            f"were regional vocational or agricultural districts (derived, "
            f"SRC-606-03). <b>{v['highest']['name']}</b> was the largest district "
            f"at <b>{commify(v['highest']['v'])}</b> (SRC-606-03). "
            f"Grade-9 exploratory enrolled <b>{commify(expl.get('v') or 0)}</b>; "
            f"the largest occupational program was <b>{top.get('name')}</b> at "
            f"<b>{commify(top.get('v') or 0)}</b> (SRC-606-03). After Dark "
            f"partnership enrollment was <b>{commify(v['after_dark']['v'])}</b>, "
            f"a published subset of Chapter 74 (SRC-606-03). All college-and-career "
            f"pathways together enrolled <b>{commify(v['all_pathways']['v'])}</b> "
            f"(SRC-606-03)."
        )
    if tool_id == "DL-07":
        a = sec["acgr_2021_22"]
        d = sec["oss_suspension_2020_21"]
        return (
            f"The public high school 4-year graduation rate was "
            f"<b>{a['us']}%</b> in 2021-22 (SRC-607-03). Massachusetts was "
            f"<b>{a['ma']['v']}%</b>, rank {a['ma']['rank']} of {a['ma']['n']} "
            f"(derived, SRC-607-03). Out-of-school suspensions reached "
            f"<b>{d['us']}%</b> of U.S. public-school students in 2020-21; "
            f"Massachusetts was <b>{d['ma']['v']}%</b> (SRC-607-04)."
        )
    if tool_id == "DL-08":
        s = sec["sat_2023"]
        f = sec["faculty_fall_2023_us"]
        part = s["participation_pct"]
        return (
            f"The SAT mean total score for 2023 graduates was <b>{int(s['us'])}</b> "
            f"in the United States, with {part['us']}% of graduates taking the test "
            f"(SRC-608-02). Massachusetts scored <b>{int(s['ma']['v'])}</b> "
            f"({part['ma']}% taking the test) (SRC-608-02). Degree-granting "
            f"institutions employed <b>{commify(f['us'])}</b> faculty in Fall 2023 "
            f"(SRC-608-03)."
        )
    if tool_id == "DL-09":
        t = sec["teachers_fte_fall_2022"]
        return (
            f"Public schools employed <b>{commify(t['us'])}</b> full-time-equivalent "
            f"teachers in Fall 2022 (SRC-609-02). Massachusetts employed "
            f"<b>{commify(t['ma']['v'])}</b>, rank {t['ma']['rank']} of "
            f"{t['ma']['n']} (derived, SRC-609-02)."
        )
    if tool_id == "DL-14":
        q = sec["qcew_avg_weekly_wage_2025q4"]
        return (
            f"Average weekly wages were <b>${commify(q['us'])}</b> in the "
            f"United States in 2025 Q4 (derived, SRC-614-02). Massachusetts "
            f"was <b>${commify(q['ma']['v'])}</b>, rank {q['ma']['rank']} of "
            f"{q['ma']['n']} (derived, SRC-614-02)."
        )
    if tool_id == "DL-15":
        p = sec["personal_income_2025"]
        pc = p["per_capita"]
        return (
            f"Personal income was <b>{usd_prose(p['us_dollars'])}</b> in 2025 "
            f"(SRC-615-02). Massachusetts was <b>{usd_prose(p['ma']['v'] * 1_000_000)}</b>, "
            f"rank {p['ma']['rank']} of {p['ma']['n']} (derived, SRC-615-02). "
            f"Per capita personal income was <b>${commify(pc['us'])}</b> in the "
            f"United States and <b>${commify(pc['ma']['v'])}</b> in Massachusetts, "
            f"rank {pc['ma']['rank']} of {pc['ma']['n']} (derived, SRC-615-02)."
        )
    if tool_id == "DL-16":
        h = sec["fhfa_hpi_annual_change_2025"]
        return (
            f"The FHFA all-transactions house-price index rose "
            f"<b>{h['ma']['v']}%</b> in Massachusetts in 2025, rank "
            f"{h['ma']['rank']} of {h['ma']['n']} on annual change "
            f"(derived, SRC-616-02)."
            + (
                f" Florida changed <b>{h['fl']['v']}%</b>, rank "
                f"{h['fl']['rank']} of {h['fl']['n']} (derived, SRC-616-02)."
                if h.get("fl") else ""
            )
            + f" {h['highest']['name']} had the largest "
            f"increase at {h['highest']['v']}% (SRC-616-02)."
        )
    if tool_id == "DL-20":
        c = sec["ma_county_taxpayer_migration_2022_23"]
        fl = sec.get("fl_county_taxpayer_migration_2022_23") or {}
        us = sec.get("us_county_taxpayer_migration_2022_23") or {}
        bits = [
            f"Among Massachusetts counties, <b>{c['highest']['name']}</b> had "
            f"the largest net domestic taxpayer inflow ({commify(c['highest']['v'])} "
            f"returns) and <b>{c['lowest']['name']}</b> the largest net outflow "
            f"({commify(c['lowest']['v'])}) in 2022-23 (derived, SRC-620-02)."
        ]
        if fl.get("highest") and fl.get("lowest"):
            bits.append(
                f"Among Florida counties, <b>{fl['highest']['name']}</b> had "
                f"the largest net inflow ({commify(fl['highest']['v'])} returns) "
                f"and <b>{fl['lowest']['name']}</b> the largest net outflow "
                f"({commify(fl['lowest']['v'])}) (derived, SRC-620-02)."
            )
        if us.get("highest") and us.get("lowest"):
            bits.append(
                f"Among {commify(us.get('n_counties') or 0)} U.S. counties, "
                f"<b>{us['highest']['name']}</b> had the largest net inflow "
                f"({commify(us['highest']['v'])} returns) and "
                f"<b>{us['lowest']['name']}</b> the largest net outflow "
                f"({commify(us['lowest']['v'])}) (derived, SRC-620-02)."
            )
        pairs = sec.get("state_pair_flows_2022_23") or {}
        ma_fl = pairs.get("ma_to_fl") or {}
        top = (pairs.get("ma_out_top") or [{}])[0]
        if ma_fl.get("returns") is not None and top.get("name"):
            bits.append(
                f"The largest destination for Massachusetts filers was "
                f"<b>{top['name']}</b> at <b>{commify(top['returns'])}</b> "
                f"returns (SRC-620-01)."
                if top.get("st") == "FL"
                else
                f"The largest destination for Massachusetts filers was "
                f"<b>{top['name']}</b> at <b>{commify(top['returns'])}</b> "
                f"returns. Massachusetts to Florida was "
                f"<b>{commify(ma_fl['returns'])}</b> returns (SRC-620-01)."
            )
        return " ".join(bits)
    if tool_id == "DL-24":
        s = sec["seds_consumption_2024"]
        return (
            f"Total energy consumption was <b>{commify(s['us'])}</b> billion Btu "
            f"in the United States in 2024 (SRC-624-02). Massachusetts consumed "
            f"<b>{commify(s['ma']['v'])}</b> billion Btu, rank {s['ma']['rank']} "
            f"of {s['ma']['n']} (derived, SRC-624-02)."
        )
    if tool_id == "DL-27":
        b = sec["boston_operating_budget_fy26"]
        return (
            f"The FY26 adopted operating appropriation was "
            f"<b>{usd_prose(b['fy26_appropriation'])}</b> "
            f"(SRC-627-02). <b>{b['highest']['name']}</b> was the largest "
            f"department at <b>{usd_prose(b['highest']['v'])}</b> (SRC-627-02)."
        )
    return ""


STRIP_PHRASES = {
    "DL-06": [],
    "DL-07": [
        "NAEP scores and discipline files are pending on this page.",
        "NAEP, completion, and discipline remain pending.",
    ],
    "DL-08": [
        "Admissions-test and faculty files are pending.",
        "Admissions tests, faculty, and IPEDS outcomes remain pending.",
        "State faculty counts are not in that national table.",
        "A state-level faculty table is not in the current Digest xlsx set.",
        "That Digest table is national; it has no state column.",
    ],
    "DL-09": [
        "Staff counts are pending on this page.",
        "Education-staff files remain pending.",
    ],
    "DL-14": [],
    "DL-15": [
        "Personal income and NAICS detail are pending on this page.",
        "Personal income remains pending.",
        "Large states lead on raw output; personal income is pending.",
    ],
    "DL-16": [],
    "DL-20": [
        "County-to-county files are pending.",
        "County files remain pending.",
    ],
    "DL-24": [
        "EIA SEDS consumption and production remain pending.",
        "consumption and production remain pending.",
        "Consumption and production from SEDS are pending.",
        "remain pending because those files were not reachable this pass.",
    ],
    "DL-27": [
        "The adopted budget is pending.",
        "Adopted-budget views are pending on this page.",
        "The adopted-budget file remains pending.",
    ],
}


def enrich(app, ledger):
    """Attach verified later views to a live ledger and rewrite pending copy."""
    if ledger.get("status") != "live":
        return ledger
    tid = app["id"]
    if tid not in SECONDARY and tid not in MORE_SECONDARY:
        return ledger
    sec = {}
    if tid in SECONDARY:
        sec.update(SECONDARY[tid]())
    if tid in MORE_SECONDARY:
        sec.update(MORE_SECONDARY[tid]())
    derived = ledger.setdefault("derived", {})
    existing = derived.get("secondary") or {}
    existing.update(sec)
    derived["secondary"] = existing
    for phrase in list(STRIP_PHRASES.get(tid, [])) + list(MORE_STRIP.get(tid, [])):
        if ledger.get("lead"):
            ledger["lead"] = ledger["lead"].replace(phrase, "")
        if ledger.get("vintage_note"):
            ledger["vintage_note"] = ledger["vintage_note"].replace(phrase, "")
        for k in ledger.get("kpis") or []:
            for field in ("detail", "why"):
                if k.get(field):
                    k[field] = k[field].replace(phrase, "").strip()
    appendix = " ".join(
        p for p in (lead_appendix(tid, sec), more_lead(tid, sec)) if p
    ).strip()
    if appendix:
        lead = " ".join((ledger.get("lead") or "").split())
        marker = appendix[:56]
        if marker and marker in lead:
            lead = lead[:lead.index(marker)].strip()
        if appendix not in lead:
            ledger["lead"] = (lead + " " + appendix).strip()
    if tid == "DL-06" and "ma_chapter74_cte" in sec:
        v = sec["ma_chapter74_cte"]
        kpis = list(ledger.get("kpis") or [])
        kpis.append(_kpi(
            "MA Chapter 74 CTE, 2025-26",
            commify(v["v"]),
            (
                f"{v['share'] * 100:.1f} percent of high-school enrollment. "
                f"{v['districts']} districts, {v['schools']} schools "
                f"(SRC-606-03)."
            ),
            "The vocational-technical stock the research portfolio asked for.",
            "DESE / E2C pathways enrollment (SRC-606-03)",
        ))
        ledger["kpis"] = kpis[:2] + [kpis[-1]]
    if tid == "DL-13" and "bed_births_deaths" in sec:
        b = sec["bed_births_deaths"]
        ma = b.get("ma") or {}
        fl = b.get("fl") or {}
        kpis = list(ledger.get("kpis") or [])
        ma_label = f"MA establishment birth rate, {ma.get('births_as_of')}"
        fl_label = f"FL establishment birth rate, {fl.get('births_as_of')}"
        kpis = [k for k in kpis if k.get("label") not in (ma_label, fl_label)]
        kpis.append(_kpi(
            ma_label,
            f"{ma.get('birth_rate_pct')}%",
            (
                f"{commify(ma.get('births') or 0)} private-sector births "
                f"(SRC-613-02). Deaths are published through "
                f"{ma.get('deaths_as_of')} at {ma.get('death_rate_pct')} percent "
                f"({commify(ma.get('deaths') or 0)} establishments)."
            ),
            "The birth-versus-death rate the formation page now charts.",
            "BLS Business Employment Dynamics (SRC-613-02)",
        ))
        if fl.get("birth_rate_pct") is not None:
            kpis.append(_kpi(
                fl_label,
                f"{fl.get('birth_rate_pct')}%",
                (
                    f"{commify(fl.get('births') or 0)} private-sector births "
                    f"(SRC-613-02). Deaths are published through "
                    f"{fl.get('deaths_as_of')} at {fl.get('death_rate_pct')} percent "
                    f"({commify(fl.get('deaths') or 0)} establishments)."
                ),
                "Florida on the same BLS establishment birth-rate series.",
                "BLS Business Employment Dynamics (SRC-613-02)",
            ))
        ledger["kpis"] = kpis
    if tid == "DL-06" and "mcas_2025" in sec:
        m = sec["mcas_2025"]
        kpis = list(ledger.get("kpis") or [])
        kpis.append(_kpi(
            "MCAS grades 3-8 ELA, 2025",
            f"{m['ela_3_8_pct']}%",
            (
                f"Met or exceeded expectations. Math was {m['math_3_8_pct']}% "
                f"(SRC-606-04)."
            ),
            "The statewide Next Generation MCAS stock.",
            "DESE / E2C Next Generation MCAS (SRC-606-04)",
        ))
        ledger["kpis"] = kpis
    if tid == "DL-13":
        bed = sec.get("bed_births_deaths") or {}
        states = bed.get("states") or {}
        birth, death = {}, {}
        for st, series in states.items():
            b, d = [], []
            for p in series or []:
                q = p.get("q")
                if p.get("birth_rate_pct") is not None and q:
                    b.append({"q": q, "v": p["birth_rate_pct"]})
                if p.get("death_rate_pct") is not None and q:
                    d.append({"q": q, "v": p["death_rate_pct"]})
            if len(b) >= 2:
                birth[st] = b
            if len(d) >= 2:
                death[st] = d
        wins = {}
        if birth:
            wins.update(windows_from_trend(
                birth, src="SRC-613-02", unit="percent",
                ns=(4, 9), label_stem="Establishment birth rate",
                named_ends=["2024 Q3"], prefix="bed_birth_rate",
            ))
        if death:
            wins.update(windows_from_trend(
                death, src="SRC-613-02", unit="percent",
                ns=(4, 9), label_stem="Establishment death rate",
                named_ends=["2024 Q3"], prefix="bed_death_rate",
            ))
        attach_windows(
            ledger, wins,
            note="Prefer these over recomputing. BED window means and ranks cite (derived, SRC-613-02).",
        )
        w9 = wins.get("bed_birth_rate_t9_2024q3")
        if w9:
            bed["window_9q_2024q3"] = {
                "ma": w9.get("ma"), "us": w9.get("us"), "fl": w9.get("fl"),
                "highest": w9.get("highest"), "lowest": w9.get("lowest"),
                "end": w9.get("end"), "n_periods": 9,
            }
    extra_note = (
        f" Later views compiled {REVISED} are stored under derived.secondary."
    )
    ledger["vintage_note"] = (
        " ".join((ledger.get("vintage_note") or "").split()) + extra_note
    ).strip()
    return ledger


# ---------------------------------------------------------------------------
# New live builders (former stubs)
# ---------------------------------------------------------------------------

def build_hospitals(app):
    """DL-10: CMS Hospital General Information for Massachusetts facilities."""
    body = json.dumps({
        "conditions": [
            {"resource": "t", "property": "state", "value": "MA", "operator": "="}
        ],
        "limit": 200,
        "offset": 0,
    }).encode()
    req = urllib.request.Request(
        CMS_QUERY,
        data=body,
        headers={"User-Agent": UA, "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        payload = json.loads(resp.read())
    rows = payload.get("results") or []
    if payload.get("count") != VERIFY_CMS_MA_HOSPITALS or len(rows) != VERIFY_CMS_MA_HOSPITALS:
        sys.exit(
            f"FATAL: CMS MA hospital count is {payload.get('count')} / {len(rows)}, "
            f"expected {VERIFY_CMS_MA_HOSPITALS}"
        )
    rated = {}
    meta = {}
    by_type = defaultdict(int)
    by_own = defaultdict(int)
    stars = defaultdict(int)
    emergency = 0
    for r in rows:
        name = (r.get("facility_name") or "").strip()
        if not name:
            continue
        rating = parse_num(r.get("hospital_overall_rating"))
        rec = {
            "city": r.get("citytown"),
            "type": r.get("hospital_type"),
            "ownership": r.get("hospital_ownership"),
            "emergency": r.get("emergency_services"),
            "rating": int(rating) if rating is not None else None,
        }
        meta[name] = rec
        by_type[rec["type"] or "Unspecified"] += 1
        by_own[rec["ownership"] or "Unspecified"] += 1
        if rec["emergency"] == "Yes":
            emergency += 1
        if rating is not None:
            rated[name] = rating
            stars[int(rating)] += 1
    ranked = rank_named(rated, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = int(rec["v"])
        rec.update({k: meta[rec["name"]].get(k) for k in ("city", "type", "ownership")})
    unrated = [n for n, m in meta.items() if m["rating"] is None]
    five = stars.get(5, 0)
    as_of = "2026-08"
    as_of_label = "CMS Hospital General Information, retrieved Aug 2026"
    kpis = [
        _kpi(
            "Massachusetts hospitals",
            commify(len(meta)),
            (
                f"{commify(len(rated))} have a CMS overall star rating; "
                f"{commify(len(unrated))} are unrated (SRC-610-02)."
            ),
            "The facility count behind the Massachusetts hospital tracker.",
            "CMS Hospital General Information (SRC-610-02)",
        ),
        _kpi(
            "Five-star ratings",
            commify(five),
            f"{emergency} facilities report emergency services (SRC-610-02).",
            "Star ratings are CMS overall ratings, not CHIA relative prices.",
            "CMS Hospital General Information (SRC-610-02)",
        ),
        _kpi(
            "Most common type",
            max(by_type, key=by_type.get),
            (
                f"{by_type[max(by_type, key=by_type.get)]} of {len(meta)} "
                f"facilities (SRC-610-02). CHIA relative prices remain pending."
            ),
            "Type mix is the first cut of the hospital file.",
            "CMS Hospital General Information (SRC-610-02)",
        ),
    ]
    lead = (
        f"CMS lists <b>{commify(len(meta))}</b> hospitals in Massachusetts "
        f"(SRC-610-02). <b>{commify(five)}</b> have a five-star overall rating; "
        f"{commify(len(rated))} have any star rating (SRC-610-02). CHIA relative "
        f"prices remain pending on this page."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from the CMS provider-data Hospital General "
            f"Information datastore (dataset xubh-q36u), filtered to state=MA. "
            f"Count equals {VERIFY_CMS_MA_HOSPITALS}. The ranking is the CMS "
            f"overall star rating (5 to 1) among rated facilities. CHIA "
            f"relative-price files remain pending."
        ),
        metric="cms_ma_hospital_overall_rating",
        metric_label="CMS overall hospital star rating, Massachusetts facilities",
        unit="star rating (1-5)",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "n_hospitals": len(meta),
            "n_rated": len(rated),
            "n_unrated": len(unrated),
            "five_star": five,
            "emergency_services": emergency,
            "by_type": dict(by_type),
            "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"], "city": ranked[0].get("city")},
        },
        src_note="SRC-610-02",
        extra={"derived": {"unrated_count": len(unrated), "by_ownership": dict(by_own)}},
    )


def build_transit(app):
    """DL-22: FTA NTD monthly unlinked passenger trips by agency, latest month."""
    latest = _soda(URL_NTD, {"$select": "max(date) as d"})
    date_s = (latest[0].get("d") or "")[:10]
    if date_s != "2026-06-01":
        sys.exit(f"FATAL: NTD latest month is {date_s}, expected 2026-06-01")
    us = _soda(URL_NTD, {
        "$select": "sum(upt) as upt",
        "$where": f"date='{date_s}T00:00:00.000'",
    })
    us_val = parse_num(us[0].get("upt"))
    if us_val is None or abs(us_val - VERIFY_NTD_US_JUN_2026) > 1:
        sys.exit(f"FATAL: NTD US June 2026 UPT is {us_val}")
    rows = _soda(URL_NTD, {
        "$select": "agency,state,sum(upt) as upt",
        "$where": f"date='{date_s}T00:00:00.000'",
        "$group": "agency,state",
        "$order": "upt DESC",
        "$limit": "5000",
    })
    values = {}
    states = {}
    for r in rows:
        name = (r.get("agency") or "").strip()
        st = (r.get("state") or "").strip()
        v = parse_num(r.get("upt"))
        if not name or v is None or v <= 0:
            continue
        # Disambiguate agencies that share a name across states.
        key = name if name not in values else f"{name} ({st})"
        values[key] = v
        states[key] = st
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = int(round(rec["v"]))
        rec["state"] = states.get(rec["name"])
    ma_rows = [r for r in ranked if r.get("state") == "MA"]
    if not ma_rows:
        sys.exit("FATAL: NTD June 2026 has no Massachusetts agency")
    mbta = next((r for r in ma_rows if "Bay Transportation" in r["name"]), ma_rows[0])
    ma_total = sum(r["v"] for r in ma_rows)
    as_of = "2026-06"
    as_of_label = "June 2026"
    kpis = [
        _kpi(
            "U.S. unlinked trips, June 2026",
            commify(us_val),
            f"{commify(len(ranked))} agencies reported trips (SRC-622-01).",
            "The national monthly ridership stock.",
            "FTA NTD complete monthly ridership (SRC-622-01)",
        ),
        _kpi(
            "MBTA",
            commify(mbta["v"]),
            (
                f"Rank {mbta['rank']} of {mbta['n']} U.S. agencies "
                f"(derived, SRC-622-01). Massachusetts agencies together "
                f"reported {commify(ma_total)} trips."
            ),
            "The MBTA is the Massachusetts comparison; mode-level views stay on DL-03.",
            "FTA NTD complete monthly ridership (SRC-622-01)",
        ),
        _kpi(
            "Largest U.S. agency",
            ranked[0]["name"],
            f"{commify(ranked[0]['v'])} unlinked trips (SRC-622-01).",
            "Agency-level comparison is the former Compare Systems view.",
            "FTA NTD complete monthly ridership (SRC-622-01)",
        ),
    ]
    lead = (
        f"U.S. transit agencies reported <b>{commify(us_val)}</b> unlinked "
        f"passenger trips in June 2026 (SRC-622-01). The MBTA reported "
        f"<b>{commify(mbta['v'])}</b>, rank {mbta['rank']} of {mbta['n']} "
        f"agencies (derived, SRC-622-01). <b>{ranked[0]['name']}</b> was "
        f"the largest agency (SRC-622-01)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from FTA NTD Complete Monthly Ridership "
            f"(Socrata 8bui-9xvu). Headline is unlinked passenger trips "
            f"summed by agency for {as_of_label}. U.S. total equals "
            f"{VERIFY_NTD_US_JUN_2026:,}. Mode-by-mode MBTA reliability "
            f"and cost stay on DL-03."
        ),
        metric="ntd_agency_upt_latest_month",
        metric_label="Unlinked passenger trips by transit agency, June 2026",
        unit="unlinked passenger trips",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "us": {"v": int(us_val)},
            "mbta": {"name": mbta["name"], "v": mbta["v"], "rank": mbta["rank"], "n": mbta["n"]},
            "ma_agencies": len(ma_rows),
            "ma_total": ma_total,
            "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"], "state": ranked[0].get("state")},
        },
        src_note="SRC-622-01",
        extra={"derived": {"massachusetts_agencies": [
            {"name": r["name"], "v": r["v"], "rank": r["rank"]} for r in ma_rows
        ]}},
    )


def build_payroll(app):
    """DL-30: CTHRU Commonwealth payroll by department, calendar 2025."""
    years = _soda("https://cthru.data.socrata.com/resource/9ttk-7vz6.json", {
        "$select": "year,count(*) as n,sum(pay_total_actual) as pay",
        "$group": "year",
        "$order": "year",
    })
    y2025 = next((r for r in years if str(r.get("year")) == "2025"), None)
    if not y2025:
        sys.exit("FATAL: CTHRU payroll missing calendar 2025")
    n_emp = int(float(y2025["n"]))
    total = parse_num(y2025["pay"])
    if total is None or n_emp < 100000 or total < 1e10:
        sys.exit(f"FATAL: CTHRU 2025 payroll n={n_emp} pay={total}")
    depts = _soda("https://cthru.data.socrata.com/resource/9ttk-7vz6.json", {
        "$select": "department_division,count(*) as n,sum(pay_total_actual) as pay",
        "$where": "year='2025'",
        "$group": "department_division",
        "$order": "pay DESC",
        "$limit": "5000",
    })
    values = {}
    counts = {}
    for r in depts:
        name = (r.get("department_division") or "").strip()
        v = parse_num(r.get("pay"))
        if not name or v is None:
            continue
        values[name] = v
        counts[name] = int(float(r.get("n") or 0))
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["employees"] = counts.get(rec["name"], 0)
    hi = ranked[0]
    # Vendor / all-object spending, last complete fiscal year.
    spend_years = _soda(URL_SPEND, {
        "$select": "budget_fiscal_year,sum(amount) as amt",
        "$group": "budget_fiscal_year",
        "$order": "budget_fiscal_year",
    })
    fy2025 = next((r for r in spend_years if str(r.get("budget_fiscal_year")) == "2025"), None)
    spend = parse_num(fy2025["amt"]) if fy2025 else None
    if spend is None or spend < 1e10:
        sys.exit(f"FATAL: CTHRU spending FY2025 is {spend}")
    as_of = "2025-12"
    as_of_label = "Calendar year 2025"
    kpis = [
        _kpi(
            "Commonwealth payroll, 2025",
            usd_prose(total),
            f"{commify(n_emp)} employee rows across {commify(len(ranked))} departments (SRC-630-01).",
            "The statewide earnings stock, excluding municipal payroll.",
            "CTHRU Commonwealth payroll v4 (SRC-630-01)",
        ),
        _kpi(
            "Largest department",
            hi["name"],
            f"{usd_prose(hi['v'])} across {commify(hi['employees'])} employee rows (SRC-630-01).",
            "Higher education and transportation dominate the raw dollar ranking.",
            "CTHRU Commonwealth payroll v4 (SRC-630-01)",
        ),
        _kpi(
            "CTHRU recorded spending, FY2025",
            usd_prose(spend),
            (
                "All object classes in the Comptroller spending file, including "
                "payroll transfers. This is not a vendor-only extract (SRC-630-02)."
            ),
            "The spending file is the former vendor-payments view, published as a total.",
            "CTHRU Comptroller spending (SRC-630-02)",
        ),
    ]
    lead = (
        f"Commonwealth payroll totaled <b>{usd_prose(total)}</b> in calendar "
        f"2025 across <b>{commify(n_emp)}</b> employee rows (SRC-630-01). "
        f"<b>{hi['name']}</b> was the largest department at "
        f"<b>{usd_prose(hi['v'])}</b> (SRC-630-01). Comptroller-recorded "
        f"spending was <b>{usd_prose(spend)}</b> in fiscal 2025 (SRC-630-02)."
    )
    return finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt {REVISED} from CTHRU Commonwealth Of Massachusetts Payroll "
            f"v4 (Socrata 9ttk-7vz6), calendar 2025, summed by department_division. "
            f"Named employees are not published here; named House and Senate pay "
            f"sits on Legislature Pay. Spending is the Comptroller "
            f"file pegc-naaa, budget fiscal year 2025, all object classes. "
            f"Calendar 2026 payroll is year-to-date and is not the headline."
        ),
        metric="cthru_department_payroll_2025",
        metric_label="Commonwealth payroll by department, calendar 2025",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "total": total,
            "employees": n_emp,
            "departments": len(ranked),
            "highest": {"name": hi["name"], "v": hi["v"], "employees": hi["employees"]},
            "spending_fy2025": spend,
        },
        src_note="SRC-630-01",
        extra={"derived": {
            "spending_fy2025": {
                "v": spend,
                "src": "SRC-630-02",
                "label": "CTHRU Comptroller spending, all object classes, FY2025",
            }
        }},
    )


# Two-path checks against a prior complete pull of calendar 2025.
VERIFY_LEG_2025_TOTAL = 26508845.89
VERIFY_LEG_2025_TOP = 223719.88
VERIFY_LEG_HOUSE_N = 179
VERIFY_LEG_SENATE_N = 43
VERIFY_LEG_HOUSE_MEDIAN = 111864.14
VERIFY_LEG_SENATE_MEDIAN = 156726.07


def _median(vals):
    nums = sorted(v for v in vals if v is not None)
    n = len(nums)
    if n == 0:
        return None
    if n % 2:
        return nums[n // 2]
    return (nums[n // 2 - 1] + nums[n // 2]) / 2


def _title_name(part):
    text = (part or "").strip()
    if not text:
        return ""
    if text.isupper() or text.islower():
        return text.title()
    return text


def _leg_pay(row, key):
    v = parse_num(row.get(key))
    return 0.0 if v is None else float(v)


def build_legislature_pay(app):
    """DL-32: named House and Senate pay from CTHRU, calendar 2025."""
    raw = _soda(URL_CTHRU, {
        "$select": (
            "name_first,name_last,department_code,position_title,"
            "pay_base_actual,aa1,a14,pay_other_actual,pay_total_actual,"
            "annual_rate,year"
        ),
        "$where": (
            "year='2025' AND department_code in('HOU','SEN') "
            "AND position_title in('Representative','Senator')"
        ),
        "$limit": "5000",
    })
    if not raw:
        sys.exit("FATAL: CTHRU legislator payroll returned no calendar 2025 rows")
    house = [
        r for r in raw
        if r.get("department_code") == "HOU" and r.get("position_title") == "Representative"
    ]
    senate = [
        r for r in raw
        if r.get("department_code") == "SEN" and r.get("position_title") == "Senator"
    ]
    if len(house) != VERIFY_LEG_HOUSE_N:
        sys.exit(f"FATAL: House Representative rows {len(house)}, expected {VERIFY_LEG_HOUSE_N}")
    if len(senate) != VERIFY_LEG_SENATE_N:
        sys.exit(f"FATAL: Senate Senator rows {len(senate)}, expected {VERIFY_LEG_SENATE_N}")

    total = sum(_leg_pay(r, "pay_total_actual") for r in raw)
    base = sum(_leg_pay(r, "pay_base_actual") for r in raw)
    a14 = sum(_leg_pay(r, "a14") for r in raw)
    aa1 = sum(_leg_pay(r, "aa1") for r in raw)
    if abs(total - VERIFY_LEG_2025_TOTAL) > 0.05:
        sys.exit(f"FATAL: legislator 2025 total {total}, expected {VERIFY_LEG_2025_TOTAL}")

    house_pays = [_leg_pay(r, "pay_total_actual") for r in house]
    senate_pays = [_leg_pay(r, "pay_total_actual") for r in senate]
    house_med = _median(house_pays)
    senate_med = _median(senate_pays)
    if house_med is None or abs(house_med - VERIFY_LEG_HOUSE_MEDIAN) > 0.05:
        sys.exit(f"FATAL: House median {house_med}, expected {VERIFY_LEG_HOUSE_MEDIAN}")
    if senate_med is None or abs(senate_med - VERIFY_LEG_SENATE_MEDIAN) > 0.05:
        sys.exit(f"FATAL: Senate median {senate_med}, expected {VERIFY_LEG_SENATE_MEDIAN}")

    grouped = defaultdict(list)
    for r in raw:
        key = (
            (r.get("name_first") or "").strip().lower(),
            (r.get("name_last") or "").strip().lower(),
        )
        grouped[key].append(r)

    display_used = set()
    values = {}
    extras = {}
    for recs in grouped.values():
        first = _title_name(recs[0].get("name_first"))
        last = _title_name(recs[0].get("name_last"))
        name = " ".join(p for p in (first, last) if p)
        if name in display_used:
            sys.exit(f"FATAL: duplicate legislator display name {name}")
        display_used.add(name)
        tot = sum(_leg_pay(r, "pay_total_actual") for r in recs)
        base_p = sum(_leg_pay(r, "pay_base_actual") for r in recs)
        a14_p = sum(_leg_pay(r, "a14") for r in recs)
        aa1_p = sum(_leg_pay(r, "aa1") for r in recs)
        depts = {r.get("department_code") for r in recs}
        titles = {r.get("position_title") for r in recs}
        primary = max(recs, key=lambda r: _leg_pay(r, "pay_total_actual"))
        if depts == {"HOU", "SEN"} or titles == {"Representative", "Senator"}:
            chamber = "House and Senate"
            title = "Representative and Senator"
        elif primary.get("department_code") == "SEN" or primary.get("position_title") == "Senator":
            chamber = "Senate"
            title = "Senator"
        else:
            chamber = "House"
            title = "Representative"
        values[name] = tot
        extras[name] = {
            "base": round(base_p, 2),
            "a14": round(a14_p, 2),
            "aa1": round(aa1_p, 2),
            "chamber": chamber,
            "title": title,
            "first": first,
            "last": last,
            "n_stints": len(recs),
            "annual_rate": round(max(_leg_pay(r, "annual_rate") for r in recs), 2),
        }

    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: extras[n]["last"][:8])
    for rec in ranked:
        rec.update(extras[rec["name"]])
        rec["v"] = round(rec["v"], 2)

    top = ranked[0]
    if abs(top["v"] - VERIFY_LEG_2025_TOP) > 0.05:
        sys.exit(f"FATAL: top legislator pay {top['v']}, expected {VERIFY_LEG_2025_TOP}")
    tied = [r for r in ranked if abs(r["v"] - VERIFY_LEG_2025_TOP) <= 0.05]
    if len(tied) < 2:
        sys.exit(f"FATAL: expected Speaker and Senate President tie at {VERIFY_LEG_2025_TOP}")

    y2026 = _soda(URL_CTHRU, {
        "$select": "count(*) as n,sum(pay_total_actual) as pay",
        "$where": (
            "year='2026' AND department_code in('HOU','SEN') "
            "AND position_title in('Representative','Senator')"
        ),
    })
    ytd_n = int(float((y2026[0] or {}).get("n") or 0)) if y2026 else 0
    ytd_pay = parse_num((y2026[0] or {}).get("pay")) if y2026 else None

    as_of = "2025-12"
    as_of_label = "Calendar year 2025"
    hi_names = " and ".join(r["name"] for r in tied[:2])
    kpis = [
        _kpi(
            "Highest 2025 pay",
            usd_prose(VERIFY_LEG_2025_TOP),
            f"{hi_names}, the House Speaker and Senate President (SRC-632-01).",
            "Leadership extras sit in the Comptroller supplemental (AA1) bucket.",
            "CTHRU named House and Senate payroll (SRC-632-01)",
        ),
        _kpi(
            "House median, 2025",
            usd_prose(house_med),
            f"{commify(len(house))} Representative rows on the House payroll (SRC-632-01).",
            "The typical House check, including partial-year replacements.",
            "CTHRU named House and Senate payroll (SRC-632-01)",
        ),
        _kpi(
            "Senate median, 2025",
            usd_prose(senate_med),
            f"{commify(len(senate))} Senator rows on the Senate payroll (SRC-632-01).",
            "The typical Senate check, including partial-year replacements.",
            "CTHRU named House and Senate payroll (SRC-632-01)",
        ),
    ]
    lead = (
        f"People paid as a Massachusetts Representative or Senator received "
        f"<b>{usd_prose(total)}</b> in calendar 2025 (SRC-632-01). "
        f"<b>{hi_names}</b> each received <b>{usd_prose(VERIFY_LEG_2025_TOP)}</b>, "
        f"the House Speaker and Senate President totals (SRC-632-01). "
        f"The House median was <b>{usd_prose(house_med)}</b>; the Senate median "
        f"was <b>{usd_prose(senate_med)}</b> (SRC-632-01). Employer-paid health "
        f"and pension contributions are not named-employee lines on this file."
    )
    components = [
        {"name": "Base salary", "v": round(base, 2)},
        {"name": "Supplemental (AA1)", "v": round(aa1, 2)},
        {"name": "Stipends (A14)", "v": round(a14, 2)},
    ]
    ledger = finish_live(
        app,
        as_of=as_of,
        as_of_label=as_of_label,
        vintage_note=(
            f"Rebuilt Aug 16, 2026 from CTHRU Commonwealth Of Massachusetts Payroll "
            f"v4 (Socrata 9ttk-7vz6), calendar 2025. Rows are people with "
            f"department_code HOU or SEN and position_title Representative or "
            f"Senator. Directory rows collapse one person across chambers. "
            f"{len(raw)} payroll rows become {len(ranked)} people because of "
            f"mid-year replacements and chamber switches. Total pay equals "
            f"base plus AA1 (Comptroller Salaries, supplemental) plus A14 "
            f"(Stipends, bonus pay, and awards). Buyout and overtime are $0 "
            f"on these rows. Employer-paid GIC and MSERS amounts are not on "
            f"this file. Calendar 2026 is year-to-date "
            f"({commify(ytd_n)} rows"
            + (f", {usd_prose(ytd_pay)}" if ytd_pay is not None else "")
            + ") and is not the headline."
        ),
        metric="ma_legislator_pay_2025",
        metric_label="Massachusetts legislator pay, calendar 2025",
        unit="dollars",
        lead=lead,
        kpis=kpis,
        ranked=ranked,
        trend={},
        latest={
            "total": round(total, 2),
            "base": round(base, 2),
            "a14": round(a14, 2),
            "aa1": round(aa1, 2),
            "n_people": len(ranked),
            "n_rows": len(raw),
            "house": {
                "n": len(house),
                "median": house_med,
                "total": round(sum(house_pays), 2),
            },
            "senate": {
                "n": len(senate),
                "median": senate_med,
                "total": round(sum(senate_pays), 2),
            },
            "highest": [
                {"name": r["name"], "v": r["v"], "chamber": r["chamber"], "title": r["title"]}
                for r in tied
            ],
            "components": components,
            "ytd_2026": {"n": ytd_n, "pay": ytd_pay},
        },
        src_note="SRC-632-01",
        extra={"derived": {
            "components": components,
            "chamber_medians": [
                {"name": "House median", "v": house_med},
                {"name": "Senate median", "v": senate_med},
            ],
        }},
    )
    ledger["page"]["revised"] = "Aug 16, 2026"
    return ledger


BUILDERS = {
    "DL-10": build_hospitals,
    "DL-22": build_transit,
    "DL-30": build_payroll,
    "DL-32": build_legislature_pay,
}
