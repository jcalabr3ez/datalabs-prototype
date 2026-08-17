#!/usr/bin/env python3
"""Public later views that finish the live suite pages.

Each helper fetches a published file and returns verified figures only.
Series whose files are blocked, PDF-only, or Pioneer-method-only stay out.
"""
from __future__ import annotations

import csv
import io
import json
import sys
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict

from openpyxl import load_workbook

from suite_hollow import hollow_lead, hollow_secondary
from suite_substance import (
    sec_aslg_2022,
    sec_aspp_2025,
    sec_boston_earners,
    sec_discipline_race,
    sec_expulsion,
    sec_gov_orgs,
    sec_he_finance,
    sec_he_students,
    sec_ipeds_grad_by_state,
    sec_k12_staff,
    sec_ma_demographics,
    sec_pop_age_race,
    sec_pop_components,
    sec_public_he_faculty,
    sec_qcew_employment,
)
from suite_common import (
    LEDGER_DIR,
    RANKED,
    STATE_NAMES,
    ST_TO_FIPS,
    UA,
    commify,
    fetch,
    fetch_text,
    fl_cell,
    geo_to_st,
    parse_num,
    pct,
    rank_named,
    rank_rows,
    snap_pack,
    usd_prose,
    yoy_pct,
)

E2C_MCAS = "https://educationtocareer.data.mass.gov/resource/i9w6-niyt.json"
E2C_ATTEND = "https://educationtocareer.data.mass.gov/resource/ak6h-9k7x.json"
E2C_DROPOUT = "https://educationtocareer.data.mass.gov/resource/cmm7-ttbg.json"
E2C_FINANCE = "https://educationtocareer.data.mass.gov/resource/er3w-dyti.json"
URL_NAEP = "https://www.nationsreportcard.gov/Dataservice/GetAdhocData.aspx"
DIGEST_326 = "https://nces.ed.gov/programs/digest/d24/tables/xls/tabn326.10.xlsx"
URL_OIG_MFCU = (
    "https://oig.hhs.gov/documents/evaluation/11556/"
    "Statistical_Chart_-_2025.xlsx"
)
URL_UI = "https://oui.doleta.gov/unemploy/csv/ar539.csv"
URL_SAGDP = "https://apps.bea.gov/regional/zip/SAGDP.zip"
URL_CS_BOS = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=BOXRSA"
URL_CS_MIA = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=MIXRSA"
URL_RUCC = (
    "https://www.ers.usda.gov/media/5768/2023-rural-urban-continuum-codes.csv"
    "?v=23323"
)
URL_IRS_COUNTY = "https://www.irs.gov/pub/irs-soi/22incyallagi.csv"
URL_NTD_ANN = "https://data.transportation.gov/resource/ekg5-frzt.json"
URL_OPENFEMA = (
    "https://www.fema.gov/api/open/v2/PublicAssistanceGrantAwardActivities.csv"
)
URL_NRI = (
    "https://services.arcgis.com/XG15cJAlne2vxtgt/arcgis/rest/services/"
    "National_Risk_Index_Counties/FeatureServer/0/query"
)
URL_NOAA = (
    "https://www.ncei.noaa.gov/access/monitoring/climate-at-a-glance/"
    "statewide/time-series/{code}/{var}/all/12/2000-2024.csv"
)
URL_SEDS_PROD = "https://www.eia.gov/state/seds/sep_prod/xls/Prod_dataset.xlsx"
URL_QUASI = "https://cthru.data.socrata.com/resource/tc5d-8ckm.json"
URL_SPEND = "https://cthru.data.socrata.com/resource/pegc-naaa.json"
URL_SUBEST = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2020-2025/cities/totals/sub-est2025_25.csv"
)

VERIFY_OIG_US_RECOVERIES_FY2025 = 1966591061.26
VERIFY_OIG_MA_RECOVERIES_FY2025 = 139892852.19
VERIFY_MCAS_ELA_38_2025 = 0.42
VERIFY_MCAS_MATH_38_2025 = 0.41
VERIFY_DROPOUT_MA_2025 = 0.018
VERIFY_IPEDS_6YR_2017 = 64.6  # Digest 326.10, all 4-year institutions
VERIFY_NAEP_NP_READ4_2024_MIN = 180
VERIFY_NAEP_NP_READ4_2024_MAX = 250
VERIFY_NAEP_NP_READ4_2019 = 219.4
VERIFY_NAEP_NP_READ4_2024 = 214.3
VERIFY_NAEP_LA_READ4_CHANGE_2019_2024 = 6.1
# BLS CEWBD 2025 Q4 news release (July 29, 2026): 338,000 births;
# 284,000 deaths in 2025 Q1. The national level series is in thousands.
VERIFY_US_BED_BIRTHS_THOUSANDS_2025Q4 = 338
VERIFY_US_BED_DEATHS_THOUSANDS_2025Q1 = 284
BLS_API = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
# BD + S + msa 00000 + FIPS + county 000 + industry 000000 (total private)
# + unit 1 + element 2 (establishments) + size 00 + class 07/08 + L/R + Q + 5
BD_US_BIRTHS_L = "BDS0000000000000000120007LQ5"
BD_US_DEATHS_L = "BDS0000000000000000120008LQ5"
BD_MA_BIRTHS_L = "BDS0000025000000000120007LQ5"
BD_MA_DEATHS_L = "BDS0000025000000000120008LQ5"
BD_US_BIRTHS_R = "BDS0000000000000000120007RQ5"
BD_US_DEATHS_R = "BDS0000000000000000120008RQ5"
BD_MA_BIRTHS_R = "BDS0000025000000000120007RQ5"
BD_MA_DEATHS_R = "BDS0000025000000000120008RQ5"
BD_FL_BIRTHS_L = "BDS0000012000000000120007LQ5"
BD_FL_DEATHS_L = "BDS0000012000000000120008LQ5"
BD_FL_BIRTHS_R = "BDS0000012000000000120007RQ5"
BD_FL_DEATHS_R = "BDS0000012000000000120008RQ5"


def _soda(url, params):
    q = urllib.parse.urlencode(params)
    return json.loads(fetch(url + "?" + q, timeout=120))


def _ma(ranked):
    rec = next((r for r in ranked if r.get("st") == "MA"), None)
    if not rec:
        sys.exit("FATAL: ranking is missing Massachusetts")
    return rec


def _snap(values, us_val, round_to=None):
    return snap_pack(values, us_val, round_to=round_to)


def _as_pct(v):
    if v is None:
        return None
    return round(v * 100, 1) if v <= 1.5 else round(v, 1)


# ---------------------------------------------------------------------------
# Education
# ---------------------------------------------------------------------------

def sec_public_k12_enrollment():
    """Fall enrollment from Digest 203.20, already verified on State School Scores."""
    path = LEDGER_DIR / "dl07-answers.json"
    dl07 = json.loads(path.read_text())
    latest = dl07.get("latest") or {}
    ma = latest.get("ma") or {}
    trend = (dl07.get("trend") or {}).get("MA") or []
    rows = [
        {
            "st": r["st"],
            "name": r.get("name") or r["st"],
            "v": r["v"],
            "rank": r.get("rank"),
            "n": r.get("n") or ma.get("n"),
        }
        for r in (dl07.get("rows") or [])
        if r.get("st") and r.get("v") is not None
    ]
    fl = next((r for r in rows if r.get("st") == "FL"), None)
    if not ma.get("v") or len(trend) < 2:
        return None
    return {
        "label": "Massachusetts public K-12 enrollment",
        "src": "SRC-606-02",
        "unit": "students",
        "as_of_label": "Fall 2024",
        "us": (latest.get("us") or {}).get("v"),
        "ma": {
            "v": ma.get("v"),
            "rank": ma.get("rank"),
            "n": ma.get("n"),
            "name": "Massachusetts",
        },
        "fl": (
            {"v": fl["v"], "rank": fl["rank"], "n": fl["n"], "name": fl["name"]}
            if fl else None
        ),
        "highest": latest.get("highest"),
        "lowest": latest.get("lowest"),
        "n_ranked": ma.get("n") or len(rows),
        "rows": rows,
        "trend": trend,
        "note": (
            "NCES Digest table 203.20, public elementary and secondary "
            "enrollment. Same published cells as the State School Scores ranking."
        ),
    }


def sec_mcas():
    rows = _soda(E2C_MCAS, {
        "$where": (
            "sy='2025' AND org_type='State' AND stu_grp='All Students' "
            "AND test_grade='ALL (03-08)'"
        ),
        "$limit": "10",
    })
    by_subj = {r.get("subject_code"): r for r in rows}
    ela = parse_num((by_subj.get("ELA") or {}).get("m_plus_e_pct"))
    math = parse_num((by_subj.get("MATH") or {}).get("m_plus_e_pct"))
    if ela is None or abs(ela - VERIFY_MCAS_ELA_38_2025) > 0.005:
        sys.exit(f"FATAL: MCAS 2025 state ELA 3-8 meets/exceeds is {ela}")
    if math is None or abs(math - VERIFY_MCAS_MATH_38_2025) > 0.005:
        sys.exit(f"FATAL: MCAS 2025 state math 3-8 meets/exceeds is {math}")
    g10 = _soda(E2C_MCAS, {
        "$where": (
            "sy='2025' AND org_type='State' AND stu_grp='All Students' "
            "AND test_grade='10'"
        ),
        "$limit": "10",
    })
    g10s = {r.get("subject_code"): parse_num(r.get("m_plus_e_pct")) for r in g10}
    dist = _soda(E2C_MCAS, {
        "$select": "dist_name,m_plus_e_pct,stu_cnt",
        "$where": (
            "sy='2025' AND org_type='Public School District' "
            "AND stu_grp='All Students' AND test_grade='ALL (03-08)' "
            "AND subject_code='ELA' AND m_plus_e_pct IS NOT NULL"
        ),
        "$order": "m_plus_e_pct DESC",
        "$limit": "400",
    })
    values = {}
    for r in dist:
        name = (r.get("dist_name") or "").strip()
        v = parse_num(r.get("m_plus_e_pct"))
        n = parse_num(r.get("stu_cnt"))
        if name and v is not None and n and n >= 100:
            values[name] = v
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = _as_pct(rec["v"])
    return {
        "label": "Next Generation MCAS, share meeting or exceeding, 2025",
        "src": "SRC-606-04",
        "unit": "percent",
        "as_of_label": "Spring 2025",
        "ela_3_8_pct": _as_pct(ela),
        "math_3_8_pct": _as_pct(math),
        "ela_10_pct": _as_pct(g10s.get("ELA")),
        "math_10_pct": _as_pct(g10s.get("MATH")),
        "highest_district_ela_3_8": {
            "name": ranked[0]["name"], "v": ranked[0]["v"]
        } if ranked else None,
        "lowest_district_ela_3_8": {
            "name": ranked[-1]["name"], "v": ranked[-1]["v"]
        } if ranked else None,
        "district_n": len(ranked),
        "note": "Statewide All Students. District ranking is ELA grades 3-8, districts with at least 100 tested students.",
    }


def sec_attendance():
    rows = _soda(E2C_ATTEND, {
        "$where": (
            "sy='2025' AND attend_period='End of Year' "
            "AND org_type='State' AND stu_grp='All Students'"
        ),
        "$limit": "2",
    })
    if len(rows) != 1:
        sys.exit(f"FATAL: E2C attendance state EOY 2025 has {len(rows)} rows")
    r = rows[0]
    rate = parse_num(r.get("attend_rate"))
    chron = parse_num(r.get("pct_chron_abs_10"))
    if rate is None or not (0.8 < rate < 1.0):
        sys.exit(f"FATAL: E2C attendance rate 2025 is {rate}")
    return {
        "label": "Massachusetts attendance rate, school year 2024-25",
        "src": "SRC-606-05",
        "unit": "percent",
        "as_of_label": "School year 2024-25, end of year",
        "attend_rate_pct": _as_pct(rate),
        "chronic_absent_10_pct": _as_pct(chron),
    }


def sec_dropouts():
    rows = _soda(E2C_DROPOUT, {
        "$where": "sy='2025' AND org_type='State' AND stu_grp='All Students'",
        "$limit": "2",
    })
    if len(rows) != 1:
        sys.exit(f"FATAL: E2C dropout state 2025 has {len(rows)} rows")
    r = rows[0]
    rate = parse_num(r.get("drpout_pct_all"))
    n = parse_num(r.get("drpout_cnt_all"))
    enroll = parse_num(r.get("enroll_cnt_all"))
    if rate is None or abs(rate - VERIFY_DROPOUT_MA_2025) > 0.0005:
        sys.exit(f"FATAL: E2C dropout rate 2025 is {rate}")
    return {
        "label": "Massachusetts high-school dropout rate, 2024-25",
        "src": "SRC-606-06",
        "unit": "percent",
        "as_of_label": "School year 2024-25",
        "dropout_pct": _as_pct(rate),
        "dropout_count": int(n) if n is not None else None,
        "enroll_count": int(enroll) if enroll is not None else None,
    }


def sec_district_finance():
    rows = _soda(E2C_FINANCE, {
        "$where": (
            "sy='2025' AND ind_cat='Expenditures Per Pupil' "
            "AND ind_subcat='Total Expenditures'"
        ),
        "$limit": "500",
    })
    values = {}
    for r in rows:
        name = (r.get("dist_name") or "").strip()
        v = parse_num(r.get("ind_value"))
        if name and v and 1000 < v < 200000:
            values[name] = v
    if len(values) < 200:
        sys.exit(f"FATAL: E2C district PPE 2025 parsed {len(values)} districts")
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = round(rec["v"])
    # State row if present
    state = next((r for r in ranked if r["name"].lower() in ("state", "massachusetts")), None)
    return {
        "label": "Massachusetts district total expenditures per pupil, FY 2025",
        "src": "SRC-606-07",
        "unit": "dollars per pupil",
        "as_of_label": "Fiscal year 2025",
        "districts": len(ranked),
        "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"]},
        "lowest": {"name": ranked[-1]["name"], "v": ranked[-1]["v"]},
        "state_row": state["v"] if state else None,
        "top_five": [{"name": r["name"], "v": r["v"]} for r in ranked[:5]],
    }


NAEP_JURS = ",".join(list(RANKED) + ["NP"])


def _naep_displayable(row):
    disp = row.get("isStatDisplayable")
    return disp not in (0, "0", False)


def _naep_all_years(subject, grade, subscale):
    """All published years for one NAEP scale. Year omitted so the API returns the series."""
    url = URL_NAEP + "?" + urllib.parse.urlencode({
        "type": "data",
        "subject": subject,
        "grade": str(grade),
        "subscale": subscale,
        "variable": "TOTAL",
        "jurisdiction": NAEP_JURS,
        "stattype": "MN:MN",
    })
    payload = json.loads(fetch(url, timeout=120))
    by_year = defaultdict(dict)
    us_by_year = {}
    for r in payload.get("result") or []:
        if not _naep_displayable(r):
            continue
        jur = r.get("jurisdiction")
        year = r.get("year")
        v = parse_num(r.get("value"))
        if v is None or year is None:
            continue
        year = int(year)
        if jur == "NP":
            us_by_year[year] = v
        elif jur in STATE_NAMES and jur != "US":
            by_year[year][jur] = v
    return by_year, us_by_year


def _naep_trend(by_year, us_by_year, st):
    years = sorted(set(us_by_year) | set(by_year))
    us, ma = [], []
    for y in years:
        if y in us_by_year:
            us.append({"y": y, "v": round(us_by_year[y], 1)})
        if st in by_year.get(y, {}):
            ma.append({"y": y, "v": round(by_year[y][st], 1)})
    return years, us, ma


def _naep_change(by_year, us_by_year, start, end):
    if start not in by_year or end not in by_year:
        return None
    values = {}
    cells = {}
    for st in RANKED:
        a = by_year[start].get(st)
        b = by_year[end].get(st)
        if a is None or b is None:
            continue
        values[st] = b - a
        cells[st] = (a, b)
    if len(values) < 48:
        return None
    snap = _snap(values, None if start not in us_by_year or end not in us_by_year else us_by_year[end] - us_by_year[start], round_to=1)
    ranked = rank_rows(values, higher_is_better=True)
    rows = []
    for rec in ranked:
        a, b = cells[rec["st"]]
        rows.append({
            "st": rec["st"],
            "name": rec["name"],
            "v": round(rec["v"], 1),
            "from": round(a, 1),
            "to": round(b, 1),
            "rank": rec["rank"],
            "n": rec["n"],
        })
    n_up = sum(1 for r in rows if r["v"] > 0)
    n_down = sum(1 for r in rows if r["v"] < 0)
    snap.update({
        "from_year": start,
        "to_year": end,
        "n_up": n_up,
        "n_down": n_down,
        "n_ranked": rows[0]["n"] if rows else 0,
        "rows": rows,
    })
    return snap


def sec_naep():
    specs = {
        "read4": ("reading", 4, "RRPCM", "reading", 4),
        "read8": ("reading", 8, "RRPCM", "reading", 8),
        "math4": ("mathematics", 4, "MRPCM", "math", 4),
        "math8": ("mathematics", 8, "MRPCM", "math", 8),
    }
    out = {}
    history = {}
    for key, (subj, grade, scale, short, g) in specs.items():
        by_year, us_by_year = _naep_all_years(subj, grade, scale)
        values = by_year.get(2024) or {}
        us_val = us_by_year.get(2024)
        if us_val is None:
            sys.exit(f"FATAL: NAEP 2024 {key} national public is missing")
        if len(values) < 48:
            sys.exit(f"FATAL: NAEP 2024 {key} parsed {len(values)} states")
        snap = _snap(values, us_val, round_to=1)
        snap.update({
            "label": f"NAEP {short} grade {g} average scale score, 2024",
            "src": "SRC-607-05",
            "unit": "scale score",
            "as_of_label": "2024",
        })
        out[key] = snap
        years, us_tr, ma_tr = _naep_trend(by_year, us_by_year, "MA")
        ch19 = _naep_change(by_year, us_by_year, 2019, 2024)
        ch22 = _naep_change(by_year, us_by_year, 2022, 2024)
        if not ch19 or not ch22:
            sys.exit(f"FATAL: NAEP {key} missing 2019-2024 or 2022-2024 change")
        history[key] = {
            "label": f"NAEP {short} grade {g} average scale score",
            "src": "SRC-607-05",
            "unit": "scale score",
            "years": years,
            "us": us_tr,
            "ma": ma_tr,
            "change_2019_2024": ch19,
            "change_2022_2024": {
                "us": ch22.get("us"),
                "ma": ch22.get("ma"),
                "highest": ch22.get("highest"),
                "lowest": ch22.get("lowest"),
                "n_up": ch22.get("n_up"),
                "n_down": ch22.get("n_down"),
                "n_ranked": ch22.get("n_ranked"),
                "from_year": 2022,
                "to_year": 2024,
            },
        }
    r4 = out["read4"]["us"]
    if r4 < VERIFY_NAEP_NP_READ4_2024_MIN or r4 > VERIFY_NAEP_NP_READ4_2024_MAX:
        sys.exit(f"FATAL: NAEP 2024 grade-4 reading NP is {r4}")
    if abs(r4 - VERIFY_NAEP_NP_READ4_2024) > 0.15:
        sys.exit(f"FATAL: NAEP 2024 grade-4 reading NP is {r4}, expected {VERIFY_NAEP_NP_READ4_2024}")
    us19 = next((p["v"] for p in history["read4"]["us"] if p["y"] == 2019), None)
    if us19 is None or abs(us19 - VERIFY_NAEP_NP_READ4_2019) > 0.15:
        sys.exit(f"FATAL: NAEP 2019 grade-4 reading NP is {us19}")
    la = next(
        (r for r in history["read4"]["change_2019_2024"]["rows"] if r.get("st") == "LA"),
        None,
    )
    if not la or abs(la["v"] - VERIFY_NAEP_LA_READ4_CHANGE_2019_2024) > 0.15:
        sys.exit(f"FATAL: Louisiana grade-4 reading 2019-2024 change is {la}")
    return {
        "label": "NAEP state reading and math, 1992-2024",
        "src": "SRC-607-05",
        "as_of_label": "2024",
        "note": (
            "Average scale scores for public-school students. National public is "
            "the published NP line. Change is 2024 minus the earlier year. "
            "States with no published score in either year are omitted from that change ranking."
        ),
        "series": out,
        "history": history,
    }


def sec_ipeds_outcomes():
    wb = load_workbook(io.BytesIO(fetch(DIGEST_326, timeout=90)), data_only=True)
    ws = wb.active
    # Find the 6-year, all 4-year institutions, 2017 entry cohort total.
    in_six = False
    us_val = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        label = str(row[0] or "").replace("\xa0", " ").strip()
        if "within 6 years" in label.lower():
            in_six = True
            continue
        if in_six and label.startswith("2017"):
            us_val = parse_num(row[1])
            break
    if us_val is None or abs(us_val - VERIFY_IPEDS_6YR_2017) > 0.2:
        sys.exit(f"FATAL: Digest 326.10 2017 6-year graduation rate is {us_val}")
    return {
        "label": "IPEDS 6-year bachelor's graduation rate, 2017 entry cohort",
        "src": "SRC-608-04",
        "unit": "percent",
        "as_of_label": "2017 cohort, completing by 2023",
        "us": us_val,
        "ma": None,
        "note": "National table. Digest 326.10 has no state column.",
    }


# ---------------------------------------------------------------------------
# Healthcare, labor, GDP, housing, population
# ---------------------------------------------------------------------------

def sec_mfcu():
    wb = load_workbook(io.BytesIO(fetch(URL_OIG_MFCU, timeout=90)), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[2]]
    rec_i = headers.index("Total Recoveries2")
    values, us_val = {}, None
    for row in ws.iter_rows(min_row=3, values_only=True):
        st = geo_to_st(row[0])
        v = parse_num(row[rec_i])
        if v is None:
            continue
        if str(row[0]).strip().lower() == "total":
            us_val = v
            continue
        if st and st != "US":
            values[st] = v
    if us_val is None or abs(us_val - VERIFY_OIG_US_RECOVERIES_FY2025) > 1:
        sys.exit(f"FATAL: OIG MFCU US recoveries are {us_val}")
    if abs(values.get("MA", 0) - VERIFY_OIG_MA_RECOVERIES_FY2025) > 1:
        sys.exit(f"FATAL: OIG MFCU MA recoveries are {values.get('MA')}")
    if len(values) < 51:
        sys.exit(f"FATAL: OIG MFCU FY 2025 parsed {len(values)} states")
    snap = _snap(values, us_val, round_to=0)
    snap.update({
        "label": "Medicaid Fraud Control Unit total recoveries, FY 2025",
        "src": "SRC-612-03",
        "unit": "dollars",
        "as_of_label": "Fiscal year 2025",
    })
    return snap


def sec_ui_claims():
    text = fetch_text(URL_UI, timeout=120)
    rows = list(csv.DictReader(io.StringIO(text)))
    latest = {}
    for r in rows:
        st = (r.get("st") or "").strip()
        if st not in STATE_NAMES or st == "US":
            continue
        week = (r.get("rptdate") or "")[:10]
        init = parse_num(r.get("c3"))
        cont = parse_num(r.get("c8"))
        if not week or init is None:
            continue
        prev = latest.get(st)
        if prev is None or week > prev["week"]:
            latest[st] = {"week": week, "initial": init, "continued": cont}
    weeks = {v["week"] for v in latest.values()}
    week = max(weeks)
    values = {st: rec["initial"] for st, rec in latest.items() if rec["week"] == week}
    if len(values) < 50:
        sys.exit(f"FATAL: ETA 539 {week} parsed {len(values)} states")
    us_val = sum(values.values())
    snap = _snap(values, us_val, round_to=0)
    ma_cont = latest.get("MA", {}).get("continued")
    snap.update({
        "label": f"UI initial claims, week ending {week}",
        "src": "SRC-614-03",
        "unit": "claims",
        "as_of_label": week,
        "ma_continued": int(ma_cont) if ma_cont is not None else None,
        "note": "ETA 539 cell C3 is initial claims; C8 is continued weeks claimed. U.S. is the sum of states and D.C.",
    })
    return snap


def sec_sagdp2():
    zf = zipfile.ZipFile(io.BytesIO(fetch(URL_SAGDP, timeout=180)))
    rows = list(csv.DictReader(io.TextIOWrapper(
        zf.open("SAGDP2__ALL_AREAS_1997_2025.csv"), encoding="latin-1"
    )))
    want = {
        "1": "all_industry",
        "12": "manufacturing",
        "51": "finance_insurance",
        "45": "information",
        "11": "construction",
    }
    out = {}
    for line, key in want.items():
        values, us_val = {}, None
        for r in rows:
            if str(r.get("LineCode") or "").split(".")[0] != line:
                continue
            name = r.get("GeoName")
            v = parse_num(r.get("2025"))
            if v is None:
                continue
            if name == "United States *":
                us_val = v
                continue
            st = geo_to_st(name)
            if st and st != "US":
                values[st] = v
        if us_val is None or "MA" not in values or len(values) < 40:
            sys.exit(f"FATAL: SAGDP2 line {line} US={us_val} n={len(values)}")
        snap = _snap(values, us_val, round_to=1)
        snap.update({
            "label": f"Current-dollar GDP, {key.replace('_', ' ')}, 2025",
            "src": "SRC-615-03",
            "unit": "millions of current dollars",
            "as_of_label": "Calendar year 2025",
        })
        out[key] = snap
    return {
        "label": "BEA SAGDP2 current-dollar GDP by NAICS, 2025",
        "src": "SRC-615-03",
        "as_of_label": "Calendar year 2025",
        "industries": out,
    }


def _fred_cs(url, col):
    text = fetch_text(url, timeout=60)
    rows = [(r["observation_date"], parse_num(r[col]))
            for r in csv.DictReader(io.StringIO(text))
            if parse_num(r.get(col)) is not None]
    if len(rows) < 24:
        sys.exit(f"FATAL: FRED {col} parsed {len(rows)} months")
    last_d, last_v = rows[-1]
    prev = next(
        (v for d, v in reversed(rows)
         if d[:4] == str(int(last_d[:4]) - 1) and d[5:7] == last_d[5:7]),
        None,
    )
    return {
        "as_of": last_d[:7],
        "v": round(last_v, 2),
        "yoy_pct": yoy_pct(last_v, prev),
        "trend": [{"m": d[:7], "v": round(v, 2)} for d, v in rows[-36:]],
    }


def sec_case_shiller():
    bos = _fred_cs(URL_CS_BOS, "BOXRSA")
    mia = _fred_cs(URL_CS_MIA, "MIXRSA")
    return {
        "label": "S&P/CoreLogic Case-Shiller Boston and Miami house-price indexes",
        "src": "SRC-616-03",
        "unit": "index, January 2000 = 100",
        "as_of_label": bos["as_of"],
        "boston": bos["v"],
        "yoy_pct": bos["yoy_pct"],
        "miami": mia["v"],
        "miami_as_of_label": mia["as_of"],
        "miami_yoy_pct": mia["yoy_pct"],
        "trend": bos["trend"],
        "miami_trend": mia["trend"],
        "note": (
            "Seasonally adjusted Boston MSA series BOXRSA and Miami MSA series "
            "MIXRSA via FRED. January 2000 equals 100. Case-Shiller does not "
            "publish another Massachusetts city; Miami is the Florida 20-city "
            "counterpart."
        ),
    }


def sec_rucc():
    text = fetch(URL_RUCC, timeout=60).decode("latin-1")
    metro = defaultdict(lambda: {"counties": 0, "pop": 0, "nonmetro_counties": 0, "nonmetro_pop": 0})
    pop = {}
    code = {}
    for r in csv.DictReader(io.StringIO(text)):
        fips = (r.get("FIPS") or "").zfill(5)
        st = (r.get("State") or "").strip()
        attr = r.get("Attribute")
        v = r.get("Value")
        if attr == "Population_2020":
            pop[fips] = parse_num(v) or 0
        elif attr == "RUCC_2023":
            code[fips] = parse_num(v)
    for fips, rucc in code.items():
        st = None
        # FIPS 01001 -> need state from a row; recover from first two digits via a scan
    # Re-read with state
    by_fips_st = {}
    for r in csv.DictReader(io.StringIO(text)):
        fips = (r.get("FIPS") or "").zfill(5)
        by_fips_st[fips] = (r.get("State") or "").strip()
    for fips, rucc in code.items():
        st = by_fips_st.get(fips)
        if st not in STATE_NAMES or st == "US" or rucc is None:
            continue
        p = pop.get(fips) or 0
        if rucc <= 3:
            metro[st]["counties"] += 1
            metro[st]["pop"] += p
        else:
            metro[st]["nonmetro_counties"] += 1
            metro[st]["nonmetro_pop"] += p
    values = {}
    for st, rec in metro.items():
        tot = rec["pop"] + rec["nonmetro_pop"]
        if tot:
            values[st] = 100 * rec["pop"] / tot
    if "MA" not in values:
        sys.exit("FATAL: RUCC 2023 missing Massachusetts")
    snap = _snap(values, None, round_to=1)
    ma = metro["MA"]
    snap.update({
        "label": "Share of 2020 population in metro RUCC 1-3 counties, 2023 codes",
        "src": "SRC-617-02",
        "unit": "percent",
        "as_of_label": "USDA RUCC 2023",
        "ma_metro_counties": ma["counties"],
        "ma_nonmetro_counties": ma["nonmetro_counties"],
        "us": None,
        "note": "Codes are county-level. Massachusetts has no nonmetro county under the 2023 codes.",
    })
    return snap


def _pack_irs_county(state, state_name, by_county, stubs, min_counties):
    if len(by_county) < min_counties:
        sys.exit(
            f"FATAL: IRS county TY2022 {state} parsed {len(by_county)} counties"
        )
    values = {k: v["agi"] * 1000 for k, v in by_county.items()}
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = round(rec["v"])
        rec["returns"] = int(by_county[rec["name"]]["returns"])
    return {
        "label": f"{state_name} county adjusted gross income, tax year 2022",
        "src": "SRC-621-02",
        "unit": "dollars",
        "as_of_label": "Tax year 2022",
        "counties": len(ranked),
        "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"]},
        "lowest": {"name": ranked[-1]["name"], "v": ranked[-1]["v"]},
        "size_class": {
            k: {"returns": int(v["returns"]), "agi_thousands": round(v["agi"])}
            for k, v in stubs.items()
        },
        "note": "County AGI is the sum of SOI size-of-AGI stubs. Amounts in A00100 are thousands of dollars. A dedicated AGI-percentile-by-state file is not posted.",
        "top_five": [{"name": r["name"], "v": r["v"]} for r in ranked[:5]],
    }


def sec_irs_county_bundle():
    text = fetch_text(URL_IRS_COUNTY, timeout=120)
    by_st = {
        "MA": defaultdict(lambda: {"returns": 0, "agi": 0}),
        "FL": defaultdict(lambda: {"returns": 0, "agi": 0}),
    }
    stubs = {
        "MA": defaultdict(lambda: {"returns": 0, "agi": 0}),
        "FL": defaultdict(lambda: {"returns": 0, "agi": 0}),
    }
    for r in csv.DictReader(io.StringIO(text)):
        st = r.get("STATE")
        if st not in by_st:
            continue
        fips = (r.get("COUNTYFIPS") or "").zfill(3)
        name = (r.get("COUNTYNAME") or "").strip()
        n = parse_num(r.get("N1")) or 0
        agi = parse_num(r.get("A00100")) or 0  # thousands of dollars
        stub = r.get("agi_stub")
        if fips == "000":
            stubs[st][stub]["returns"] += n
            stubs[st][stub]["agi"] += agi
            continue
        by_st[st][name]["returns"] += n
        by_st[st][name]["agi"] += agi
    return {
        "ma_county_agi_2022": _pack_irs_county(
            "MA", "Massachusetts", by_st["MA"], stubs["MA"], 10
        ),
        "fl_county_agi_2022": _pack_irs_county(
            "FL", "Florida", by_st["FL"], stubs["FL"], 50
        ),
    }


# ---------------------------------------------------------------------------
# Transit, roads, energy, municipal, payroll
# ---------------------------------------------------------------------------

def sec_ntd_annual():
    rows = _soda(URL_NTD_ANN, {
        "$select": "agency,state,sum(total_operating_expenses) as opexp,sum(fare_revenues_earned) as fares,sum(unlinked_passenger_trips) as upt",
        "$where": "report_year='2024'",
        "$group": "agency,state",
        "$order": "opexp DESC",
        "$limit": "5000",
    })
    values, fares, upt, states = {}, {}, {}, {}
    for r in rows:
        name = (r.get("agency") or "").strip()
        st = (r.get("state") or "").strip()
        o = parse_num(r.get("opexp"))
        if not name or o is None or o <= 0:
            continue
        key = name if name not in values else f"{name} ({st})"
        values[key] = o
        fares[key] = parse_num(r.get("fares")) or 0
        upt[key] = parse_num(r.get("upt")) or 0
        states[key] = st
    if len(values) < 200:
        sys.exit(f"FATAL: NTD 2024 annual metrics parsed {len(values)} agencies")
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    for rec in ranked:
        rec["v"] = round(rec["v"])
        rec["state"] = states.get(rec["name"])
        rec["fares"] = round(fares.get(rec["name"], 0))
        rec["upt"] = int(round(upt.get(rec["name"], 0)))
        rec["farebox_pct"] = round(100 * rec["fares"] / rec["v"], 1) if rec["v"] else None
        rec["cost_per_trip"] = round(rec["v"] / rec["upt"], 2) if rec["upt"] else None
    mbta = next((r for r in ranked if r.get("state") == "MA" and "Bay Transportation" in r["name"]), None)
    if not mbta:
        sys.exit("FATAL: NTD 2024 annual metrics missing the MBTA")
    us_opexp = sum(values.values())
    us_fares = sum(fares.values())
    return {
        "label": "FTA NTD agency operating expenses and farebox, report year 2024",
        "src": "SRC-622-02",
        "unit": "dollars",
        "as_of_label": "Report year 2024",
        "us_operating": round(us_opexp),
        "us_farebox_pct": round(100 * us_fares / us_opexp, 1) if us_opexp else None,
        "agencies": len(ranked),
        "mbta": {
            "name": mbta["name"],
            "operating": mbta["v"],
            "farebox_pct": mbta["farebox_pct"],
            "cost_per_trip": mbta["cost_per_trip"],
            "rank": mbta["rank"],
            "n": mbta["n"],
        },
        "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"]},
    }


def sec_openfema():
    """Sum OpenFEMA Public Assistance federal share obligated by state."""
    url = (
        "https://www.fema.gov/api/open/v2/PublicAssistanceGrantAwardActivities.csv"
        "?$select=stateAbbreviation,federalShareObligated"
    )
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    values = defaultdict(float)
    n = 0
    with urllib.request.urlopen(req, timeout=300) as resp:
        rdr = csv.DictReader(io.TextIOWrapper(resp, encoding="utf-8", newline=""))
        fields = rdr.fieldnames or []
        amt_key = next(
            (k for k in fields if "federalshareobligated" in k.lower()),
            None,
        )
        st_key = next(
            (k for k in fields if k.lower() in ("stateabbreviation", "state")),
            None,
        )
        if not amt_key or not st_key:
            sys.exit(f"FATAL: OpenFEMA columns are {fields[:12]}")
        for r in rdr:
            st = (r.get(st_key) or "").strip()
            v = parse_num(r.get(amt_key))
            if st in STATE_NAMES and st != "US" and v:
                values[st] += v
                n += 1
    if n < 1000 or "MA" not in values:
        sys.exit(f"FATAL: OpenFEMA parsed n={n} states={len(values)}")
    snap = _snap(dict(values), sum(values.values()), round_to=0)
    snap.update({
        "label": "FEMA Public Assistance federal share obligated, all disasters on file",
        "src": "SRC-623-03",
        "unit": "dollars",
        "as_of_label": "OpenFEMA extract, retrieved Aug 2026",
        "award_rows": n,
    })
    return snap


def sec_nri():
    feats = []
    offset = 0
    while True:
        url = URL_NRI + "?" + urllib.parse.urlencode({
            "where": "1=1",
            "outFields": "STATEABBRV,RISK_SCORE,RISK_RATNG",
            "returnGeometry": "false",
            "f": "json",
            "resultRecordCount": "2000",
            "resultOffset": str(offset),
        })
        payload = json.loads(fetch(url, timeout=90))
        batch = payload.get("features") or []
        feats.extend(batch)
        if len(batch) < 2000:
            break
        offset += len(batch)
        if offset > 10000:
            break
    if len(feats) < 3000:
        sys.exit(f"FATAL: NRI county query returned {len(feats)} features")
    by_st = defaultdict(list)
    for f in feats:
        a = f.get("attributes") or {}
        st = a.get("STATEABBRV")
        v = parse_num(a.get("RISK_SCORE"))
        if st in STATE_NAMES and st != "US" and v is not None:
            by_st[st].append(v)
    values = {st: sum(vs) / len(vs) for st, vs in by_st.items() if vs}
    snap = _snap(values, None, round_to=1)
    snap.update({
        "label": "FEMA National Risk Index, mean county risk score",
        "src": "SRC-623-04",
        "unit": "index score",
        "as_of_label": "NRI v1.20, December 2025",
        "us": None,
        "note": "State figure is the unweighted mean of county risk scores.",
    })
    return snap


def sec_noaa_degree_days():
    # Massachusetts climate division code is 19; contiguous U.S. is 110.
    out = {}
    for key, var, path, code in (
        ("ma_hdd", "hdd", "statewide", "19"),
        ("ma_cdd", "cdd", "statewide", "19"),
        ("us_hdd", "hdd", "national", "110"),
        ("us_cdd", "cdd", "national", "110"),
    ):
        url = (
            "https://www.ncei.noaa.gov/access/monitoring/climate-at-a-glance/"
            f"{path}/time-series/{code}/{var}/ann/12/2000-2024.csv"
        )
        try:
            text = fetch_text(url, timeout=60)
        except Exception:
            return None
        last = None
        for line in text.splitlines():
            if not line or line[0] not in "12":
                continue
            parts = line.split(",")
            if len(parts) >= 2 and parts[0].startswith("2024"):
                last = parse_num(parts[1])
        if last is None:
            return None
        out[key] = last
    return {
        "label": "NOAA heating and cooling degree days, 2024",
        "src": "SRC-623-05",
        "unit": "degree days",
        "as_of_label": "Calendar year 2024",
        "ma_hdd": out["ma_hdd"],
        "ma_cdd": out["ma_cdd"],
        "us_hdd": out["us_hdd"],
        "us_cdd": out["us_cdd"],
    }


def sec_seds_production():
    wb = load_workbook(io.BytesIO(fetch(URL_SEDS_PROD, timeout=120)), data_only=True)
    ws = wb["Data"]
    headers = [c.value for c in ws[1]]
    # Expect State, MSN, years...
    year_col = None
    state_col = msn_col = 0
    for i, h in enumerate(headers):
        if str(h).strip() == "2024":
            year_col = i
        if str(h).strip() in ("State", "StateCode"):
            state_col = i
        if str(h).strip() == "MSN":
            msn_col = i
    if year_col is None:
        # maybe first row is title
        for r0 in range(1, 6):
            headers = [c.value for c in ws[r0]]
            if "2024" in {str(h).strip() for h in headers if h is not None}:
                year_col = [i for i, h in enumerate(headers) if str(h).strip() == "2024"][0]
                state_col = next(i for i, h in enumerate(headers) if str(h).strip() in ("State", "StateCode"))
                msn_col = next(i for i, h in enumerate(headers) if str(h).strip() == "MSN")
                start = r0 + 1
                break
        else:
            sys.exit(f"FATAL: SEDS Prod_dataset has no 2024 column, headers={headers[:8]}")
    else:
        start = 2
    values, us_val = {}, None
    for row in ws.iter_rows(min_row=start, values_only=True):
        if str(row[msn_col] or "").strip() != "TEPRB":
            continue
        st = str(row[state_col] or "").strip()
        v = parse_num(row[year_col])
        if v is None:
            continue
        if st == "US":
            us_val = v
        elif st in STATE_NAMES:
            values[st] = v
    if us_val is None or "MA" not in values:
        sys.exit(f"FATAL: SEDS TEPRB 2024 US={us_val} MA={values.get('MA')}")
    snap = _snap(values, us_val, round_to=0)
    snap.update({
        "label": "SEDS total energy production, 2024 (billion Btu)",
        "src": "SRC-624-03",
        "unit": "billion Btu",
        "as_of_label": "Calendar year 2024",
    })
    return snap


def sec_pop_peers():
    text = fetch_text(URL_SUBEST, timeout=90)
    towns = []
    for r in csv.DictReader(io.StringIO(text)):
        if r.get("SUMLEV") != "061":
            continue
        p = parse_num(r.get("POPESTIMATE2025"))
        name = (r.get("NAME") or "").strip()
        if p and name:
            towns.append({"name": name, "pop": int(p)})
    if len(towns) != 351:
        sys.exit(f"FATAL: peer-set expected 351 towns, got {len(towns)}")
    towns.sort(key=lambda t: t["pop"], reverse=True)
    by_name = {t["name"]: t for t in towns}
    # Five nearest population peers for Boston and a few large cities.
    focus = [t for t in towns if t["name"] in (
        "Boston city", "Worcester city", "Springfield city",
        "Cambridge city", "Lowell city",
    )]
    peers = {}
    for t in focus:
        others = sorted(towns, key=lambda x: abs(x["pop"] - t["pop"]) if x["name"] != t["name"] else 10**12)
        peers[t["name"]] = [{"name": x["name"], "pop": x["pop"]} for x in others[:5]]
    return {
        "label": "Massachusetts municipal population peers, July 1, 2025",
        "src": "SRC-625-02",
        "as_of_label": "July 1, 2025",
        "method": "Five nearest Census 2025 populations. This is not the old Pioneer socioeconomic peer workbook.",
        "peers": peers,
        "n_towns": 351,
    }


def sec_quasi_payroll():
    years = _soda(URL_QUASI, {
        "$select": "calendar_year,count(*) as n,sum(pay_year_to_date) as pay",
        "$group": "calendar_year",
        "$order": "calendar_year",
    })
    y = next((r for r in years if str(r.get("calendar_year")) == "2025"), None)
    if not y:
        sys.exit("FATAL: CTHRU quasi-public payroll missing calendar 2025")
    total = parse_num(y.get("pay"))
    n = int(float(y.get("n") or 0))
    if total is None or total < 1e8:
        sys.exit(f"FATAL: CTHRU quasi 2025 pay={total}")
    depts = _soda(URL_QUASI, {
        "$select": "quasi_agency_name,count(*) as n,sum(pay_year_to_date) as pay",
        "$where": "calendar_year='2025'",
        "$group": "quasi_agency_name",
        "$order": "pay DESC",
        "$limit": "200",
    })
    values = {}
    for r in depts:
        name = (r.get("quasi_agency_name") or "").strip()
        v = parse_num(r.get("pay"))
        if name and v:
            values[name] = v
    ranked = rank_named(values, higher_is_better=True, st_key=lambda n: n)
    return {
        "label": "Massachusetts quasi-public payroll, calendar 2025",
        "src": "SRC-630-03",
        "unit": "dollars",
        "as_of_label": "Calendar year 2025",
        "total": total,
        "employees": n,
        "agencies": len(ranked),
        "highest": {"name": ranked[0]["name"], "v": ranked[0]["v"]},
        "top_five": [{"name": r["name"], "v": r["v"]} for r in ranked[:5]],
    }


def sec_vendor_extract():
    # Object classes that are not employee payroll transfers.
    rows = _soda(URL_SPEND, {
        "$select": "object_class,sum(amount) as amt",
        "$where": "budget_fiscal_year='2025'",
        "$group": "object_class",
        "$order": "amt DESC",
        "$limit": "50",
    })
    skip = ("payroll", "salary", "wage", "benefit", "pension", "transfer")
    vendor_classes = []
    for r in rows:
        name = (r.get("object_class") or "").strip()
        amt = parse_num(r.get("amt"))
        if not name or amt is None:
            continue
        low = name.lower()
        if any(s in low for s in skip):
            continue
        vendor_classes.append(name)
    if not vendor_classes:
        sys.exit("FATAL: CTHRU spending has no non-payroll object class")
    # Top vendors in those classes.
    where = (
        "budget_fiscal_year='2025' AND ("
        + " OR ".join(f"object_class='{c.replace(chr(39), chr(39)+chr(39))}'" for c in vendor_classes[:12])
        + ")"
    )
    vendors = _soda(URL_SPEND, {
        "$select": "vendor,sum(amount) as amt",
        "$where": where,
        "$group": "vendor",
        "$order": "amt DESC",
        "$limit": "25",
    })
    top = []
    total = 0
    for r in vendors:
        name = (r.get("vendor") or "").strip()
        v = parse_num(r.get("amt"))
        if name and v:
            top.append({"name": name, "v": v})
            total += v
    if not top:
        sys.exit("FATAL: CTHRU vendor extract returned no rows")
    return {
        "label": "CTHRU vendor payments, non-payroll object classes, FY 2025",
        "src": "SRC-630-04",
        "unit": "dollars",
        "as_of_label": "Fiscal year 2025",
        "object_classes": vendor_classes[:12],
        "top_vendor_sum": total,
        "highest": top[0],
        "top_ten": top[:10],
        "note": "Excludes object classes whose names include payroll, salary, wage, benefit, pension, or transfer.",
    }


def _bd_rate_id(st, kind):
    """kind is 'birth' or 'death'. Rate series, seasonally adjusted, quarterly."""
    fips = ST_TO_FIPS.get(st)
    if not fips:
        return None
    elem = "07" if kind == "birth" else "08"
    return f"BDS00000{fips}0000000001200{elem}RQ5"


def _bls_bd_series(ids, start="2006", end="2026", require_all=True):
    # Unregistered BLS API: 25 series and ten calendar years per call.
    ids = list(ids)
    out = {sid: [] for sid in ids}
    year = int(start)
    end_y = int(end)
    while year <= end_y:
        chunk_end = min(year + 9, end_y)
        for i in range(0, len(ids), 25):
            batch = ids[i:i + 25]
            body = json.dumps({
                "seriesid": batch,
                "startyear": str(year),
                "endyear": str(chunk_end),
            }).encode()
            req = urllib.request.Request(
                BLS_API,
                data=body,
                headers={"User-Agent": UA, "Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=120) as resp:
                payload = json.loads(resp.read())
            if payload.get("status") != "REQUEST_SUCCEEDED":
                sys.exit(f"FATAL: BLS BED API {payload.get('status')} {payload.get('message')}")
            for s in payload.get("Results", {}).get("series", []):
                sid = s["seriesID"]
                if sid not in out:
                    out[sid] = []
                seen = {(y, q) for y, q, _ in out[sid]}
                for r in s.get("data") or []:
                    raw = (r.get("value") or "").strip()
                    v = None if raw in ("", "-") else parse_num(raw)
                    y, q = int(r["year"]), int(r["period"][1:])
                    if (y, q) not in seen:
                        out[sid].append((y, q, v))
                        seen.add((y, q))
        year = chunk_end + 1
    for sid in out:
        out[sid].sort()
    missing = [i for i in ids if not out.get(i)]
    if missing and require_all:
        sys.exit(f"FATAL: BLS BED missing series {missing}")
    return out


def _bd_lookup(points, year, quarter):
    for y, q, v in points:
        if y == year and q == quarter:
            return v
    return None


def _bd_last(points):
    for y, q, v in reversed(points):
        if v is not None:
            return y, q, v
    return None


def _q_label(year, quarter):
    return f"{year} Q{quarter}"


def sec_bed_births_deaths():
    level_ids = (
        BD_US_BIRTHS_L, BD_US_DEATHS_L, BD_MA_BIRTHS_L, BD_MA_DEATHS_L,
        BD_FL_BIRTHS_L, BD_FL_DEATHS_L,
        BD_US_BIRTHS_R, BD_US_DEATHS_R, BD_MA_BIRTHS_R, BD_MA_DEATHS_R,
        BD_FL_BIRTHS_R, BD_FL_DEATHS_R,
    )
    rate_ids = []
    rate_map = {}
    for st in list(RANKED) + ["US"]:
        bid = _bd_rate_id(st, "birth")
        did = _bd_rate_id(st, "death")
        if bid:
            rate_ids.append(bid)
            rate_map[bid] = (st, "birth")
        if did:
            rate_ids.append(did)
            rate_map[did] = (st, "death")
    series = _bls_bd_series(level_ids, start="2006", end="2026", require_all=True)
    rates = _bls_bd_series(rate_ids, start="2006", end="2026", require_all=False)
    series.update(rates)
    us_b4 = _bd_lookup(series[BD_US_BIRTHS_L], 2025, 4)
    us_d1 = _bd_lookup(series[BD_US_DEATHS_L], 2025, 1)
    if us_b4 != VERIFY_US_BED_BIRTHS_THOUSANDS_2025Q4:
        sys.exit(
            f"FATAL: BLS US 2025 Q4 establishment births are {us_b4}, "
            f"expected {VERIFY_US_BED_BIRTHS_THOUSANDS_2025Q4} thousand "
            "(CEWBD 2025 Q4 news release)"
        )
    if us_d1 != VERIFY_US_BED_DEATHS_THOUSANDS_2025Q1:
        sys.exit(
            f"FATAL: BLS US 2025 Q1 establishment deaths are {us_d1}, "
            f"expected {VERIFY_US_BED_DEATHS_THOUSANDS_2025Q1} thousand "
            "(CEWBD 2025 Q4 news release)"
        )
    by_q = {}
    fields = (
        (BD_US_BIRTHS_L, "us_births_thousands"),
        (BD_US_DEATHS_L, "us_deaths_thousands"),
        (BD_MA_BIRTHS_L, "ma_births"),
        (BD_MA_DEATHS_L, "ma_deaths"),
        (BD_FL_BIRTHS_L, "fl_births"),
        (BD_FL_DEATHS_L, "fl_deaths"),
        (BD_US_BIRTHS_R, "us_birth_rate_pct"),
        (BD_US_DEATHS_R, "us_death_rate_pct"),
        (BD_MA_BIRTHS_R, "ma_birth_rate_pct"),
        (BD_MA_DEATHS_R, "ma_death_rate_pct"),
        (BD_FL_BIRTHS_R, "fl_birth_rate_pct"),
        (BD_FL_DEATHS_R, "fl_death_rate_pct"),
    )
    for sid, key in fields:
        for y, q, v in series[sid]:
            rec = by_q.setdefault((y, q), {"q": _q_label(y, q)})
            rec[key] = int(v) if v is not None and key.endswith(("thousands", "births", "deaths")) and "rate" not in key else v
            if v is not None and key.endswith("rate_pct"):
                rec[key] = round(v, 1)
    trend = [by_q[k] for k in sorted(by_q)]
    by, bq, bv = _bd_last(series[BD_MA_BIRTHS_R])
    dy, dq, dv = _bd_last(series[BD_MA_DEATHS_R])
    us_by, us_bq, us_bv = _bd_last(series[BD_US_BIRTHS_R])
    us_dy, us_dq, us_dv = _bd_last(series[BD_US_DEATHS_R])
    fl_by, fl_bq, fl_bv = _bd_last(series[BD_FL_BIRTHS_R])
    fl_dy, fl_dq, fl_dv = _bd_last(series[BD_FL_DEATHS_R])
    ma_b_n = _bd_lookup(series[BD_MA_BIRTHS_L], by, bq)
    ma_d_n = _bd_lookup(series[BD_MA_DEATHS_L], dy, dq)
    fl_b_n = _bd_lookup(series[BD_FL_BIRTHS_L], fl_by, fl_bq)
    fl_d_n = _bd_lookup(series[BD_FL_DEATHS_L], fl_dy, fl_dq)
    overlap = next(
        (
            rec for rec in reversed(trend)
            if rec.get("ma_birth_rate_pct") is not None
            and rec.get("ma_death_rate_pct") is not None
        ),
        None,
    )
    fl_overlap = next(
        (
            rec for rec in reversed(trend)
            if rec.get("fl_birth_rate_pct") is not None
            and rec.get("fl_death_rate_pct") is not None
        ),
        None,
    )
    return {
        "label": "Private-sector establishment birth and death rates",
        "src": "SRC-613-02",
        "unit": "percent of establishments",
        "count_unit_us": "thousands of establishments",
        "count_unit_ma": "establishments",
        "as_of_label": _q_label(by, bq),
        "deaths_as_of_label": _q_label(dy, dq),
        "us": {
            "birth_rate_pct": us_bv,
            "births_thousands": int(_bd_lookup(series[BD_US_BIRTHS_L], us_by, us_bq) or 0),
            "births_as_of": _q_label(us_by, us_bq),
            "death_rate_pct": us_dv,
            "deaths_thousands": int(_bd_lookup(series[BD_US_DEATHS_L], us_dy, us_dq) or 0),
            "deaths_as_of": _q_label(us_dy, us_dq),
        },
        "ma": {
            "birth_rate_pct": bv,
            "births": int(ma_b_n) if ma_b_n is not None else None,
            "births_as_of": _q_label(by, bq),
            "death_rate_pct": dv,
            "deaths": int(ma_d_n) if ma_d_n is not None else None,
            "deaths_as_of": _q_label(dy, dq),
        },
        "fl": {
            "birth_rate_pct": fl_bv,
            "births": int(fl_b_n) if fl_b_n is not None else None,
            "births_as_of": _q_label(fl_by, fl_bq),
            "death_rate_pct": fl_dv,
            "deaths": int(fl_d_n) if fl_d_n is not None else None,
            "deaths_as_of": _q_label(fl_dy, fl_dq),
        },
        "overlap": overlap,
        "fl_overlap": fl_overlap,
        "trend": trend,
        "states": _bed_state_cube(rates, rate_map),
        "note": (
            "BLS Business Employment Dynamics, total private, seasonally adjusted. "
            "Births are a subset of openings; deaths are a subset of closings and "
            "lag three quarters. Rates are the component as a percent of the "
            "average of current and prior-quarter establishment counts. U.S. "
            "counts are thousands of establishments, matching the BLS news release. "
            "State rates cover every published jurisdiction on the same series. "
            "The states cube is the all-jurisdiction quarterly file; the trend "
            "array keeps the United States, Massachusetts, and Florida overlay."
        ),
    }


def _bed_state_cube(rates, rate_map):
    """{st: [{q, birth_rate_pct, death_rate_pct}, ...]} for modelSlice."""
    by_st = {}
    for sid, points in rates.items():
        rec = rate_map.get(sid)
        if not rec:
            continue
        st, kind = rec
        key = "birth_rate_pct" if kind == "birth" else "death_rate_pct"
        for y, q, v in points:
            row = by_st.setdefault(st, {}).setdefault((y, q), {"q": _q_label(y, q)})
            if v is not None:
                row[key] = round(v, 1)
    cube = {}
    for st, recs in by_st.items():
        series = [recs[k] for k in sorted(recs)]
        if len(series) >= 2:
            cube[st] = series
    return cube


MORE_SECONDARY = {
    "DL-06": lambda: {
        k: v for k, v in {
            "public_k12_enrollment": sec_public_k12_enrollment(),
            "mcas_2025": sec_mcas(),
            "attendance_2025": sec_attendance(),
            "dropouts_2025": sec_dropouts(),
            "district_finance_fy2025": sec_district_finance(),
            **sec_ma_demographics(),
        }.items() if v
    },
    "DL-07": lambda: {"naep_2024": sec_naep(), **hollow_secondary("DL-07"), **sec_expulsion(), **sec_discipline_race()},
    "DL-08": lambda: {
        "ipeds_6yr_grad_2017": sec_ipeds_outcomes(),
        **hollow_secondary("DL-08"),
        **sec_public_he_faculty(),
        **sec_he_finance(),
        **sec_he_students(),
        **sec_ipeds_grad_by_state(),
    },
    "DL-10": lambda: hollow_secondary("DL-10"),
    "DL-12": lambda: {"mfcu_recoveries_fy2025": sec_mfcu()},
    "DL-13": lambda: {"bed_births_deaths": sec_bed_births_deaths()},
    "DL-14": lambda: {"ui_initial_claims": sec_ui_claims(), **hollow_secondary("DL-14"), **sec_qcew_employment()},
    "DL-19": lambda: hollow_secondary("DL-19"),
    "DL-15": lambda: {"sagdp2_naics_2025": sec_sagdp2()},
    "DL-16": lambda: {"case_shiller_boston": sec_case_shiller()},
    "DL-17": lambda: {"rucc_2023": sec_rucc(), **sec_pop_age_race(), **sec_pop_components()},
    "DL-21": lambda: {**sec_irs_county_bundle(), **hollow_secondary("DL-21")},
    "DL-22": lambda: {"ntd_annual_2024": sec_ntd_annual()},
    "DL-23": lambda: {
        k: v for k, v in {
            "fema_pa_obligations": sec_openfema(),
            "nri_mean_county_score": sec_nri(),
            "noaa_degree_days_2024": sec_noaa_degree_days(),
        }.items() if v
    },
    "DL-24": lambda: {"seds_production_2024": sec_seds_production()},
    "DL-25": lambda: {"population_peers_2025": sec_pop_peers(), **hollow_secondary("DL-25")},
    "DL-26": lambda: {"district_ppe_fy2025": sec_district_finance(), **hollow_secondary("DL-26")},
    "DL-28": lambda: hollow_secondary("DL-28"),
    "DL-29": lambda: {
        **hollow_secondary("DL-29"),
        **sec_aslg_2022(),
        **sec_gov_orgs(),
        **sec_aspp_2025(),
    },
    "DL-31": lambda: hollow_secondary("DL-31"),
    "DL-30": lambda: {
        "quasi_payroll_2025": sec_quasi_payroll(),
        "vendor_extract_fy2025": sec_vendor_extract(),
    },
}


def more_lead(tool_id, sec):
    parts = []
    if tool_id == "DL-06":
        m = sec.get("mcas_2025") or {}
        a = sec.get("attendance_2025") or {}
        d = sec.get("dropouts_2025") or {}
        f = sec.get("district_finance_fy2025") or {}
        parts.append(
            f"On the 2025 Next Generation MCAS, <b>{m.get('ela_3_8_pct')}%</b> of "
            f"grades 3-8 students met or exceeded expectations in English language "
            f"arts and <b>{m.get('math_3_8_pct')}%</b> in mathematics (SRC-606-04). "
            f"Grade 10 was <b>{m.get('ela_10_pct')}%</b> ELA and "
            f"<b>{m.get('math_10_pct')}%</b> math (SRC-606-04). The attendance "
            f"rate was <b>{a.get('attend_rate_pct')}%</b> in 2024-25; "
            f"<b>{a.get('chronic_absent_10_pct')}%</b> of students were chronically "
            f"absent (SRC-606-05). The high-school dropout rate was "
            f"<b>{d.get('dropout_pct')}%</b>, <b>{commify(d.get('dropout_count') or 0)}</b> "
            f"students (SRC-606-06). Among districts, <b>{f.get('highest', {}).get('name')}</b> "
            f"had the highest total expenditures per pupil at "
            f"<b>${commify(f.get('highest', {}).get('v') or 0)}</b> (SRC-606-07). "
            f"Waitlists and lottery outcomes are not a published statewide table."
        )
        demo = sec.get("ma_enrollment_demographics_2026") or {}
        li = next((r for r in (demo.get("selected") or []) if r.get("name") == "Low income"), {})
        el = next((r for r in (demo.get("selected") or []) if r.get("name") == "English learners"), {})
        if li.get("v") is not None:
            parts.append(
                f"In 2025-26, <b>{li.get('v')}%</b> of Massachusetts public-school "
                f"students were low income and <b>{el.get('v')}%</b> were English "
                f"learners (SRC-606-08)."
            )
        k = next((r for r in (demo.get("grades") or []) if r.get("name") == "Kindergarten"), {})
        if k.get("v") is not None:
            parts.append(
                f"Kindergarten enrolled <b>{commify(k.get('v'))}</b> students "
                f"(SRC-606-08)."
            )
    if tool_id == "DL-07":
        n = (sec.get("naep_2024") or {}).get("series") or {}
        r4, m4 = n.get("read4") or {}, n.get("math4") or {}
        r8, m8 = n.get("read8") or {}, n.get("math8") or {}
        parts.append(
            f"On the 2024 NAEP, grade-4 reading averaged <b>{r4.get('us')}</b> "
            f"in the national public sample; Massachusetts scored "
            f"<b>{(r4.get('ma') or {}).get('v')}</b>, rank "
            f"{(r4.get('ma') or {}).get('rank')} of {(r4.get('ma') or {}).get('n')} "
            f"(SRC-607-05). Grade-4 math was <b>{m4.get('us')}</b> nationally and "
            f"<b>{(m4.get('ma') or {}).get('v')}</b> in Massachusetts "
            f"(SRC-607-05). Grade-8 reading was <b>{r8.get('us')}</b> / "
            f"<b>{(r8.get('ma') or {}).get('v')}</b>; grade-8 math was "
            f"<b>{m8.get('us')}</b> / <b>{(m8.get('ma') or {}).get('v')}</b> "
            f"(SRC-607-05)."
        )
        hist = ((sec.get("naep_2024") or {}).get("history") or {})
        r4c = (hist.get("read4") or {}).get("change_2019_2024") or {}
        m8c = (hist.get("math8") or {}).get("change_2019_2024") or {}
        r4_hi = r4c.get("highest") or {}
        if r4c.get("us") is not None:
            parts.append(
                f"From 2019 to 2024, national public grade-4 reading changed "
                f"<b>{r4c['us']:+.1f}</b> points; "
                f"<b>{r4c.get('n_up')}</b> of {r4c.get('n_ranked')} states rose "
                f"and <b>{r4c.get('n_down')}</b> fell (SRC-607-05). "
                f"<b>{r4_hi.get('name')}</b> had the largest grade-4 reading "
                f"gain at <b>{r4_hi.get('v'):+.1f}</b> (SRC-607-05). "
                f"Grade-8 math changed <b>{m8c.get('us'):+.1f}</b> nationally; "
                f"<b>{m8c.get('n_up')}</b> states rose (SRC-607-05)."
            )
        ex = sec.get("expulsion_2020_21") or {}
        if (ex.get("ma") or {}).get("v") is not None:
            parts.append(
                f"Expulsions reached <b>{ex.get('us')}%</b> of U.S. public-school "
                f"students in 2020-21; Massachusetts was <b>{(ex.get('ma') or {}).get('v')}%</b> "
                f"(SRC-607-07)."
            )
        race = ((sec.get("discipline_race_2020_21") or {}).get("oss") or {})
        us_b = next((r for r in (race.get("us") or []) if r.get("name") == "Black"), {})
        ma_b = next((r for r in (race.get("ma") or []) if r.get("name") == "Black"), {})
        if us_b.get("v") is not None and ma_b.get("v") is not None:
            parts.append(
                f"Among Black students, the out-of-school suspension share was "
                f"<b>{us_b.get('v')}%</b> nationally and <b>{ma_b.get('v')}%</b> "
                f"in Massachusetts (SRC-607-04)."
            )
    if tool_id == "DL-08":
        g = sec.get("ipeds_6yr_grad_by_state_2017") or {}
        fac = sec.get("public_fte_faculty_fall_2023") or {}
        tui = sec.get("he_public4_tuition_2022_23") or {}
        if g.get("us") is not None:
            parts.append(
                f"The IPEDS 6-year bachelor's graduation rate for the 2017 entry "
                f"cohort was <b>{g.get('us')}%</b> nationally; Massachusetts was "
                f"<b>{(g.get('ma') or {}).get('v')}%</b>, rank "
                f"{(g.get('ma') or {}).get('rank')} of {(g.get('ma') or {}).get('n')} "
                f"(derived, SRC-608-12)."
            )
        if (fac.get("ma") or {}).get("v") is not None:
            parts.append(
                f"Public institutions employed <b>{commify((fac.get('ma') or {}).get('v') or 0)}</b> "
                f"FTE faculty in Massachusetts in Fall 2023, rank "
                f"{(fac.get('ma') or {}).get('rank')} of {(fac.get('ma') or {}).get('n')} "
                f"(derived, SRC-608-06)."
            )
        if (tui.get("ma") or {}).get("v") is not None:
            parts.append(
                f"Public 4-year in-state tuition and fees were "
                f"<b>${commify((tui.get('ma') or {}).get('v') or 0)}</b> in "
                f"Massachusetts in 2022-23 (SRC-608-09)."
            )
    if tool_id == "DL-12":
        m = sec.get("mfcu_recoveries_fy2025") or {}
        parts.append(
            f"Medicaid Fraud Control Units recovered "
            f"<b>{usd_prose(m.get('us') or 0)}</b> in fiscal year 2025 "
            f"(SRC-612-03). Massachusetts recovered "
            f"<b>{usd_prose((m.get('ma') or {}).get('v') or 0)}</b>, rank "
            f"{(m.get('ma') or {}).get('rank')} of {(m.get('ma') or {}).get('n')} "
            f"(derived, SRC-612-03). NASBO health-chapter tables remain PDF-only."
        )
    if tool_id == "DL-13":
        b = sec.get("bed_births_deaths") or {}
        ma = b.get("ma") or {}
        ov = b.get("overlap") or {}
        parts.append(
            f"BLS Business Employment Dynamics put the Massachusetts "
            f"private-sector establishment birth rate at "
            f"<b>{ma.get('birth_rate_pct')}%</b> in {ma.get('births_as_of')} "
            f"({commify(ma.get('births') or 0)} establishments, SRC-613-02). "
            f"Deaths are published through {ma.get('deaths_as_of')}, when the "
            f"death rate was <b>{ma.get('death_rate_pct')}%</b> "
            f"({commify(ma.get('deaths') or 0)} establishments, SRC-613-02). "
            f"In the last overlapping quarter, {ov.get('q')}, the birth rate "
            f"was <b>{ov.get('ma_birth_rate_pct')}%</b> and the death rate was "
            f"<b>{ov.get('ma_death_rate_pct')}%</b> (SRC-613-02)."
        )
        fl = b.get("fl") or {}
        if fl.get("birth_rate_pct") is not None:
            parts.append(
                f"Florida's establishment birth rate was "
                f"<b>{fl.get('birth_rate_pct')}%</b> in {fl.get('births_as_of')} "
                f"({commify(fl.get('births') or 0)} establishments, SRC-613-02)."
            )
        w9 = (b.get("window_9q_2024q3") or {})
        if (w9.get("ma") or {}).get("rank"):
            parts.append(
                f"Over the 9 quarters ending 2024 Q3, Massachusetts had the "
                f"{'lowest' if w9['ma']['rank'] == w9['ma']['n'] else 'rank ' + str(w9['ma']['rank'])} "
                f"mean establishment birth rate among {w9['ma']['n']} jurisdictions "
                f"at <b>{w9['ma']['v']}%</b> (derived, SRC-613-02)."
            )
    if tool_id == "DL-14":
        u = sec.get("ui_initial_claims") or {}
        parts.append(
            f"UI initial claims were <b>{commify(u.get('us') or 0)}</b> in the "
            f"United States in the week ending {u.get('as_of_label')} "
            f"(derived, SRC-614-03). Massachusetts filed "
            f"<b>{commify((u.get('ma') or {}).get('v') or 0)}</b> initial claims "
            f"and <b>{commify(u.get('ma_continued') or 0)}</b> continued weeks "
            f"(SRC-614-03)."
        )
        qe = sec.get("qcew_employment_2025q4") or {}
        if (qe.get("ma") or {}).get("v") is not None:
            parts.append(
                f"QCEW average monthly employment was "
                f"<b>{commify((qe.get('ma') or {}).get('v') or 0)}</b> in "
                f"Massachusetts in 2025 Q4, rank {(qe.get('ma') or {}).get('rank')} of "
                f"{(qe.get('ma') or {}).get('n')} (derived, SRC-614-02)."
            )
    if tool_id == "DL-15":
        s = ((sec.get("sagdp2_naics_2025") or {}).get("industries") or {})
        a = s.get("all_industry") or {}
        mfg = s.get("manufacturing") or {}
        fin = s.get("finance_insurance") or {}
        parts.append(
            f"Current-dollar GDP was <b>{usd_prose((a.get('us') or 0) * 1_000_000)}</b> "
            f"in 2025 (SRC-615-03). Massachusetts was "
            f"<b>{usd_prose(((a.get('ma') or {}).get('v') or 0) * 1_000_000)}</b> "
            f"(SRC-615-03). Manufacturing was "
            f"<b>{usd_prose(((mfg.get('ma') or {}).get('v') or 0) * 1_000_000)}</b> "
            f"and finance and insurance "
            f"<b>{usd_prose(((fin.get('ma') or {}).get('v') or 0) * 1_000_000)}</b> "
            f"(SRC-615-03)."
        )
    if tool_id == "DL-16":
        c = sec.get("case_shiller_boston") or {}
        parts.append(
            f"The S&P/CoreLogic Case-Shiller Boston index was "
            f"<b>{c.get('boston')}</b> in {c.get('as_of_label')} "
            f"({pct(c.get('yoy_pct'))} from a year earlier) (SRC-616-03). "
            f"Boston is the only Massachusetts city in that series. "
            f"The Miami MSA index was <b>{c.get('miami')}</b> in "
            f"{c.get('miami_as_of_label')} "
            f"({pct(c.get('miami_yoy_pct'))} from a year earlier) (SRC-616-03)."
        )
    if tool_id == "DL-17":
        r = sec.get("rucc_2023") or {}
        parts.append(
            f"Under the 2023 USDA rural-urban continuum codes, "
            f"<b>{(r.get('ma') or {}).get('v')}%</b> of Massachusetts 2020 "
            f"county population lived in metro (RUCC 1-3) counties "
            f"(SRC-617-02). The state has <b>{r.get('ma_metro_counties')}</b> "
            f"metro counties and <b>{r.get('ma_nonmetro_counties')}</b> "
            f"nonmetro counties (SRC-617-02)."
        )
        a65 = sec.get("pop_age_65plus_share_2025") or {}
        if (a65.get("ma") or {}).get("v") is not None:
            parts.append(
                f"<b>{(a65.get('ma') or {}).get('v')}%</b> of Massachusetts "
                f"residents were 65 and over in 2025, rank "
                f"{(a65.get('ma') or {}).get('rank')} of {(a65.get('ma') or {}).get('n')} "
                f"(derived, SRC-617-03)."
            )
        b = sec.get("births_2025") or {}
        d = sec.get("deaths_2025") or {}
        if (b.get("ma") or {}).get("v") is not None:
            parts.append(
                f"Census estimated <b>{commify((b.get('ma') or {}).get('v') or 0)}</b> "
                f"births and <b>{commify((d.get('ma') or {}).get('v') or 0)}</b> deaths "
                f"in Massachusetts in 2025 (SRC-617-01)."
            )
    if tool_id == "DL-21":
        c = sec.get("ma_county_agi_2022") or {}
        parts.append(
            f"Among Massachusetts counties, <b>{c.get('highest', {}).get('name')}</b> "
            f"had the most adjusted gross income in tax year 2022 at "
            f"<b>{usd_prose(c.get('highest', {}).get('v') or 0)}</b> "
            f"(derived, SRC-621-02). Among Florida counties, "
            f"<b>{(sec.get('fl_county_agi_2022') or {}).get('highest', {}).get('name')}</b> "
            f"had the most AGI at "
            f"<b>{usd_prose((sec.get('fl_county_agi_2022') or {}).get('highest', {}).get('v') or 0)}</b> "
            f"(derived, SRC-621-02). A dedicated AGI-percentile-by-state file "
            f"is not posted; size-of-AGI stubs are stored under derived.secondary."
        )
    if tool_id == "DL-22":
        n = sec.get("ntd_annual_2024") or {}
        mb = n.get("mbta") or {}
        parts.append(
            f"In NTD report year 2024, U.S. agencies reported "
            f"<b>{usd_prose(n.get('us_operating') or 0)}</b> in operating "
            f"expenses, with a <b>{n.get('us_farebox_pct')}%</b> farebox "
            f"recovery rate (derived, SRC-622-02). The MBTA spent "
            f"<b>{usd_prose(mb.get('operating') or 0)}</b>, recovered "
            f"<b>{mb.get('farebox_pct')}%</b> from fares, and cost "
            f"<b>${mb.get('cost_per_trip')}</b> per unlinked trip "
            f"(derived, SRC-622-02)."
        )
    if tool_id == "DL-23":
        f = sec.get("fema_pa_obligations") or {}
        n = sec.get("nri_mean_county_score") or {}
        d = sec.get("noaa_degree_days_2024") or {}
        bits = [
            f"OpenFEMA Public Assistance records show "
            f"<b>{usd_prose((f.get('ma') or {}).get('v') or 0)}</b> in federal "
            f"share obligated to Massachusetts (SRC-623-03). The National Risk "
            f"Index mean county score is <b>{(n.get('ma') or {}).get('v')}</b>, "
            f"rank {(n.get('ma') or {}).get('rank')} of {(n.get('ma') or {}).get('n')} "
            f"(derived, SRC-623-04)."
        ]
        if d:
            bits.append(
                f"Massachusetts recorded <b>{commify(d.get('ma_hdd') or 0)}</b> "
                f"heating degree days and <b>{commify(d.get('ma_cdd') or 0)}</b> "
                f"cooling degree days in 2024 (SRC-623-05)."
            )
        parts.append(" ".join(bits))
    if tool_id == "DL-24":
        p = sec.get("seds_production_2024") or {}
        parts.append(
            f"SEDS total energy production was <b>{commify(p.get('us') or 0)}</b> "
            f"billion Btu in the United States in 2024 (SRC-624-03). Massachusetts "
            f"produced <b>{commify((p.get('ma') or {}).get('v') or 0)}</b> billion "
            f"Btu, rank {(p.get('ma') or {}).get('rank')} of {(p.get('ma') or {}).get('n')} "
            f"(derived, SRC-624-03)."
        )
    if tool_id == "DL-25":
        p = sec.get("population_peers_2025") or {}
        bos = ((p.get("peers") or {}).get("Boston city") or [{}])[0]
        parts.append(
            f"Population peers are the five Census 2025 municipalities closest "
            f"in resident count (SRC-625-02). Boston's nearest peer is "
            f"<b>{bos.get('name')}</b> at <b>{commify(bos.get('pop') or 0)}</b> "
            f"(derived, SRC-625-02). This is not the old Pioneer socioeconomic "
            f"peer workbook. A DLS levy file is not posted as a stable public CSV."
        )
    if tool_id == "DL-26":
        f = sec.get("district_ppe_fy2025") or {}
        parts.append(
            f"On DESE FY 2025 total expenditures per pupil, "
            f"<b>{f.get('highest', {}).get('name')}</b> was highest at "
            f"<b>${commify(f.get('highest', {}).get('v') or 0)}</b> among "
            f"{f.get('districts')} districts (SRC-626-02). DLS debt, levy, "
            f"revenue, and municipal crime files are not posted as stable "
            f"public CSVs on this pass."
        )
    if tool_id == "DL-30":
        q = sec.get("quasi_payroll_2025") or {}
        v = sec.get("vendor_extract_fy2025") or {}
        parts.append(
            f"Quasi-public payroll totaled <b>{usd_prose(q.get('total') or 0)}</b> "
            f"in calendar 2025 across <b>{commify(q.get('employees') or 0)}</b> "
            f"rows (SRC-630-03). <b>{q.get('highest', {}).get('name')}</b> was "
            f"the largest quasi-public employer (SRC-630-03). The largest "
            f"non-payroll vendor in FY 2025 was "
            f"<b>{v.get('highest', {}).get('name')}</b> at "
            f"<b>{usd_prose(v.get('highest', {}).get('v') or 0)}</b> "
            f"(SRC-630-04)."
        )
    if tool_id == "DL-27":
        e = (sec.get("boston_top_earners_2025") or {}).get("highest") or {}
        if e.get("name"):
            parts.append(
                f"<b>{e.get('name')}</b> was the highest named City of Boston "
                f"earner in 2025 at <b>{usd_prose(e.get('v') or 0)}</b> "
                f"(SRC-627-01)."
            )
    if tool_id == "DL-29":
        r = sec.get("aslg_revenue_2022") or {}
        p = sec.get("aspp_holdings_2025") or {}
        g = sec.get("gov_units_2022") or {}
        if (r.get("ma") or {}).get("v") is not None:
            parts.append(
                f"State and local revenue was "
                f"<b>{usd_prose((r.get('ma') or {}).get('v') or 0)}</b> in "
                f"Massachusetts in 2022, rank {(r.get('ma') or {}).get('rank')} of "
                f"{(r.get('ma') or {}).get('n')} (derived, SRC-629-05)."
            )
        if (p.get("ma") or {}).get("v") is not None:
            parts.append(
                f"Public pension systems held "
                f"<b>{usd_prose((p.get('ma') or {}).get('v') or 0)}</b> in cash "
                f"and investments (derived, SRC-629-07)."
            )
        if (g.get("ma") or {}).get("v") is not None:
            parts.append(
                f"The 2022 Census of Governments counted "
                f"<b>{commify((g.get('ma') or {}).get('v') or 0)}</b> government "
                f"units in Massachusetts (SRC-629-06)."
            )
    extra = hollow_lead(tool_id, sec)
    if extra:
        parts.append(extra)
    return " ".join(parts)


MORE_STRIP = {
    "DL-06": [
        "District MCAS files are pending on this page.",
        "MCAS, attendance, and district finance remain pending.",
        "MCAS remains pending.",
    ],
    "DL-07": [
        "NAEP state scores remain pending.",
        "NAEP scores and discipline files are pending on this page.",
        "NAEP, completion, and discipline remain pending.",
    ],
    "DL-08": [
        "Admissions-test and faculty files are pending.",
        "Admissions tests, faculty, and IPEDS outcomes remain pending.",
        "State faculty counts are not in that national table.",
        "A state-level faculty table is not in the current Digest xlsx set.",
        "That Digest table is national; it has no state column.",
        "That Digest table is national and has no state column.",
    ],
    "DL-10": [
        "CHIA relative prices remain pending on this page.",
        "CHIA relative prices remain pending.",
        "CHIA relative-price files remain pending.",
    ],
    "DL-12": [
        "Broader state health spending and fraud recoveries are pending on this page.",
        "NASBO health-chapter and fraud-recovery files remain pending.",
        "Large states lead on raw dollars; fraud recoveries are pending.",
    ],
    "DL-14": [
        "UI claims remain pending.",
        "A CPS demographic extract is not in this ledger.",
        "CPS labor demographics are not posted as a machine file: decline those.",
    ],
    "DL-19": [
        "Tariff, defense, and fiscal-dependency measures are pending.",
        "Tariff, defense-impact, and fiscal-dependency measures are Pioneer products and stay pending until those methods are in the repo.",
    ],
    "DL-15": [
        "NAICS industry detail remains pending.",
        "Personal income and NAICS detail are pending on this page.",
    ],
    "DL-16": [
        "Case-Shiller city indexes remain a later view.",
        "Case-Shiller city indexes are pending on this page.",
    ],
    "DL-17": [
        "The rural-urban continuum is listed as a later view.",
        "USDA rural-urban codes are pending on this page.",
    ],
    "DL-21": [
        "County and municipal extracts are pending.",
        "County and municipal files remain pending.",
        "County, municipal, and AGI-percentile tables are listed as later views.",
        "County and municipal files are pending.",
    ],
    "DL-22": [
        "Operating cost and farebox recovery remain later views.",
        "Agency operating cost and farebox recovery are pending.",
        "Agency operating cost is pending here.",
    ],
    "DL-23": [
        "FEMA risk and degree-day files are pending.",
        "FEMA obligations and the National Risk Index are pending on this page.",
        "FEMA and NOAA files remain pending.",
        "FEMA obligations, the National Risk Index, and degree days are listed as later views.",
    ],
    "DL-24": [
        "SEDS production remains pending because that file was not in the consumption extract.",
        "SEDS production remains pending.",
        "SEDS production remains a later view.",
        "SEDS production is pending.",
    ],
    "DL-25": [
        "Tax levy and peer sets are pending on this page.",
        "DLS levy and peer-set files remain pending.",
        "Tax levy and peer sets are listed as later views.",
        "Tax levy and peer sets are pending.",
        "A DLS levy file is not posted as a stable public CSV.",
        "This is not the old Pioneer socioeconomic peer workbook.",
    ],
    "DL-26": [
        "Crime, debt, education, spending, and tax rankings are pending because those DLS and DESE files were not reachable this pass.",
        "Crime, debt, education, and levy rankings are pending.",
        "Crime, debt, education, expenditure, revenue, and tax rankings remain pending.",
        "Crime, debt, education, spending, and tax rankings are pending: decline those.",
        "DLS, DESE, and crime rankings are pending.",
    ],
    "DL-28": [
        "DOR monthly collections and tax-credit files are pending.",
        "DOR monthly reports and tax credits are pending.",
        "DOR monthly and tax-credit files remain pending.",
    ],
    "DL-29": [
        "Annual Survey of State Government Finances, NASBO rainy-day funds, and employee counts remain pending.",
        "NASBO rainy-day figures remain pending.",
        "Rainy-day funds and public-employee counts are pending.",
    ],
    "DL-31": [
        "FBI crime rates, juvenile incarceration, and IC3 are pending.",
        "FBI crime rates, juvenile incarceration, and internet-crime reports are pending on this page.",
        "FBI UCR/NIBRS, juvenile, and IC3 files remain pending.",
    ],
    "DL-30": [
        "Quasi-public payroll detail and a vendor-only extract remain later views.",
        "Quasi-public payroll detail is pending.",
        "Spending is the Comptroller all-object-class total, not a vendor-only extract.",
    ],
}
