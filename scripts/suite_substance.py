#!/usr/bin/env python3
"""Compile Tableau companions that public files already support.

Merges new derived.secondary keys into existing live ledgers. Does not
rebuild primary rankings, invent Pioneer-method indexes, or restore
archived topics. Every figure is a published cell or a labeled derived
rollup of published cells.
"""
from __future__ import annotations

import csv
import io
import json
import sys
import urllib.parse
import zipfile
from collections import defaultdict

from openpyxl import load_workbook

from suite_common import (
    FIPS_TO_ST,
    LEDGER_DIR,
    RANKED,
    STATE_NAMES,
    commify,
    fetch,
    fetch_text,
    fl_cell,
    geo_to_st,
    ledger_path,
    load_apps,
    parse_num,
    rank_named,
    rank_rows,
    snap_pack,
    usd_prose,
    write_ledger,
)
def _wb(url, timeout=120):
    return load_workbook(io.BytesIO(fetch(url, timeout=timeout)), data_only=True)

PAGE_REVISED = "Aug 16, 2026"

DIGEST_314 = "https://nces.ed.gov/programs/digest/d24/tables/xls/tabn314.50.xlsx"
DIGEST_333 = "https://nces.ed.gov/programs/digest/d22/tables/xls/tabn333.30.xlsx"
DIGEST_334 = "https://nces.ed.gov/programs/digest/d22/tables/xls/tabn334.20.xlsx"
DIGEST_330 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn330.20.xlsx"
DIGEST_213 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn213.20.xlsx"
DIGEST_319 = "https://nces.ed.gov/programs/digest/d22/tables/xls/tabn319.20.xlsx"
DIGEST_317 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn317.20.xlsx"
DIGEST_233 = "https://nces.ed.gov/programs/digest/d23/tables/xls/tabn233.40.xlsx"
IPEDS_HD = "https://nces.ed.gov/ipeds/datacenter/data/HD2023.zip"
IPEDS_GR = "https://nces.ed.gov/ipeds/datacenter/data/GR2023.zip"
CENSUS_AGE = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2020-2025/state/asrh/sc-est2025-alldata6.csv"
)
E2C_ENROLL = "https://educationtocareer.data.mass.gov/resource/t8td-gens.json"
URL_QCEW = "https://data.bls.gov/cew/data/api/2025/4/industry/10.csv"
URL_ASLG = "https://www2.census.gov/programs-surveys/gov-finances/tables/2022/22slsstab1.xlsx"
URL_GUS = "https://www2.census.gov/programs-surveys/gus/tables/2022/cog2022_cg2200org01.zip"
URL_ASPP = (
    "https://www2.census.gov/programs-surveys/aspp/datasets/2025/"
    "aspp-historical-datasets/ASPP_Unit_File_2025.csv"
)
URL_BOS_2025 = (
    "https://data.boston.gov/datastore/dump/29b3544f-752a-4cb1-a6af-a1de153d20a0"
)
BOS_YEAR_DUMPS = [
    (2015, "941c9de4-fb91-41bb-ad5a-43a35f5dc80f"),
    (2016, "2ff6343f-850d-46e7-98d1-aca79b619fd6"),
    (2017, "8368bd3d-3633-4927-8355-2a2f9811ab4f"),
    (2018, "70129b87-bd4e-49bb-aa09-77644da73503"),
    (2019, "31358fd1-849a-48e0-8285-e813f6efbdf1"),
    (2020, "3bdfe6dc-3a81-49ce-accc-22161e2f7e74"),
    (2021, "e2e2c23a-6fc7-4456-8751-5321d8aa869b"),
    (2022, "ec5aaf93-1509-4641-9310-28e62e028457"),
    (2023, "63ac638b-36c4-487d-9453-1d83eb5090d2"),
    (2024, "6b3c5333-1dcb-4b3d-9cd7-6a03fb526da7"),
    (2025, "29b3544f-752a-4cb1-a6af-a1de153d20a0"),
]

VERIFY_US_STAFF_314 = 2023852
VERIFY_US_FACULTY_314 = 694542
VERIFY_US_STU_FAC = 14.4
VERIFY_US_STATE_APPROP_2021 = 75516740.741  # thousands
VERIFY_US_HE_EXP_2021 = 419696081.648  # thousands
VERIFY_US_TUITION_2223 = 9749.987
VERIFY_US_K12_STAFF = 6795469.766418726
VERIFY_US_BA_2021 = 2066445
VERIFY_US_INSTITUTIONS = 3722
VERIFY_US_EXPULSION = 0.05938007341602887
VERIFY_US_IPEDS_6YR = 64.6
VERIFY_E2C_SY2026 = 900490
VERIFY_GUS_US = 90888
VERIFY_US_BIRTHS_2025 = 3620461
VERIFY_US_DEATHS_2025 = 3101603
VERIFY_US_INTL_MIG_2025 = 1262202
URL_PEP = (
    "https://www2.census.gov/programs-surveys/popest/datasets/"
    "2020-2025/state/totals/NST-EST2025-ALLDATA.csv"
)
GRADE_FIELDS = [
    ("pk_cnt", "Pre-kindergarten"),
    ("k_cnt", "Kindergarten"),
    ("g1_cnt", "Grade 1"),
    ("g2_cnt", "Grade 2"),
    ("g3_cnt", "Grade 3"),
    ("g4_cnt", "Grade 4"),
    ("g5_cnt", "Grade 5"),
    ("g6_cnt", "Grade 6"),
    ("g7_cnt", "Grade 7"),
    ("g8_cnt", "Grade 8"),
    ("g9_cnt", "Grade 9"),
    ("g10_cnt", "Grade 10"),
    ("g11_cnt", "Grade 11"),
    ("g12_cnt", "Grade 12"),
    ("sp_cnt", "Special education beyond grade 12"),
]
OSS_RACE_COLS = [
    (4, "American Indian/Alaska Native"),
    (5, "Asian"),
    (6, "Black"),
    (7, "Hispanic"),
    (8, "Pacific Islander"),
    (9, "White"),
    (10, "Two or more races"),
]
EXPEL_RACE_COLS = [
    (14, "American Indian/Alaska Native"),
    (15, "Asian"),
    (16, "Black"),
    (17, "Hispanic"),
    (18, "Pacific Islander"),
    (19, "White"),
    (20, "Two or more races"),
]


def _digest_state_pairs(ws, start_row, col, us_row=None, min_states=48):
    values, us_val = {}, None
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if i < start_row and i != us_row:
            continue
        st = geo_to_st(row[0])
        if not st:
            continue
        v = parse_num(row[col])
        if v is None:
            continue
        if st == "US":
            if us_val is None:
                us_val = v
        elif st not in values:
            values[st] = v
    if len(values) < min_states:
        sys.exit(f"FATAL: {ws.title} col {col} parsed {len(values)} states")
    return values, us_val


def _snap(values, us_val, round_to=None, higher_is_better=True):
    return snap_pack(values, us_val, round_to=round_to, higher_is_better=higher_is_better)


def sec_public_he_faculty():
    ws = _wb(DIGEST_314).active
    staff, us_staff = _digest_state_pairs(ws, 6, 1, us_row=5)
    fac, us_fac = _digest_state_pairs(ws, 6, 4, us_row=5)
    ratio, us_ratio = _digest_state_pairs(ws, 6, 12, us_row=5)
    if us_staff is None or abs(us_staff - VERIFY_US_STAFF_314) > 1:
        sys.exit(f"FATAL: Digest 314.50 US FTE staff is {us_staff}")
    if us_fac is None or abs(us_fac - VERIFY_US_FACULTY_314) > 1:
        sys.exit(f"FATAL: Digest 314.50 US FTE faculty is {us_fac}")
    if us_ratio is None or abs(us_ratio - VERIFY_US_STU_FAC) > 0.05:
        sys.exit(f"FATAL: Digest 314.50 US students per faculty is {us_ratio}")
    staff_s = _snap({k: round(v) for k, v in staff.items()}, round(us_staff))
    fac_s = _snap({k: round(v) for k, v in fac.items()}, round(us_fac))
    ratio_s = _snap({k: round(v, 1) for k, v in ratio.items()}, round(us_ratio, 1), higher_is_better=False)
    meta = {
        "src": "SRC-608-06",
        "as_of_label": "Fall 2023",
        "note": "NCES Digest table 314.50, public degree-granting institutions.",
    }
    staff_s.update({**meta, "label": "Public FTE staff, Fall 2023", "unit": "full-time-equivalent staff"})
    fac_s.update({**meta, "label": "Public FTE faculty, Fall 2023", "unit": "full-time-equivalent faculty"})
    ratio_s.update({
        **meta,
        "label": "FTE students per FTE faculty, public institutions, Fall 2023",
        "unit": "students per faculty",
        "note": "NCES Digest table 314.50. A lower ratio is fewer students per faculty member.",
    })
    return {
        "public_fte_staff_fall_2023": staff_s,
        "public_fte_faculty_fall_2023": fac_s,
        "students_per_faculty_fall_2023": ratio_s,
    }


def sec_he_finance():
    ws = _wb(DIGEST_333).active
    state_ap, us_state = _digest_state_pairs(ws, 7, 6, us_row=7)
    local_ap, us_local = _digest_state_pairs(ws, 7, 12, us_row=7)
    if us_state is None or abs(us_state - VERIFY_US_STATE_APPROP_2021) > 1:
        sys.exit(f"FATAL: Digest 333.30 US state appropriations 2020-21 are {us_state}")
    # Published cells are thousands of dollars.
    state_d = {k: v * 1000 for k, v in state_ap.items()}
    local_d = {k: v * 1000 for k, v in local_ap.items()}
    total_d = {k: state_d[k] + local_d.get(k, 0) for k in state_d}
    us_state_d = us_state * 1000
    us_local_d = (us_local or 0) * 1000
    us_tot = us_state_d + us_local_d
    state_s = _snap({k: round(v) for k, v in state_d.items()}, round(us_state_d))
    tot_s = _snap({k: round(v) for k, v in total_d.items()}, round(us_tot))
    state_s.update({
        "label": "State appropriations to public higher education, 2020-21",
        "src": "SRC-608-07",
        "unit": "dollars",
        "as_of_label": "Academic year 2020-21",
        "note": "NCES Digest table 333.30. Published in thousands of dollars; shown in dollars.",
    })
    tot_s.update({
        "label": "State and local appropriations to public higher education, 2020-21",
        "src": "SRC-608-07",
        "unit": "dollars",
        "as_of_label": "Academic year 2020-21",
        "note": "NCES Digest table 333.30. State plus local appropriations. Published in thousands of dollars; shown in dollars.",
    })

    ws = _wb(DIGEST_334).active
    exp, us_exp = _digest_state_pairs(ws, 8, 6, us_row=7)
    if us_exp is None or abs(us_exp - VERIFY_US_HE_EXP_2021) > 1:
        sys.exit(f"FATAL: Digest 334.20 US expenditures 2020-21 are {us_exp}")
    exp_d = {k: v * 1000 for k, v in exp.items()}
    exp_s = _snap({k: round(v) for k, v in exp_d.items()}, round(us_exp * 1000))
    exp_s.update({
        "label": "Public higher-education expenditures, 2020-21",
        "src": "SRC-608-08",
        "unit": "dollars",
        "as_of_label": "Academic year 2020-21",
        "note": "NCES Digest table 334.20, all public institutions. Published in thousands of dollars; shown in dollars.",
    })

    ws = _wb(DIGEST_330).active
    tui, us_tui = _digest_state_pairs(ws, 8, 4, us_row=7)
    if us_tui is None or abs(us_tui - VERIFY_US_TUITION_2223) > 1:
        sys.exit(f"FATAL: Digest 330.20 US in-state tuition 2022-23 is {us_tui}")
    tui_s = _snap({k: round(v) for k, v in tui.items()}, round(us_tui))
    tui_s.update({
        "label": "Public 4-year in-state tuition and required fees, 2022-23",
        "src": "SRC-608-09",
        "unit": "dollars",
        "as_of_label": "Academic year 2022-23",
        "note": "NCES Digest table 330.20. In-state tuition and required fees at public 4-year institutions.",
    })
    return {
        "he_state_appropriations_2020_21": state_s,
        "he_state_local_appropriations_2020_21": tot_s,
        "he_expenditures_2020_21": exp_s,
        "he_public4_tuition_2022_23": tui_s,
    }


def sec_he_students():
    ws = _wb(DIGEST_319).active
    ba, us_ba = _digest_state_pairs(ws, 6, 10, us_row=5)
    if us_ba is None or abs(us_ba - VERIFY_US_BA_2021) > 1:
        sys.exit(f"FATAL: Digest 319.20 US bachelor's 2020-21 is {us_ba}")
    ba_s = _snap({k: round(v) for k, v in ba.items()}, round(us_ba))
    ba_s.update({
        "label": "Bachelor's degrees conferred, 2020-21",
        "src": "SRC-608-10",
        "unit": "degrees",
        "as_of_label": "Academic year 2020-21",
        "note": "NCES Digest table 319.20.",
    })

    ws = _wb(DIGEST_317).active
    inst, us_inst = _digest_state_pairs(ws, 6, 1, us_row=5)
    if us_inst is None or abs(us_inst - VERIFY_US_INSTITUTIONS) > 0.5:
        sys.exit(f"FATAL: Digest 317.20 US institutions are {us_inst}")
    inst_s = _snap({k: int(v) for k, v in inst.items()}, int(us_inst))
    inst_s.update({
        "label": "Degree-granting postsecondary institutions, 2022-23",
        "src": "SRC-608-11",
        "unit": "institutions",
        "as_of_label": "Academic year 2022-23",
        "note": "NCES Digest table 317.20.",
    })
    return {
        "bachelors_conferred_2020_21": ba_s,
        "degree_granting_institutions_2022_23": inst_s,
    }


def sec_ipeds_grad_by_state():
    hd = zipfile.ZipFile(io.BytesIO(fetch(IPEDS_HD, timeout=180)))
    gr = zipfile.ZipFile(io.BytesIO(fetch(IPEDS_GR, timeout=180)))
    hd_name = next(n for n in hd.namelist() if n.lower().endswith(".csv"))
    gr_name = next(n for n in gr.namelist() if n.lower().endswith(".csv"))

    def _rows(raw):
        text = raw.decode("utf-8-sig", "replace")
        rdr = csv.DictReader(io.StringIO(text))
        rdr.fieldnames = [((h or "").lstrip("\ufeff").strip()) for h in (rdr.fieldnames or [])]
        return rdr

    stabbr = {}
    for r in _rows(hd.read(hd_name)):
        uid = (r.get("UNITID") or "").strip()
        st = (r.get("STABBR") or "").strip()
        deg = str(r.get("DEGGRANT") or "").strip()
        if uid and st in STATE_NAMES and deg == "1":
            stabbr[uid] = st
    cohort = defaultdict(float)
    completers = defaultdict(float)
    for r in _rows(gr.read(gr_name)):
        uid = (r.get("UNITID") or "").strip()
        st = stabbr.get(uid)
        if not st:
            continue
        gtype = str(r.get("GRTYPE") or "").strip()
        n = parse_num(r.get("GRTOTLT"))
        if n is None:
            continue
        if gtype == "8":
            cohort[st] += n
        elif gtype == "12":
            completers[st] += n
    rates = {}
    for st in RANKED:
        if cohort.get(st) and completers.get(st) is not None:
            rates[st] = 100.0 * completers[st] / cohort[st]
    if "MA" not in rates or len(rates) < 48:
        sys.exit(f"FATAL: IPEDS GR2023 parsed {len(rates)} state bachelor's rates")
    us_coh = sum(cohort.values())
    us_comp = sum(completers.values())
    us_rate = 100.0 * us_comp / us_coh if us_coh else None
    if us_rate is None or abs(us_rate - VERIFY_US_IPEDS_6YR) > 0.5:
        sys.exit(f"FATAL: IPEDS bachelor's 150% rate is {us_rate}, expected ~{VERIFY_US_IPEDS_6YR}")
    snap = _snap({k: round(v, 1) for k, v in rates.items()}, round(us_rate, 1))
    snap.update({
        "label": "IPEDS 6-year bachelor's graduation rate, 2017 cohort",
        "src": "SRC-608-12",
        "unit": "percent",
        "as_of_label": "2017 cohort, completing by 2023",
        "note": (
            "IPEDS GR2023 joined to HD2023 on UNITID. Bachelor's or equivalent "
            "subcohort at degree-granting 4-year institutions: completers within "
            "150% of normal time (GRTYPE 12) divided by the adjusted cohort "
            "(GRTYPE 8). The national rollup is "
            f"{round(us_rate, 1)} percent, matching Digest 326.10."
        ),
    })
    return {"ipeds_6yr_grad_by_state_2017": snap}


def sec_k12_staff():
    ws = _wb(DIGEST_213).active
    total, us_tot = _digest_state_pairs(ws, 8, 1, us_row=7)
    aides, us_aides = _digest_state_pairs(ws, 8, 8, us_row=7)
    if us_tot is None or abs(us_tot - VERIFY_US_K12_STAFF) > 1:
        sys.exit(f"FATAL: Digest 213.20 US staff is {us_tot}")
    tot_s = _snap({k: round(v) for k, v in total.items()}, round(us_tot))
    aide_s = _snap({k: round(v) for k, v in aides.items()}, round(us_aides) if us_aides else None)
    tot_s.update({
        "label": "Public elementary and secondary staff (FTE), Fall 2022",
        "src": "SRC-609-03",
        "unit": "full-time-equivalent staff",
        "as_of_label": "Fall 2022",
        "note": "NCES Digest table 213.20. Teachers remain on the companion teacher table.",
    })
    aide_s.update({
        "label": "Public-school instructional aides (FTE), Fall 2022",
        "src": "SRC-609-03",
        "unit": "full-time-equivalent aides",
        "as_of_label": "Fall 2022",
        "note": "NCES Digest table 213.20.",
    })
    return {
        "k12_staff_fte_fall_2022": tot_s,
        "k12_aides_fte_fall_2022": aide_s,
    }


def sec_expulsion():
    ws = _wb(DIGEST_233).active
    values, us_val = _digest_state_pairs(ws, 7, 11, us_row=6, min_states=40)
    if us_val is None or abs(us_val - VERIFY_US_EXPULSION) > 0.001:
        sys.exit(f"FATAL: Digest 233.40 US expulsion share is {us_val}")
    snap = _snap({k: round(v, 2) for k, v in values.items()}, round(us_val, 2), higher_is_better=False)
    snap.update({
        "label": "Share of public-school students expelled, 2020-21",
        "src": "SRC-607-07",
        "unit": "percent",
        "as_of_label": "School year 2020-21",
        "note": "NCES Digest table 233.40. Higher is more students expelled, not a performance ranking.",
    })
    return {"expulsion_2020_21": snap}


def sec_ma_demographics():
    q = urllib.parse.urlencode({
        "$where": "org_type='State'",
        "$order": "sy DESC",
        "$limit": "6",
    })
    raw = fetch(E2C_ENROLL + "?" + q, timeout=90)
    rows = json.loads(raw.decode("utf-8"))
    latest = next(
        (
            r for r in rows
            if str(r.get("sy")) == "2026"
            and str(r.get("org_code") or "") == "00000000"
        ),
        None,
    )
    if not latest:
        sys.exit("FATAL: E2C t8td-gens has no state row")
    total = parse_num(latest.get("total_cnt"))
    if total is None or abs(total - VERIFY_E2C_SY2026) > 1:
        sys.exit(f"FATAL: E2C SY2026 state enrollment is {total}")

    def pct_field(key):
        v = parse_num(latest.get(key))
        if v is None:
            return None
        return round(v * 100, 1) if v <= 1.5 else round(v, 1)

    race = [
        {"name": "White", "v": pct_field("wh_pct")},
        {"name": "Hispanic or Latino", "v": pct_field("hl_pct")},
        {"name": "Black or African American", "v": pct_field("baa_pct")},
        {"name": "Asian", "v": pct_field("as_pct")},
        {"name": "Multi-race, non-Hispanic", "v": pct_field("mnhl_pct")},
        {"name": "American Indian or Alaska Native", "v": pct_field("aian_pct")},
        {"name": "Native Hawaiian or Pacific Islander", "v": pct_field("nhpi_pct")},
    ]
    selected = [
        {"name": "High needs", "v": pct_field("hn_pct"), "count": parse_num(latest.get("hn_cnt"))},
        {"name": "Low income", "v": pct_field("li_pct"), "count": parse_num(latest.get("li_cnt"))},
        {"name": "First language not English", "v": pct_field("flne_pct"), "count": parse_num(latest.get("flne_cnt"))},
        {"name": "English learners", "v": pct_field("el_pct"), "count": parse_num(latest.get("el_cnt"))},
        {"name": "Students with disabilities", "v": pct_field("swd_pct"), "count": parse_num(latest.get("swd_cnt"))},
    ]
    grades = []
    for key, name in GRADE_FIELDS:
        v = parse_num(latest.get(key))
        if v is None:
            continue
        grades.append({"name": name, "v": int(v)})
    grade_sum = sum(g["v"] for g in grades)
    if grades and abs(grade_sum - total) > 5:
        sys.exit(f"FATAL: E2C grade counts sum to {grade_sum} vs total {total}")
    return {
        "ma_enrollment_demographics_2026": {
            "label": "Massachusetts public-school enrollment by race and selected populations, 2025-26",
            "src": "SRC-606-08",
            "unit": "percent",
            "as_of_label": "School year 2025-26",
            "total": int(total),
            "race": [r for r in race if r["v"] is not None],
            "selected": [r for r in selected if r["v"] is not None],
            "grades": grades,
            "note": "DESE / E2C Enrollment: Grade, Race/Ethnicity, Gender, and Selected Populations. Statewide All Students.",
        }
    }


def sec_qcew_employment():
    text = fetch_text(URL_QCEW, timeout=120)
    emp = {}
    for r in csv.DictReader(io.StringIO(text)):
        if not (
            r.get("own_code") == "0"
            and r.get("agglvl_code") == "50"
            and r.get("industry_code") == "10"
        ):
            continue
        fips = (r.get("area_fips") or "").zfill(5)
        st = FIPS_TO_ST.get(fips[:2])
        if not st or st == "US":
            continue
        e1 = parse_num(r.get("month1_emplvl"))
        e2 = parse_num(r.get("month2_emplvl"))
        e3 = parse_num(r.get("month3_emplvl"))
        if None in (e1, e2, e3):
            continue
        emp[st] = (e1 + e2 + e3) / 3
    if "MA" not in emp or len(emp) < 50:
        sys.exit(f"FATAL: QCEW employment parsed {len(emp)} states")
    us_emp = sum(emp.values())
    snap = _snap({k: round(v) for k, v in emp.items()}, round(us_emp))
    snap.update({
        "label": "Average monthly employment, all industries, 2025 Q4",
        "src": "SRC-614-02",
        "unit": "jobs",
        "as_of_label": "2025 Q4",
        "note": "BLS QCEW statewide all-ownership, all industries. The U.S. total is the sum of published state cells (derived, SRC-614-02).",
    })
    return {"qcew_employment_2025q4": snap}


def sec_pop_age_race():
    text = fetch_text(CENSUS_AGE, timeout=180)
    age = defaultdict(lambda: {"0_17": 0.0, "18_64": 0.0, "65plus": 0.0, "total": 0.0})
    race = defaultdict(lambda: {
        "white_nh": 0.0, "black": 0.0, "asian": 0.0, "hispanic": 0.0, "total": 0.0,
    })
    for r in csv.DictReader(io.StringIO(text)):
        if str(r.get("SUMLEV") or "").zfill(3) != "040":
            continue
        fips = str(r.get("STATE") or "").zfill(2)
        st = FIPS_TO_ST.get(fips)
        if not st or st == "US":
            continue
        sex = str(r.get("SEX") or "").strip()
        origin = str(r.get("ORIGIN") or "").strip()
        race_c = str(r.get("RACE") or "").strip()
        age_n = parse_num(r.get("AGE"))
        pop = parse_num(r.get("POPESTIMATE2025"))
        if pop is None or age_n is None:
            continue
        if sex == "0" and origin == "0":
            age[st]["total"] += pop
            if age_n <= 17:
                age[st]["0_17"] += pop
            elif age_n <= 64:
                age[st]["18_64"] += pop
            else:
                age[st]["65plus"] += pop
            if race_c == "2":
                race[st]["black"] += pop
            elif race_c == "4":
                race[st]["asian"] += pop
        if sex == "0" and origin == "1" and race_c == "1":
            race[st]["white_nh"] += pop
        if sex == "0" and origin == "2":
            race[st]["hispanic"] += pop
        if sex == "0" and origin == "0":
            race[st]["total"] += pop
    if "MA" not in age or age["MA"]["total"] < 6_000_000:
        sys.exit(f"FATAL: Census SC-EST2025 MA total is {age.get('MA')}")
    share65 = {st: 100.0 * rec["65plus"] / rec["total"] for st, rec in age.items() if rec["total"]}
    share17 = {st: 100.0 * rec["0_17"] / rec["total"] for st, rec in age.items() if rec["total"]}
    hisp = {st: 100.0 * rec["hispanic"] / rec["total"] for st, rec in race.items() if rec["total"]}
    white = {st: 100.0 * rec["white_nh"] / rec["total"] for st, rec in race.items() if rec["total"]}
    s65 = _snap({k: round(v, 1) for k, v in share65.items()}, None)
    s17 = _snap({k: round(v, 1) for k, v in share17.items()}, None)
    sh = _snap({k: round(v, 1) for k, v in hisp.items()}, None)
    sw = _snap({k: round(v, 1) for k, v in white.items()}, None)
    note = "Census vintage 2025 state characteristics (SC-EST2025-ALLDATA6). Shares are derived from published age-sex-race cells."
    s65.update({"label": "Share of population age 65 and over, 2025", "src": "SRC-617-03", "unit": "percent", "as_of_label": "July 1, 2025", "note": note})
    s17.update({"label": "Share of population age 17 and under, 2025", "src": "SRC-617-03", "unit": "percent", "as_of_label": "July 1, 2025", "note": note})
    sh.update({"label": "Hispanic or Latino share of population, 2025", "src": "SRC-617-03", "unit": "percent", "as_of_label": "July 1, 2025", "note": note})
    sw.update({"label": "White non-Hispanic share of population, 2025", "src": "SRC-617-03", "unit": "percent", "as_of_label": "July 1, 2025", "note": note})
    ma_age = age["MA"]
    return {
        "pop_age_65plus_share_2025": s65,
        "pop_age_0_17_share_2025": s17,
        "pop_hispanic_share_2025": sh,
        "pop_white_nh_share_2025": sw,
        "pop_age_race_ma_2025": {
            "label": "Massachusetts population by age, 2025",
            "src": "SRC-617-03",
            "as_of_label": "July 1, 2025",
            "total": round(ma_age["total"]),
            "age_0_17": round(ma_age["0_17"]),
            "age_18_64": round(ma_age["18_64"]),
            "age_65plus": round(ma_age["65plus"]),
            "hispanic_pct": sh["ma"]["v"],
            "white_nh_pct": sw["ma"]["v"],
            "note": note,
        },
    }


def sec_pop_components():
    text = fetch_text(URL_PEP, timeout=120)
    births, deaths, natural, intl, change = {}, {}, {}, {}, {}
    us = {}
    for r in csv.DictReader(io.StringIO(text)):
        name = (r.get("NAME") or "").strip()
        b = parse_num(r.get("BIRTHS2025"))
        d = parse_num(r.get("DEATHS2025"))
        nchg = parse_num(r.get("NATURALCHG2025"))
        im = parse_num(r.get("INTERNATIONALMIG2025"))
        pch = parse_num(r.get("NPOPCHG_2025"))
        if None in (b, d, im):
            continue
        if name == "United States":
            us = {
                "births": int(b),
                "deaths": int(d),
                "natural": int(nchg) if nchg is not None else int(b - d),
                "intl": int(im),
                "change": int(pch) if pch is not None else None,
            }
            continue
        if str(r.get("SUMLEV") or "") != "040":
            continue
        st = next((k for k, v in STATE_NAMES.items() if v == name), None)
        if not st:
            continue
        births[st] = int(b)
        deaths[st] = int(d)
        natural[st] = int(nchg) if nchg is not None else int(b - d)
        intl[st] = int(im)
        if pch is not None:
            change[st] = int(pch)
    if us.get("births") != VERIFY_US_BIRTHS_2025:
        sys.exit(f"FATAL: NST US births 2025 are {us.get('births')}")
    if us.get("deaths") != VERIFY_US_DEATHS_2025:
        sys.exit(f"FATAL: NST US deaths 2025 are {us.get('deaths')}")
    if us.get("intl") != VERIFY_US_INTL_MIG_2025:
        sys.exit(f"FATAL: NST US international migration 2025 is {us.get('intl')}")
    if "MA" not in births:
        sys.exit("FATAL: NST 2025 has no Massachusetts births")
    note = (
        "Census vintage 2025 NST-EST2025-ALLDATA. Births, deaths, natural "
        "change, international migration, and NPOPCHG_2025 are published cells."
    )
    b_s = _snap(births, us["births"])
    d_s = _snap(deaths, us["deaths"])
    n_s = _snap(natural, us["natural"])
    i_s = _snap(intl, us["intl"])
    b_s.update({"label": "Births, 2025", "src": "SRC-617-01", "unit": "people", "as_of_label": "July 1, 2025", "note": note})
    d_s.update({"label": "Deaths, 2025", "src": "SRC-617-01", "unit": "people", "as_of_label": "July 1, 2025", "note": note})
    n_s.update({"label": "Natural change, 2025", "src": "SRC-617-01", "unit": "people", "as_of_label": "July 1, 2025", "note": note})
    i_s.update({"label": "International migration, 2025", "src": "SRC-617-01", "unit": "people", "as_of_label": "July 1, 2025", "note": note})
    out = {
        "births_2025": b_s,
        "deaths_2025": d_s,
        "natural_change_2025": n_s,
        "international_mig_2025": i_s,
    }
    if us.get("change") is not None and len(change) >= 50:
        c_s = _snap(change, us["change"])
        c_s.update({"label": "Population change, 2025", "src": "SRC-617-01", "unit": "people", "as_of_label": "July 1, 2025", "note": note})
        out["pop_change_2025"] = c_s
    return out


def _digest_row_by_geo(ws, start_row):
    found = {}
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        st = geo_to_st(row[0])
        if st:
            found[st] = row
    return found


def sec_discipline_race():
    ws = _wb(DIGEST_233).active
    rows = _digest_row_by_geo(ws, 6)
    us_row, ma_row = rows.get("US"), rows.get("MA")
    if not us_row or not ma_row:
        sys.exit("FATAL: Digest 233.40 missing US or MA for race columns")

    def pack(row, cols):
        out = []
        for col, name in cols:
            if col >= len(row):
                continue
            v = parse_num(row[col])
            if v is None:
                continue
            out.append({"name": name, "v": round(v, 2)})
        return out

    oss_us, oss_ma = pack(us_row, OSS_RACE_COLS), pack(ma_row, OSS_RACE_COLS)
    exp_us, exp_ma = pack(us_row, EXPEL_RACE_COLS), pack(ma_row, EXPEL_RACE_COLS)
    if len(oss_us) < 5 or len(oss_ma) < 5:
        sys.exit(f"FATAL: Digest 233.40 OSS race parsed us={len(oss_us)} ma={len(oss_ma)}")
    return {
        "discipline_race_2020_21": {
            "label": "Out-of-school suspension and expulsion share by race, 2020-21",
            "src": "SRC-607-04",
            "unit": "percent",
            "as_of_label": "School year 2020-21",
            "oss": {"us": oss_us, "ma": oss_ma},
            "expulsion": {"us": exp_us, "ma": exp_ma},
            "note": (
                "NCES Digest table 233.40. Shares are of students in that "
                "racial or ethnic group. In-school suspension is not a column "
                "on this table."
            ),
        }
    }


def sec_aslg_2022():
    wb = load_workbook(io.BytesIO(fetch(URL_ASLG, timeout=120)), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[9]]
    geo_cols = []
    for i, h in enumerate(headers):
        if not h:
            continue
        st = geo_to_st(h)
        if st:
            geo_cols.append((st, i))
    if not geo_cols or geo_cols[0][0] != "US":
        # First named geo should be United States; remaining are states.
        pass
    rev_row = exp_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_col=2, values_only=True), 1):
        lab = str(row[1] or "").replace("\xa0", " ").strip()
        lab = "".join(ch for ch in lab if not ch.isdigit()).strip().lower()
        if lab == "revenue" and rev_row is None:
            rev_row = i
        if lab == "expenditure" and exp_row is None:
            exp_row = i
    if rev_row is None or exp_row is None:
        sys.exit(f"FATAL: ASLG 2022 missing revenue/expenditure rows ({rev_row}, {exp_row})")
    rev_vals, exp_vals, us_rev, us_exp = {}, {}, None, None
    rev_line = [c.value for c in ws[rev_row]]
    exp_line = [c.value for c in ws[exp_row]]
    for st, col in geo_cols:
        rv = parse_num(rev_line[col])
        ev = parse_num(exp_line[col])
        if rv is None or ev is None:
            continue
        # Published in thousands of dollars.
        if st == "US":
            us_rev, us_exp = rv * 1000, ev * 1000
        else:
            rev_vals[st] = rv * 1000
            exp_vals[st] = ev * 1000
    if us_rev is None or us_rev < 3e12 or "MA" not in rev_vals:
        sys.exit(f"FATAL: ASLG 2022 US revenue is {us_rev} states={len(rev_vals)}")
    rev_s = _snap({k: round(v) for k, v in rev_vals.items()}, round(us_rev))
    exp_s = _snap({k: round(v) for k, v in exp_vals.items()}, round(us_exp))
    note = "Census of Governments Finance 2022 table 1, state-and-local amount. Published in thousands of dollars; shown in dollars."
    rev_s.update({"label": "State and local general-plus-utility revenue, 2022", "src": "SRC-629-05", "unit": "dollars", "as_of_label": "Fiscal year 2022", "note": note})
    exp_s.update({"label": "State and local expenditure, 2022", "src": "SRC-629-05", "unit": "dollars", "as_of_label": "Fiscal year 2022", "note": note})
    return {
        "aslg_revenue_2022": rev_s,
        "aslg_expenditure_2022": exp_s,
    }


def sec_gov_orgs():
    zf = zipfile.ZipFile(io.BytesIO(fetch(URL_GUS, timeout=90)))
    name = next(n for n in zf.namelist() if n.endswith("Data.xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(name)), data_only=True)
    ws = wb.active
    values, us_val = {}, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[2]) != "2022" or str(row[4]) != "GO0001":
            continue
        st = FIPS_TO_ST.get(str(row[9] or "").zfill(2))
        v = parse_num(row[3])
        if v is None:
            continue
        if st == "US" or str(row[9]) == "00":
            us_val = v
        elif st:
            values[st] = v
    if us_val is None or abs(us_val - VERIFY_GUS_US) > 0.5:
        sys.exit(f"FATAL: GUS 2022 US government units are {us_val}")
    snap = _snap({k: int(v) for k, v in values.items()}, int(us_val))
    snap.update({
        "label": "Federal, state, and local government units, 2022",
        "src": "SRC-629-06",
        "unit": "governments",
        "as_of_label": "2022 Census of Governments",
        "note": "Census of Governments organization table CG2200ORG01, AGG_DESC GO0001.",
    })
    return {"gov_units_2022": snap}


def sec_aspp_2025():
    text = fetch_text(URL_ASPP, timeout=180)
    holdings = defaultdict(float)
    members = defaultdict(float)
    for r in csv.DictReader(io.StringIO(text)):
        st = (r.get("STATE") or "").strip()
        if st not in STATE_NAMES or st == "US":
            continue
        code = (r.get("ITEM_CODE") or "").strip()
        val = parse_num(r.get("ITEM_VALUE"))
        w = parse_num(r.get("FINAL_WEIGHT")) or 1
        if val is None:
            continue
        if code == "RZ01":
            holdings[st] += val * w
        elif code == "RJ01":
            members[st] += val * w
    if "MA" not in holdings or len(holdings) < 48:
        sys.exit(f"FATAL: ASPP 2025 holdings parsed {len(holdings)} states")
    us_h = sum(holdings.values())
    us_m = sum(members.values())
    if us_h < 1e12:
        sys.exit(f"FATAL: ASPP 2025 US holdings are {us_h}")
    h_s = _snap({k: round(v) for k, v in holdings.items()}, round(us_h))
    m_s = _snap({k: round(v) for k, v in members.items()}, round(us_m))
    note = (
        "Census Annual Survey of Public Pensions 2025 unit file. "
        "RZ01 is total cash and investments; RJ01 is total membership. "
        "2025 is a sample year; published FINAL_WEIGHT is applied (derived, SRC-629-07)."
    )
    h_s.update({"label": "Public pension cash and investments, 2025", "src": "SRC-629-07", "unit": "dollars", "as_of_label": "2025", "note": note})
    m_s.update({"label": "Public pension membership, 2025", "src": "SRC-629-07", "unit": "members", "as_of_label": "2025", "note": note})
    return {
        "aspp_holdings_2025": h_s,
        "aspp_membership_2025": m_s,
    }


def _bos_gross_col(fieldnames):
    for key in fieldnames or []:
        k = key.strip().upper().replace("_", " ")
        if k in ("TOTAL GROSS", "TOTAL EARNINGS", "TOTAL_GROSS", "TOTAL EARNINGS "):
            return key
        if "TOTAL" in k and ("GROSS" in k or "EARN" in k):
            return key
    return None


def sec_boston_earners():
    text = fetch_text(URL_BOS_2025, timeout=180)
    rdr = csv.DictReader(io.StringIO(text))
    people = []
    for r in rdr:
        name = (r.get("NAME") or "").strip()
        dept = (r.get("DEPARTMENT_NAME") or "").strip()
        title = (r.get("TITLE") or "").strip()
        gross = parse_num(r.get("TOTAL GROSS"))
        if not name or gross is None:
            continue
        people.append({"name": name, "department": dept, "title": title, "v": gross})
    people.sort(key=lambda p: p["v"], reverse=True)
    if len(people) < 1000 or people[0]["v"] < 200000:
        sys.exit(f"FATAL: Boston 2025 named earnings parsed n={len(people)}")
    top = people[:12]
    trend = []
    for year, rid in BOS_YEAR_DUMPS:
        url = f"https://data.boston.gov/datastore/dump/{rid}"
        try:
            body = fetch_text(url, timeout=180)
        except Exception:
            continue
        rows = csv.DictReader(io.StringIO(body))
        col = _bos_gross_col(rows.fieldnames)
        if not col:
            continue
        total = 0.0
        n = 0
        for r in rows:
            v = parse_num(r.get(col))
            if v is None:
                continue
            total += v
            n += 1
        if n > 500 and total > 1e8:
            trend.append({"y": year, "v": round(total), "employees": n})
    out = {
        "boston_top_earners_2025": {
            "label": "Highest City of Boston earnings, calendar 2025",
            "src": "SRC-627-01",
            "unit": "dollars",
            "as_of_label": "Calendar year 2025",
            "top": top,
            "highest": {"name": top[0]["name"], "v": top[0]["v"], "department": top[0]["department"], "title": top[0]["title"]},
            "note": "City of Boston employee earnings report 2025. TOTAL GROSS, named employees.",
        }
    }
    if len(trend) >= 3:
        out["boston_payroll_trend"] = {
            "label": "City of Boston total earnings, 2015 to 2025",
            "src": "SRC-627-01",
            "unit": "dollars",
            "trend": trend,
            "note": "Yearly CKAN dumps of the employee earnings report. TOTAL GROSS (or the year's total-earnings column) summed across named rows.",
        }
    return out


SUBSTANCE = {
    "DL-06": sec_ma_demographics,
    "DL-07": lambda: {**sec_expulsion(), **sec_discipline_race()},
    "DL-08": lambda: {
        **sec_public_he_faculty(),
        **sec_he_finance(),
        **sec_he_students(),
        **sec_ipeds_grad_by_state(),
    },
    "DL-09": sec_k12_staff,
    "DL-14": sec_qcew_employment,
    "DL-17": lambda: {**sec_pop_age_race(), **sec_pop_components()},
    "DL-27": sec_boston_earners,
    "DL-29": lambda: {
        **sec_aslg_2022(),
        **sec_gov_orgs(),
        **sec_aspp_2025(),
    },
}

SOURCES = {
    "DL-06": [{
        "id": "SRC-606-08",
        "name": "DESE / E2C enrollment by grade, race/ethnicity, gender, and selected populations",
        "cadence": "Annual school year",
        "url": "https://educationtocareer.data.mass.gov/resource/t8td-gens.json",
    }],
    "DL-07": [{
        "id": "SRC-607-07",
        "name": "NCES Digest table 233.40, expulsion share",
        "cadence": "Periodic",
        "url": DIGEST_233,
    }],
    "DL-08": [
        {"id": "SRC-608-06", "name": "NCES Digest table 314.50, public FTE staff and faculty by state", "cadence": "Annual", "url": DIGEST_314},
        {"id": "SRC-608-07", "name": "NCES Digest table 333.30, state and local appropriations for public higher education", "cadence": "Annual", "url": DIGEST_333},
        {"id": "SRC-608-08", "name": "NCES Digest table 334.20, public higher-education expenditures by state", "cadence": "Annual", "url": DIGEST_334},
        {"id": "SRC-608-09", "name": "NCES Digest table 330.20, public 4-year in-state tuition and fees", "cadence": "Annual", "url": DIGEST_330},
        {"id": "SRC-608-10", "name": "NCES Digest table 319.20, degrees conferred by state", "cadence": "Annual", "url": DIGEST_319},
        {"id": "SRC-608-11", "name": "NCES Digest table 317.20, degree-granting institutions by state", "cadence": "Annual", "url": DIGEST_317},
        {"id": "SRC-608-12", "name": "IPEDS GR2023 and HD2023, 6-year bachelor's graduation rate by state", "cadence": "Annual", "url": IPEDS_GR},
    ],
    "DL-09": [{
        "id": "SRC-609-03",
        "name": "NCES Digest table 213.20, public-school staff by assignment",
        "cadence": "Annual",
        "url": DIGEST_213,
    }],
    "DL-17": [{
        "id": "SRC-617-03",
        "name": "Census vintage 2025 state population by age, sex, and race",
        "cadence": "Annual",
        "url": CENSUS_AGE,
    }],
    "DL-29": [
        {"id": "SRC-629-05", "name": "Census of Governments Finance 2022, state and local revenue and expenditure", "cadence": "Every five years / annual estimate", "url": URL_ASLG},
        {"id": "SRC-629-06", "name": "Census of Governments 2022 organization counts (CG2200ORG01)", "cadence": "Every five years", "url": URL_GUS},
        {"id": "SRC-629-07", "name": "Census Annual Survey of Public Pensions 2025 unit file", "cadence": "Annual", "url": URL_ASPP},
    ],
}


def attach_sources(ledger, app, extra):
    smap = ledger.setdefault("source_id_map", {})
    for src in extra:
        smap[src["id"]] = {
            "name": src["name"],
            "cadence": src["cadence"],
            "url": src["url"],
            "supports": app.get("scope") or "",
        }


def merge_tool(app):
    tid = app["id"]
    path = ledger_path(tid)
    ledger = json.loads(path.read_text(encoding="utf-8"))
    if ledger.get("status") != "live":
        return ledger, []
    print(f"  compile {tid} ...", flush=True)
    added = SUBSTANCE[tid]()
    sec = ledger.setdefault("derived", {}).setdefault("secondary", {})
    sec.update(added)
    for rec in sec.values():
        if isinstance(rec, dict):
            rec.pop("rows", None)
    bits = []
    demo = added.get("ma_enrollment_demographics_2026") or {}
    k = next((r for r in (demo.get("grades") or []) if r.get("name") == "Kindergarten"), {})
    if k.get("v") is not None:
        bits.append(
            f"Kindergarten enrolled <b>{commify(k['v'])}</b> students (SRC-606-08)."
        )
    race = (added.get("discipline_race_2020_21") or {}).get("oss") or {}
    us_b = next((r for r in (race.get("us") or []) if r.get("name") == "Black"), {})
    ma_b = next((r for r in (race.get("ma") or []) if r.get("name") == "Black"), {})
    if us_b.get("v") is not None and ma_b.get("v") is not None:
        bits.append(
            f"Among Black students, the out-of-school suspension share was "
            f"<b>{us_b.get('v')}%</b> nationally and <b>{ma_b.get('v')}%</b> "
            f"in Massachusetts (SRC-607-04)."
        )
    b = added.get("births_2025") or {}
    d = added.get("deaths_2025") or {}
    if (b.get("ma") or {}).get("v") is not None:
        bits.append(
            f"Census estimated <b>{commify((b.get('ma') or {}).get('v') or 0)}</b> "
            f"births and <b>{commify((d.get('ma') or {}).get('v') or 0)}</b> deaths "
            f"in Massachusetts in 2025 (SRC-617-01)."
        )
    lead = ledger.get("lead") or ""
    for bit in bits:
        if bit[:48] not in lead:
            lead = (lead + " " + bit).strip()
    ledger["lead"] = lead
    attach_sources(ledger, app, SOURCES.get(tid) or [])
    for key in ("q", "scope", "exclusions"):
        if app.get(key):
            ledger[key] = app[key]
    ledger.setdefault("page", {})["revised"] = PAGE_REVISED
    note = ledger.get("vintage_note") or ""
    extra = f" Tableau companions compiled {PAGE_REVISED} are stored under derived.secondary."
    if extra.strip() not in note:
        ledger["vintage_note"] = (note + extra).strip()
    write_ledger(ledger)
    print(f"    added {', '.join(added)}", flush=True)
    return ledger, list(added)


def main():
    apps = {a["id"]: a for a in load_apps()}
    wanted = sys.argv[1:] or list(SUBSTANCE)
    for tid in wanted:
        if tid not in SUBSTANCE:
            sys.exit(f"FATAL: no substance compiler for {tid}")
        merge_tool(apps[tid])


if __name__ == "__main__":
    main()
